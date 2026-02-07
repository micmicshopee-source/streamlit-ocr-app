from __future__ import annotations

import streamlit as st
try:
    from streamlit.errors import StreamlitSecretNotFoundError
except ImportError:
    try:
        from streamlit.runtime.secrets import StreamlitSecretNotFoundError
    except ImportError:
        StreamlitSecretNotFoundError = None
import google.generativeai as genai
from PIL import Image, ImageEnhance
import pandas as pd
import json
import os
import io
import time
import base64
import requests
import altair as alt
import sqlite3
from datetime import datetime, timedelta
import hashlib
import shutil
import secrets as _secrets_module
import re
from openpyxl.styles import Alignment, Font

# 密碼雜湊：優先使用 bcrypt（AUTH-01），無則退回 SHA256
try:
    import bcrypt
    _USE_BCRYPT = True
except ImportError:
    _USE_BCRYPT = False
    bcrypt = None

# PDF 生成庫（僅 fpdf2）
try:
    from fpdf import FPDF
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# 若無 .streamlit/secrets.toml 則建立空檔，避免 Streamlit 報 No secrets found
def _ensure_secrets_file():
    try:
        app_dir = os.path.dirname(os.path.abspath(__file__))
        streamlit_dir = os.path.join(app_dir, ".streamlit")
        secrets_path = os.path.join(streamlit_dir, "secrets.toml")
        if not os.path.isfile(secrets_path):
            os.makedirs(streamlit_dir, exist_ok=True)
            with open(secrets_path, "w", encoding="utf-8") as f:
                f.write("# Optional: GEMINI_API_KEY, USERS, etc.\n")
    except Exception:
        pass
_ensure_secrets_file()

# --- 1. 系統佈局與初始化 ---
st.set_page_config(page_title="上班族小工具 | 發票報帳・辦公小幫手", page_icon="🧾", layout="wide")

# --- 主題：Premium Dark（Google Black #0F0F0F / 卡片 #1E1E1E / 4px·8px 網格 / 導航 Hover 過渡）---
def _inject_premium_dark_css():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    css_path = os.path.join(base_dir, "premium_dark.css")
    if os.path.isfile(css_path):
        with open(css_path, "r", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)
_inject_premium_dark_css()

if "db_error" not in st.session_state: st.session_state.db_error = None
if "db_path_mode" not in st.session_state: st.session_state.db_path_mode = "💾 本地磁碟"
if "use_memory_mode" not in st.session_state: st.session_state.use_memory_mode = False
if "local_invoices" not in st.session_state: st.session_state.local_invoices = []
if "local_batches" not in st.session_state: st.session_state.local_batches = []
if "image_storage_dir" not in st.session_state: 
    base_dir = os.path.dirname(os.path.abspath(__file__))
    st.session_state.image_storage_dir = os.path.join(base_dir, "invoice_images")
    os.makedirs(st.session_state.image_storage_dir, exist_ok=True)
if "last_edited_df_hash" not in st.session_state: st.session_state.last_edited_df_hash = None
# 登入狀態管理（多用戶版本）
# 注意：Streamlit 的 session_state 在頁面刷新時應保持（同一瀏覽器會話）
if "authenticated" not in st.session_state: 
    st.session_state.authenticated = False
if "user_email" not in st.session_state: 
    st.session_state.user_email = None
# AUTH-04：Session 過期（登入時間，用於逾時檢查）
if "login_at" not in st.session_state:
    st.session_state.login_at = None
# 左側小工具導航：目前選中的工具（invoice=發票報帳，其餘為規劃中）
if "current_tool" not in st.session_state: 
    st.session_state.current_tool = "invoice"
# 刪除確認狀態（修復 Bug #2）
if "show_delete_confirm" not in st.session_state: st.session_state.show_delete_confirm = False
# 公司資訊（用於 PDF 導出）
if "company_name" not in st.session_state: st.session_state.company_name = ""
if "company_ubn" not in st.session_state: st.session_state.company_ubn = ""
# 主列表 + 詳情抽屜：選中發票 index、列表分頁；詳情彈出框用 id
if "detail_invoice_index" not in st.session_state: st.session_state.detail_invoice_index = None
if "invoice_master_page" not in st.session_state: st.session_state.invoice_master_page = 0
if "detail_invoice_id" not in st.session_state: st.session_state.detail_invoice_id = None

# --- 安全讀取 Streamlit Secrets（無 secrets.toml 時不報錯）---
def _load_secrets_from_app_dir():
    """從 app.py 所在目錄的 .streamlit/secrets.toml 讀取，供 Streamlit 未載入時備援。"""
    cache = getattr(_load_secrets_from_app_dir, "_cache", None)
    if cache is not None:
        return cache
    out = {}
    try:
        app_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(app_dir, ".streamlit", "secrets.toml")
        if not os.path.isfile(path):
            _load_secrets_from_app_dir._cache = out
            return out
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                v = v.strip()
                if (v.startswith('"') and v.endswith('"')) or (v.startswith("'") and v.endswith("'")):
                    v = v[1:-1].replace("\\n", "\n").replace("\\t", "\t")
                out[k] = v
    except Exception:
        pass
    _load_secrets_from_app_dir._cache = out
    return out

def _safe_secrets_get(key, default=None):
    """若無 .streamlit/secrets.toml 或缺少 key，返回 default，不拋錯。先試 st.secrets，再試應用目錄檔案。"""
    try:
        if key in st.secrets:
            return st.secrets[key]
    except Exception as e:
        if StreamlitSecretNotFoundError is not None and isinstance(e, StreamlitSecretNotFoundError):
            pass
        elif type(e).__name__ == "StreamlitSecretNotFoundError":
            pass
        else:
            raise
    # 備援：從 app 目錄的 .streamlit/secrets.toml 直接讀取（避免因工作目錄不同而讀不到）
    fallback = _load_secrets_from_app_dir()
    return fallback.get(key, default)

# --- 1.4. 密碼雜湊與強度（AUTH-01, AUTH-02）---
# bcrypt 雜湊前綴，用於辨識新格式；舊為純 64 字元 hex（SHA256）
_LEGACY_HASH_PREFIX = "sha256:"

def hash_password(password: str) -> str:
    """產生密碼雜湊：優先 bcrypt，否則 SHA256（相容舊資料）。"""
    if _USE_BCRYPT and bcrypt:
        return bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    return _LEGACY_HASH_PREFIX + hashlib.sha256(password.encode("utf-8")).hexdigest()

def verify_password(password: str, stored_hash: str) -> bool:
    """驗證密碼：支援 bcrypt、帶前綴的 SHA256、與舊版純 64 字元 hex（無前綴）。"""
    if not stored_hash:
        return False
    legacy_hex = hashlib.sha256(password.encode("utf-8")).hexdigest()
    if stored_hash.startswith(_LEGACY_HASH_PREFIX):
        return stored_hash == _LEGACY_HASH_PREFIX + legacy_hex
    # 舊版：純 64 字元 hex（無前綴）
    if len(stored_hash) == 64 and all(c in "0123456789abcdef" for c in stored_hash.lower()):
        return stored_hash.lower() == legacy_hex
    if _USE_BCRYPT and bcrypt:
        try:
            return bcrypt.checkpw(password.encode("utf-8"), stored_hash.encode("utf-8"))
        except Exception:
            return False
    return False

def validate_password_strength(password: str) -> tuple[bool, str]:
    """
    AUTH-02：密碼強度檢核。至少 8 字元，且含大小寫、數字、符號其中至少兩類。
    回傳 (ok, message)。
    """
    if len(password) < 8:
        return False, "密碼至少需要 8 個字元"
    has_upper = any(c.isupper() for c in password)
    has_lower = any(c.islower() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = bool(re.search(r'[!@#$%^&*(),.?":{}|<>_\-+=[\]\\;/\'`~]', password))
    categories = sum([has_upper, has_lower, has_digit, has_special])
    if categories < 2:
        return False, "密碼須包含以下其中至少兩類：大寫、小寫、數字、符號"
    return True, ""

# --- 1.5. 註冊函數 ---
def register_user(email: str, password: str):
    """註冊新用戶（寫入 SQLite users 表）；密碼須通過強度檢核。"""
    email = email.strip()
    if not email or not password:
        return False, "電子郵件與密碼不可為空"
    
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        return False, "電子郵件格式不正確"
    
    ok, msg = validate_password_strength(password)
    if not ok:
        return False, msg
    
    # 初始化資料庫（確保 users 表存在）
    init_db()
    path = get_db_path()
    is_uri = path.startswith("file:") and "mode=memory" in path
    
    try:
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        
        # 檢查是否已存在
        cursor.execute("SELECT id FROM users WHERE email = ?", (email,))
        if cursor.fetchone():
            conn.close()
            return False, "此電子郵件已註冊，請直接登入"
        
        # 寫入新用戶
        cursor.execute(
            "INSERT INTO users (email, password_hash) VALUES (?, ?)",
            (email, hash_password(password)),
        )
        conn.commit()
        conn.close()
        return True, "註冊成功，已為您自動登入"
    except Exception as e:
        return False, f"註冊失敗: {str(e)}"

# --- 1.6. 用戶驗證函數（多用戶版本：優先查資料庫）---
def verify_user(email, password):
    """
    驗證用戶登錄
    優先順序：
    1. 查詢 SQLite users 表（註冊用戶）
    2. Streamlit Secrets
    3. 環境變數
    """
    import re
    
    # 驗證電子郵件格式
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(email_pattern, email):
        return False, "電子郵件格式不正確"
    
    email = email.strip()
    
    # AUTH-03：登入失敗鎖定檢查
    if is_login_locked(email):
        return False, "帳號暫時鎖定，請 15 分鐘後再試"
    
    # ① 優先查詢 SQLite users 表（註冊用戶）
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        
        cursor.execute(
            "SELECT id, password_hash FROM users WHERE email = ?", (email,)
        )
        row = cursor.fetchone()
        if row:
            user_id, stored_hash = row
            if not stored_hash:
                conn.close()
                return False, "此帳號僅支援第三方登入，請使用 Google / LINE / Facebook 登入"
            
            if verify_password(password, stored_hash):
                _clear_login_attempts(email)
                cursor.execute(
                    "UPDATE users SET last_login = ? WHERE id = ?",
                    (datetime.now().isoformat(), user_id),
                )
                conn.commit()
                conn.close()
                return True, "登入成功"
            else:
                _record_login_attempt(email, False)
                conn.close()
                return False, "電子郵件或密碼錯誤"
        conn.close()
    except Exception as e:
        # 資料庫查詢失敗，繼續使用其他方式
        pass
    
    # ② 使用 Streamlit Secrets（若無 secrets.toml 或無 USERS 則跳過，不報錯）
    users = _safe_secrets_get("USERS")
    if users is not None:
        if isinstance(users, dict):
            # 格式：{"user@example.com": "password", ...}
            if email in users:
                if users[email] == password or users[email] == "":
                    return True, "登入成功"
        elif isinstance(users, str):
            # 格式：字串，每行一個 "email:password"
            for line in users.strip().split('\n'):
                if ':' in line:
                    user_email, user_password = line.split(':', 1)
                    if user_email.strip() == email:
                        if user_password.strip() == password or user_password.strip() == "":
                            return True, "登入成功"
    
    # 其次使用環境變數
    env_users = os.getenv("USERS")
    if env_users:
        for line in env_users.strip().split('\n'):
            if ':' in line:
                user_email, user_password = line.split(':', 1)
                if user_email.strip() == email:
                    if user_password.strip() == password or user_password.strip() == "":
                        return True, "登入成功"
    
    # 生產環境：不提供默認測試帳號，必須通過註冊或 Secrets 配置
    return False, "電子郵件或密碼錯誤"


def user_exists_in_db(email):
    """檢查該郵箱是否已在本系統（SQLite users 表）註冊。僅限資料庫註冊用戶可重設密碼。"""
    if not email or not email.strip():
        return False
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM users WHERE email = ?", (email.strip(),))
        exists = cursor.fetchone() is not None
        conn.close()
        return exists
    except Exception:
        return False


def update_user_password(email, new_password):
    """重設本系統註冊用戶的密碼（僅限 SQLite users 表）；密碼須通過強度檢核。回傳 (success, message)。"""
    email = email.strip()
    if not email or not new_password:
        return False, "電子郵件與新密碼不可為空"
    ok, msg = validate_password_strength(new_password)
    if not ok:
        return False, msg
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute(
            "UPDATE users SET password_hash = ? WHERE email = ?",
            (hash_password(new_password), email),
        )
        conn.commit()
        updated = cursor.rowcount > 0
        conn.close()
        if updated:
            return True, "密碼已更新，請使用新密碼登入"
        return False, "找不到該電子郵件的註冊帳號，請先註冊或確認電子郵件是否正確"
    except Exception as e:
        return False, f"更新失敗: {str(e)}"


# --- AUTH-03：登入失敗鎖定 ---
def _record_login_attempt(account_key: str, success: bool):
    """記錄一次登入嘗試（account_key 為 email 或 IP）。"""
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO login_attempts (account_key, success) VALUES (?, ?)",
            (account_key, 1 if success else 0),
        )
        conn.commit()
        conn.close()
    except Exception:
        pass

def _clear_login_attempts(account_key: str):
    """登入成功後清除該帳號的嘗試記錄。"""
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM login_attempts WHERE account_key = ?", (account_key,))
        conn.commit()
        conn.close()
    except Exception:
        pass

def is_login_locked(account_key: str, max_attempts: int = 5, lock_minutes: int = 15) -> bool:
    """檢查該帳號是否在鎖定期內（最近 max_attempts 次皆失敗且最後一次在 lock_minutes 分鐘內）。"""
    try:
        init_db()
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cutoff = (datetime.now() - timedelta(minutes=lock_minutes)).isoformat()
        cursor.execute(
            """SELECT COUNT(*) FROM login_attempts
               WHERE account_key = ? AND success = 0 AND attempt_at >= ?""",
            (account_key, cutoff),
        )
        count = cursor.fetchone()[0]
        conn.close()
        return count >= max_attempts
    except Exception:
        return False


# --- 1.7. 登入頁面（多用戶版本：含註冊功能）---
def login_page():
    """顯示登入頁面（台灣版：含註冊、忘記密碼與 Google 登入規劃）"""
    # AUTH-05：CSRF 用 token，顯示表單時產生
    if "login_csrf_token" not in st.session_state:
        st.session_state.login_csrf_token = _secrets_module.token_hex(16)
    col1, col2, col3 = st.columns([1, 2.5, 1])
    with col2:
        st.markdown('<div class="login-container">', unsafe_allow_html=True)
        st.title("🔐 上班族小工具")
        st.markdown('<p>登入以使用發票報帳與更多辦公小幫手</p>', unsafe_allow_html=True)
        st.caption("您的資料僅供您本人使用，我們不會分享給第三方。")
        
        # 第三方登入：若使用者已點選某一個，顯示對應授權連結
        oauth_pending = st.session_state.get("oauth_pending")
        if oauth_pending:
            url, err = None, None
            label = ""
            if oauth_pending == "google":
                url, err = build_oauth_url_google()
                label = "Google"
            elif oauth_pending == "line":
                url, err = build_oauth_url_line()
                label = "LINE"
            elif oauth_pending == "facebook":
                url, err = build_oauth_url_facebook()
                label = "Facebook"
            if err or not url:
                st.warning(err or "無法取得登入連結")
                if st.button("← 返回", key="oauth_back"):
                    st.session_state.pop("oauth_pending", None)
                    st.rerun()
            else:
                st.info(f"請點擊以下連結以 **{label}** 帳號登入：")
                st.markdown(f'<a href="{url}" target="_self" class="login-oauth-link">以 {label} 登入</a>', unsafe_allow_html=True)
                st.caption("開啟後將導向授權頁，完成後會回到本頁。")
                if st.button("← 返回登入頁", key="oauth_cancel"):
                    st.session_state.pop("oauth_pending", None)
                    st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
            return
        
        # 選擇登入或註冊模式
        if "login_mode" not in st.session_state:
            st.session_state.login_mode = "登入"
        if "show_forgot_password" not in st.session_state:
            st.session_state.show_forgot_password = False
        
        mode = st.radio(
            "選擇操作", 
            ["登入", "註冊"], 
            horizontal=True,
            key="mode_selector",
            index=0 if st.session_state.login_mode == "登入" else 1
        )
        st.session_state.login_mode = mode
        if mode != "登入":
            st.session_state.show_forgot_password = False
        
        st.markdown("---")
        
        if mode == "登入":
            if st.session_state.show_forgot_password:
                st.subheader("🔑 重設密碼")
                st.caption("僅限在本站註冊的帳號可重設密碼（Secrets 設定的帳號請聯繫管理員）")
                reset_email = st.text_input("📧 註冊時使用的電子郵件", key="reset_email", 
                                           placeholder="you@example.com", label_visibility="visible")
                new_pw = st.text_input("🔒 新密碼（至少 8 字元，含大小寫/數字/符號其中兩類）", type="password", key="reset_new_pw", 
                                      label_visibility="visible")
                new_pw_confirm = st.text_input("🔒 再輸入一次新密碼", type="password", key="reset_confirm", 
                                               label_visibility="visible")
                r1, r2 = st.columns(2)
                with r1:
                    if st.button("✅ 重設密碼", type="primary", use_container_width=True, key="btn_reset_pw"):
                        if not reset_email:
                            st.error("❌ 請輸入電子郵件")
                        elif not user_exists_in_db(reset_email):
                            st.error("❌ 找不到該電子郵件的註冊帳號，請確認是否在本站註冊過")
                        elif not new_pw:
                            st.error("❌ 請輸入新密碼")
                        else:
                            ok_pw, msg_pw = validate_password_strength(new_pw)
                            if not ok_pw:
                                st.error(f"❌ {msg_pw}")
                            elif new_pw != new_pw_confirm:
                                st.error("❌ 兩次輸入的密碼不一致")
                            else:
                                ok, msg = update_user_password(reset_email.strip(), new_pw)
                                if ok:
                                    st.success(f"✅ {msg}")
                                    st.session_state.show_forgot_password = False
                                    time.sleep(0.8)
                                    st.rerun()
                                else:
                                    st.error(f"❌ {msg}")
                with r2:
                    if st.button("← 返回登入", use_container_width=True, key="btn_back_login"):
                        st.session_state.show_forgot_password = False
                        st.rerun()
            else:
                email = st.text_input("📧 電子郵件", key="login_email", label_visibility="visible", 
                                     placeholder="you@example.com")
                password = st.text_input("🔑 密碼", type="password", key="login_password", 
                                        label_visibility="visible")
                if st.button("忘記密碼？", key="link_forgot_pw"):
                    st.session_state.show_forgot_password = True
                    st.rerun()
                
                col_btn1, col_btn2 = st.columns([1, 1])
                with col_btn1:
                    if st.button("🔑 登入", type="primary", use_container_width=True):
                        if not email:
                            st.error("❌ 請輸入電子郵件")
                        elif not password:
                            st.error("❌ 請輸入密碼")
                        else:
                            success, message = verify_user(email.strip(), password)
                            if success:
                                st.session_state.authenticated = True
                                st.session_state.user_email = email.strip()
                                st.session_state.login_at = datetime.now().isoformat()
                                st.session_state.pop("login_csrf_token", None)
                                st.success(f"✅ {message}")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.error(f"❌ {message}")
                                if "電子郵件或密碼錯誤" in message or "郵箱或密碼錯誤" in message:
                                    st.info("💡 若忘記密碼，可點上方「忘記密碼？」重設（僅限本站註冊帳號）")
                
                with col_btn2:
                    # 第三方登入：Google / LINE / Facebook（需在 Secrets 或環境變數設定對應金鑰）
                    oauth_col1, oauth_col2, oauth_col3 = st.columns(3)
                    with oauth_col1:
                        if st.button("🔵 Google", use_container_width=True, key="btn_google"):
                            st.session_state.oauth_pending = "google"
                            st.rerun()
                    with oauth_col2:
                        if st.button("🟢 LINE", use_container_width=True, key="btn_line"):
                            st.session_state.oauth_pending = "line"
                            st.rerun()
                    with oauth_col3:
                        if st.button("🔷 Facebook", use_container_width=True, key="btn_facebook"):
                            st.session_state.oauth_pending = "facebook"
                            st.rerun()
                    st.caption("若未設定金鑰，請使用上方電子郵件與密碼登入。")
            
        else:  # 註冊模式
            email = st.text_input("📧 新帳號電子郵件", key="reg_email", label_visibility="visible", 
                                 placeholder="you@example.com")
            password = st.text_input("🔒 密碼（至少 8 字元，含大小寫/數字/符號其中兩類）", type="password", key="reg_password", 
                                    label_visibility="visible")
            confirm = st.text_input("🔒 再輸入一次密碼", type="password", key="reg_confirm", 
                                   label_visibility="visible")
            
            if st.button("✅ 建立帳號", type="primary", use_container_width=True):
                if not email:
                    st.error("❌ 請輸入電子郵件")
                elif not password:
                    st.error("❌ 請輸入密碼")
                elif password != confirm:
                    st.error("❌ 兩次密碼不一致")
                else:
                    success, message = register_user(email, password)
                    if success:
                        st.session_state.authenticated = True
                        st.session_state.user_email = email.strip()
                        st.session_state.login_at = datetime.now().isoformat()
                        st.session_state.pop("login_csrf_token", None)
                        st.success(f"✅ {message}")
                        time.sleep(0.5)
                        st.rerun()
                    else:
                        st.error(f"❌ {message}")
        
        st.markdown('</div>', unsafe_allow_html=True)

# --- 2. 終極韌性資料庫連線器 ---
def get_db_path():
    if "current_db_path" not in st.session_state:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        # 改用 invoices_v2.db 以解決舊資料庫可能發生的 Disk I/O Error 鎖定問題
        db_file = os.path.join(base_dir, "invoices_v2.db")
        
        # 確保目錄存在
        try:
            os.makedirs(base_dir, exist_ok=True)
        except Exception as e:
            st.session_state.db_error = f"無法創建目錄: {str(e)}"
            st.session_state.current_db_path = "file:invoice_mem?mode=memory&cache=shared"
            st.session_state.db_path_mode = "🧠 虛擬記憶體 (重啟會清空)"
            return st.session_state.current_db_path
        
        try:
            # 檢查寫入權限（不創建測試文件，直接嘗試創建目錄）
            # 強制使用文件數據庫，不使用內存模式
            st.session_state.current_db_path = db_file
            st.session_state.db_path_mode = "💾 本地磁碟"
            st.session_state.db_error = None
        except PermissionError as e:
            st.session_state.db_error = f"權限不足，無法寫入: {str(e)}"
            # 即使權限不足，也嘗試使用文件數據庫（可能可以讀取）
            st.session_state.current_db_path = db_file
            st.session_state.db_path_mode = "⚠️ 只讀模式"
        except Exception as e:
            st.session_state.db_error = f"無法使用文件數據庫: {str(e)}"
            # 最後才使用內存模式
            st.session_state.current_db_path = "file:invoice_mem?mode=memory&cache=shared"
            st.session_state.db_path_mode = "🧠 虛擬記憶體 (重啟會清空)"
    
    return st.session_state.current_db_path

def init_db():
    """初始化資料表，確保所有必要欄位存在（多用戶版本：含 users 表）"""
    if st.session_state.use_memory_mode:
        return True  # 使用內存模式，跳過數據庫初始化
    
    path = get_db_path()
    # 判斷是否為URI模式（只有明確包含mode=memory才是URI）
    # 普通文件路徑（如 invoices_v2.db）不是URI
    is_uri = path.startswith("file:") and "mode=memory" in path
    try:
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        
        # ① 創建 users 表（多用戶版本；含第三方登入 ID）
        cursor.execute('''CREATE TABLE IF NOT EXISTS users
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         email TEXT UNIQUE NOT NULL,
                         password_hash TEXT,
                         google_id TEXT,
                         line_id TEXT,
                         facebook_id TEXT,
                         created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                         last_login TIMESTAMP)''')
        # 補全 users 表欄位（舊庫可能無 line_id, facebook_id）
        for col in ("line_id", "facebook_id"):
            try:
                cursor.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT")
            except Exception:
                pass
        
        # ② 登入嘗試表（AUTH-03：失敗鎖定）
        cursor.execute('''CREATE TABLE IF NOT EXISTS login_attempts
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         account_key TEXT NOT NULL,
                         attempt_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                         success INTEGER NOT NULL)''')
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_login_attempts_key ON login_attempts(account_key)")
        except Exception:
            pass
        
        # ③ 創建 invoices 表（多用戶版本：使用 user_email）
        cursor.execute('''CREATE TABLE IF NOT EXISTS invoices
                        (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                        user_email TEXT NOT NULL,
                        file_name TEXT, date TEXT, invoice_number TEXT, seller_name TEXT, seller_ubn TEXT,
                        subtotal REAL, tax REAL, total REAL, category TEXT, subject TEXT, status TEXT,
                        note TEXT,
                        image_path TEXT, image_data BLOB,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        
        # ③ 補全欄位（兼容舊版本）
        # 檢查是否有 user_id 欄位，如果有則遷移到 user_email
        try:
            cursor.execute("SELECT user_id FROM invoices LIMIT 1")
            # 如果有 user_id，嘗試遷移數據
            cursor.execute("UPDATE invoices SET user_email = user_id WHERE user_email IS NULL OR user_email = ''")
            # 刪除舊的 user_id 欄位（SQLite不支持直接刪除，但可以忽略）
        except:
            pass
        
        # 添加 user_email 欄位（如果不存在）
        try:
            cursor.execute("ALTER TABLE invoices ADD COLUMN user_email TEXT")
        except:
            pass
        
        # 添加 note 欄位（備註）（如果不存在）
        try:
            cursor.execute("ALTER TABLE invoices ADD COLUMN note TEXT")
        except:
            pass
        
        # 為 user_email 創建索引（提高查詢效率）
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_user_email ON invoices(user_email)")
        except:
            pass
        
        # 補全其他欄位
        for col, c_type in {'status': "TEXT", 'seller_ubn': "TEXT", 
                            'image_path': "TEXT", 'image_data': "BLOB"}.items():
            try: 
                cursor.execute(f"ALTER TABLE invoices ADD COLUMN {col} {c_type}")
            except: 
                pass
        
        # 邏輯架構說明書：modified_at、batch_id、tax_type
        for col, c_type in {'modified_at': "TIMESTAMP", 'batch_id': "INTEGER", 'tax_type': "TEXT DEFAULT '5%'"}.items():
            try:
                cursor.execute(f"ALTER TABLE invoices ADD COLUMN {col} {c_type}")
            except:
                pass
        
        # 創建 batches 表（上傳組：同一次 OCR 或導入為一組）
        cursor.execute('''CREATE TABLE IF NOT EXISTS batches
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         user_email TEXT NOT NULL,
                         source TEXT NOT NULL,
                         created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                         invoice_count INTEGER)''')
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_batches_user ON batches(user_email)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_invoices_batch ON invoices(batch_id)")
        except Exception:
            pass
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.session_state.db_error = f"初始化失敗: {str(e)}"
        return False


# --- 下一階段：Google / LINE / Facebook OAuth 登入 ---
from urllib.parse import urlencode

def _get_oauth_redirect_uri():
    """取得 OAuth 回調網址（須與各平台後台設定一致）。"""
    base = _safe_secrets_get("OAUTH_REDIRECT_URI") or os.getenv("OAUTH_REDIRECT_URI")
    if base:
        return base.rstrip("/") + "/"
    # Streamlit 預設：本機或雲端由執行環境決定
    try:
        from streamlit.web.server.server import SERVER_PORT
        return f"http://localhost:{SERVER_PORT}/"
    except Exception:
        return "http://localhost:8501/"

def _build_oauth_state(provider: str) -> str:
    """產生 state 並存入 session，回傳 state 字串。"""
    token = _secrets_module.token_hex(16)
    state = f"{provider}:{token}"
    if "oauth_state" not in st.session_state:
        st.session_state.oauth_state = {}
    st.session_state.oauth_state[provider] = token
    return state

def _verify_oauth_state(provider: str, state: str) -> bool:
    """驗證 state 與 session 內儲存一致。"""
    if not state or ":" not in state:
        return False
    parts = state.split(":", 1)
    if parts[0] != provider:
        return False
    return st.session_state.get("oauth_state", {}).get(provider) == parts[1]

def build_oauth_url_google():
    """建立 Google OAuth 授權 URL。"""
    client_id = _safe_secrets_get("GOOGLE_CLIENT_ID") or os.getenv("GOOGLE_CLIENT_ID")
    if not client_id:
        return None, "未設定 GOOGLE_CLIENT_ID"
    redirect_uri = _get_oauth_redirect_uri()
    state = _build_oauth_state("google")
    scope = "openid email profile"
    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": scope,
        "state": state,
        "access_type": "offline",
        "prompt": "consent",
    }
    return "https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params), None

def build_oauth_url_line():
    """建立 LINE Login 授權 URL。"""
    client_id = _safe_secrets_get("LINE_CHANNEL_ID") or os.getenv("LINE_CHANNEL_ID")
    if not client_id:
        return None, "未設定 LINE_CHANNEL_ID"
    redirect_uri = _get_oauth_redirect_uri()
    state = _build_oauth_state("line")
    params = {
        "response_type": "code",
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "state": state,
        "scope": "profile openid",
    }
    return "https://access.line.me/oauth2/v2.1/authorize?" + urlencode(params), None

def build_oauth_url_facebook():
    """建立 Facebook Login 授權 URL。"""
    app_id = _safe_secrets_get("FACEBOOK_APP_ID") or os.getenv("FACEBOOK_APP_ID")
    if not app_id:
        return None, "未設定 FACEBOOK_APP_ID"
    redirect_uri = _get_oauth_redirect_uri()
    state = _build_oauth_state("facebook")
    params = {
        "client_id": app_id,
        "redirect_uri": redirect_uri,
        "state": state,
        "scope": "email,public_profile",
        "response_type": "code",
    }
    return "https://www.facebook.com/v18.0/dialog/oauth?" + urlencode(params), None

def _oauth_find_or_create_user(provider: str, email: str, provider_user_id: str) -> tuple[bool, str]:
    """
    依第三方 ID 或 email 查詢用戶；不存在則建立。回傳 (success, user_email 或錯誤訊息)。
    email 可為空（LINE 可能無 email），此時以佔位 email 建立。
    """
    init_db()
    path = get_db_path()
    is_uri = path.startswith("file:") and "mode=memory" in path
    id_col = {"google": "google_id", "line": "line_id", "facebook": "facebook_id"}.get(provider)
    if not id_col:
        return False, "不支援的登入方式"
    try:
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        # 先以第三方 ID 查詢
        cursor.execute(f"SELECT id, email FROM users WHERE {id_col} = ?", (provider_user_id,))
        row = cursor.fetchone()
        if row:
            conn.close()
            return True, row[1]
        # 再以 email 查詢（若有的話），並綁定該第三方 ID
        if email:
            cursor.execute("SELECT id, email FROM users WHERE email = ?", (email,))
            row = cursor.fetchone()
            if row:
                cursor.execute(f"UPDATE users SET {id_col} = ?, last_login = ? WHERE id = ?",
                              (provider_user_id, datetime.now().isoformat(), row[0]))
                conn.commit()
                conn.close()
                return True, row[1]
        # 建立新用戶
        new_email = email if email else f"{provider}_{provider_user_id}@oauth.local"
        cursor.execute(
            f"INSERT INTO users (email, password_hash, {id_col}, last_login) VALUES (?, NULL, ?, ?)",
            (new_email, provider_user_id, datetime.now().isoformat()),
        )
        conn.commit()
        conn.close()
        return True, new_email
    except Exception as e:
        return False, str(e)

def handle_oauth_callback_google(code: str, state: str) -> tuple[bool, str]:
    """處理 Google OAuth callback：以 code 換 token，取得使用者，查詢或建立用戶。回傳 (success, user_email 或錯誤訊息)。"""
    if not _verify_oauth_state("google", state):
        return False, "登入驗證已過期，請重新點選 Google 登入"
    client_id = _safe_secrets_get("GOOGLE_CLIENT_ID") or os.getenv("GOOGLE_CLIENT_ID")
    client_secret = _safe_secrets_get("GOOGLE_CLIENT_SECRET") or os.getenv("GOOGLE_CLIENT_SECRET")
    if not client_id or not client_secret:
        return False, "未設定 Google 登入參數"
    redirect_uri = _get_oauth_redirect_uri()
    try:
        r = requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uri": redirect_uri,
                "grant_type": "authorization_code",
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        access_token = data.get("access_token")
        if not access_token:
            return False, "無法取得存取權杖"
        info = requests.get(
            "https://www.googleapis.com/oauth2/v2/userinfo",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        info.raise_for_status()
        user_info = info.json()
        email = (user_info.get("email") or "").strip()
        sub = user_info.get("id") or ""
        if not sub:
            return False, "無法取得 Google 帳號資訊"
        return _oauth_find_or_create_user("google", email, sub)
    except requests.RequestException as e:
        return False, "登入連線失敗，請稍後再試"
    except Exception as e:
        return False, "登入失敗，請稍後再試"

def handle_oauth_callback_line(code: str, state: str) -> tuple[bool, str]:
    """處理 LINE OAuth callback。"""
    if not _verify_oauth_state("line", state):
        return False, "登入驗證已過期，請重新點選 LINE 登入"
    client_id = _safe_secrets_get("LINE_CHANNEL_ID") or os.getenv("LINE_CHANNEL_ID")
    client_secret = _safe_secrets_get("LINE_CHANNEL_SECRET") or os.getenv("LINE_CHANNEL_SECRET")
    if not client_id or not client_secret:
        return False, "未設定 LINE 登入參數"
    redirect_uri = _get_oauth_redirect_uri()
    try:
        r = requests.post(
            "https://api.line.me/oauth2/v2.1/token",
            data={
                "grant_type": "authorization_code",
                "code": code,
                "redirect_uri": redirect_uri,
                "client_id": client_id,
                "client_secret": client_secret,
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        access_token = data.get("access_token")
        if not access_token:
            return False, "無法取得存取權杖"
        profile = requests.get(
            "https://api.line.me/v2/profile",
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        profile.raise_for_status()
        user_info = profile.json()
        user_id = user_info.get("userId") or ""
        if not user_id:
            return False, "無法取得 LINE 帳號資訊"
        email = (user_info.get("email") or "").strip()
        return _oauth_find_or_create_user("line", email, user_id)
    except requests.RequestException:
        return False, "登入連線失敗，請稍後再試"
    except Exception:
        return False, "登入失敗，請稍後再試"

def handle_oauth_callback_facebook(code: str, state: str) -> tuple[bool, str]:
    """處理 Facebook OAuth callback。"""
    if not _verify_oauth_state("facebook", state):
        return False, "登入驗證已過期，請重新點選 Facebook 登入"
    app_id = _safe_secrets_get("FACEBOOK_APP_ID") or os.getenv("FACEBOOK_APP_ID")
    app_secret = _safe_secrets_get("FACEBOOK_APP_SECRET") or os.getenv("FACEBOOK_APP_SECRET")
    if not app_id or not app_secret:
        return False, "未設定 Facebook 登入參數"
    redirect_uri = _get_oauth_redirect_uri()
    try:
        r = requests.get(
            "https://graph.facebook.com/v18.0/oauth/access_token",
            params={
                "client_id": app_id,
                "client_secret": app_secret,
                "redirect_uri": redirect_uri,
                "code": code,
            },
            timeout=15,
        )
        r.raise_for_status()
        data = r.json()
        access_token = data.get("access_token")
        if not access_token:
            return False, "無法取得存取權杖"
        info = requests.get(
            "https://graph.facebook.com/me",
            params={"fields": "id,name,email"},
            headers={"Authorization": f"Bearer {access_token}"},
            timeout=10,
        )
        info.raise_for_status()
        user_info = info.json()
        fb_id = user_info.get("id") or ""
        if not fb_id:
            return False, "無法取得 Facebook 帳號資訊"
        email = (user_info.get("email") or "").strip()
        return _oauth_find_or_create_user("facebook", email, fb_id)
    except requests.RequestException:
        return False, "登入連線失敗，請稍後再試"
    except Exception:
        return False, "登入失敗，請稍後再試"

def run_query(query, params=(), is_select=True):
    """
    執行資料庫查詢（多用戶版本：自動添加 user_email 隔離）
    - SELECT: 自動添加 WHERE user_email = ? 條件（如果查詢 invoices 表且沒有 user_email 條件）
    - INSERT: 自動添加 user_email 參數（如果插入 invoices 表）
    """
    user_email = st.session_state.get('user_email')
    if not user_email:
        user_email = "default_user"  # 未登錄時使用默認用戶
    
    # 如果使用內存模式，使用 session_state 存儲
    if st.session_state.use_memory_mode:
        if is_select:
            # 處理 SELECT 查詢 - 自動過濾 user_email
            # 修復 Bug #1: 使用安全的字符串比較，避免 SQL 注入風險
            # 雖然是內存模式，但保持一致的編碼風格和安全性
            df = pd.DataFrame([inv for inv in st.session_state.local_invoices 
                             if inv.get('user_email', inv.get('user_id', 'default_user')) == user_email])
            
            # 簡單的 ORDER BY 處理
            if "ORDER BY id DESC" in query.upper():
                if not df.empty and 'id' in df.columns:
                    df = df.sort_values('id', ascending=False)
            
            return df
        else:
            # INSERT 查詢會在調用處處理，確保包含 user_email
            return True
    
    # 使用數據庫
    path = get_db_path()
    # 判斷是否為URI模式（只有明確包含mode=memory或file:前綴才是URI）
    is_uri = (path.startswith("file:") and "mode=memory" in path) or path.startswith("file:invoice_mem")
    try:
        conn = sqlite3.connect(path, timeout=30, check_same_thread=False, uri=is_uri)
        cursor = conn.cursor()
        
        if is_select:
            # 多用戶隔離：自動添加 user_email 條件（僅對 invoices 表）
            modified_query = query
            modified_params = list(params)
            
            if "FROM invoices" in query.upper() and "user_email" not in query.upper() and "WHERE" not in query.upper():
                # 如果查詢 invoices 表且沒有 WHERE 條件，添加 user_email 過濾
                modified_query = query + " WHERE user_email = ?"
                modified_params = [user_email] + list(params)
            elif "FROM invoices" in query.upper() and "user_email" not in query.upper() and "WHERE" in query.upper():
                # 如果有 WHERE 條件但沒有 user_email，添加 AND user_email = ?
                # 找到 WHERE 的位置，在其後添加
                where_pos = query.upper().find("WHERE")
                modified_query = query[:where_pos+5] + " user_email = ? AND " + query[where_pos+5:]
                modified_params = [user_email] + list(params)
            
            try:
                df = pd.read_sql_query(modified_query, conn, params=tuple(modified_params))
            except Exception as e:
                # 關鍵修復：如果發現沒表，自動初始化並重試
                if "no such table" in str(e).lower():
                    if init_db():
                        df = pd.read_sql_query(modified_query, conn, params=tuple(modified_params))
                    else:
                        # 初始化失敗，切換到內存模式
                        st.session_state.use_memory_mode = True
                        return run_query(query, params, is_select)
                else: 
                    raise e
            conn.close()
            st.session_state.db_error = None
            return df
        else:
            # 非SELECT查询：INSERT, UPDATE, DELETE等
            # 對於 INSERT invoices，確保包含 user_email
            modified_query = query
            modified_params = list(params)
            
            if "INSERT INTO invoices" in query.upper() and "user_email" not in query.upper():
                # 自動添加 user_email 到 INSERT 語句
                # 找到 VALUES 的位置
                values_pos = query.upper().find("VALUES")
                if values_pos > 0:
                    # 在列名中添加 user_email
                    insert_part = query[:values_pos]
                    values_part = query[values_pos:]
                    
                    # 在列名列表中添加 user_email
                    if "(" in insert_part:
                        last_paren = insert_part.rfind(")")
                        insert_part = insert_part[:last_paren] + ", user_email" + insert_part[last_paren:]
                    
                    # 在 VALUES 中添加 user_email 值
                    if "(" in values_part:
                        first_paren = values_part.find("(")
                        last_paren = values_part.rfind(")")
                        values_part = values_part[:first_paren+1] + "?, " + values_part[first_paren+1:last_paren] + ", ?" + values_part[last_paren:]
                    
                    modified_query = insert_part + values_part
                    modified_params = [user_email] + list(params)
            
            # 對於 UPDATE 和 DELETE，確保包含 user_email 條件
            if ("UPDATE invoices" in query.upper() or "DELETE FROM invoices" in query.upper()) and "user_email" not in query.upper():
                if "WHERE" in query.upper():
                    where_pos = query.upper().find("WHERE")
                    modified_query = query[:where_pos+5] + " user_email = ? AND " + query[where_pos+5:]
                    modified_params = [user_email] + list(params)
                else:
                    # 如果沒有 WHERE，添加 WHERE user_email = ?
                    modified_query = query + " WHERE user_email = ?"
                    modified_params = [user_email] + list(params)
            
            try:
                cursor.execute(modified_query, tuple(modified_params))
                conn.commit()
                # 验证是否真的执行成功
                if "INSERT" in modified_query.upper():
                    # 对于INSERT，检查影响的行数
                    if cursor.rowcount > 0:
                        conn.close()
                        return True
                    else:
                        conn.close()
                        st.session_state.db_error = "插入失敗：影響行數為0"
                        return False
                conn.close()
                return True
            except Exception as e:
                conn.rollback()
                conn.close()
                st.session_state.db_error = f"執行失敗: {str(e)}"
                return False
    except Exception as e:
        err_msg = str(e)
        st.session_state.db_error = f"連線異常: {err_msg}"
        # 如果數據庫失敗，自動切換到內存模式
        if "no such table" in err_msg.lower() or "unable to open" in err_msg.lower():
            st.session_state.use_memory_mode = True
            if is_select:
                return run_query(query, params, is_select)
        return pd.DataFrame() if is_select else False

# 程式啟動立即初始化（如果使用數據庫模式）
if not st.session_state.use_memory_mode:
    init_db()

import re

# ... (保持前面的 import 不变)

# --- 3. 核心辨識邏輯 ---
def extract_json(text):
    """從混合文本中提取有效的 JSON 物件"""
    text = text.strip()
    # 嘗試 1: 直接解析
    try:
        return json.loads(text)
    except:
        pass
        
    # 嘗試 2: 尋找 Markdown 代碼塊 ```json ... ```
    match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if match:
        try: return json.loads(match.group(1))
        except: pass
        
    # 嘗試 3: 尋找最外層的 {}
    match = re.search(r'(\{.*\})', text, re.DOTALL)
    if match:
        try: return json.loads(match.group(1))
        except: pass
        
    return None

def save_invoice_image(image_obj, file_name, user_email=None):
    """保存發票圖片到文件系統，返回圖片路徑"""
    try:
        # 創建用戶專屬目錄
        user_email = user_email or st.session_state.get('user_email', 'default_user')
        user_dir = os.path.join(st.session_state.image_storage_dir, user_email)
        os.makedirs(user_dir, exist_ok=True)
        
        # 生成唯一文件名（使用時間戳+文件名hash）
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_hash = hashlib.md5(file_name.encode()).hexdigest()[:8]
        safe_filename = f"{timestamp}_{file_hash}_{file_name}"
        safe_filename = "".join(c for c in safe_filename if c.isalnum() or c in "._-")
        
        image_path = os.path.join(user_dir, safe_filename)
        image_obj.save(image_path)
        return image_path
    except Exception as e:
        st.error(f"保存圖片失敗: {str(e)}")
        return None

def create_batch(user_email, source):
    """建立一筆上傳組（Batch），回傳 batch_id。source 為 'ocr' 或 'import'。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    if st.session_state.use_memory_mode:
        batch_id = len(st.session_state.local_batches) + 1
        st.session_state.local_batches.append({
            'id': batch_id,
            'user_email': user_email,
            'source': source,
            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'invoice_count': 0
        })
        return batch_id
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO batches (user_email, source) VALUES (?, ?)", (user_email, source))
        batch_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return batch_id
    except Exception:
        return None

def check_duplicate_invoice(invoice_number, date, user_email=None):
    """檢查是否為重複發票（根據發票號碼+日期）"""
    if not invoice_number or invoice_number == "No" or invoice_number == "N/A":
        return False, None
    
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    if st.session_state.use_memory_mode:
        # 內存模式檢查（多用戶版本：使用 user_email）
        for inv in st.session_state.local_invoices:
            inv_user = inv.get('user_email', inv.get('user_id', 'default_user'))
            if (inv_user == user_email and 
                inv.get('invoice_number') == invoice_number and 
                inv.get('date') == date):
                return True, inv.get('id')
    else:
        # 數據庫模式檢查（多用戶版本：使用 user_email）
        query = "SELECT id FROM invoices WHERE user_email = ? AND invoice_number = ? AND date = ?"
        result = run_query(query, (user_email, invoice_number, date), is_select=True)
        if not result.empty:
            return True, result.iloc[0]['id']
    
    return False, None


def get_batches_for_user(user_email=None):
    """取得當前用戶的 Batch 列表（說明書 § 三：按組顯示用）。回傳 list of dict: id, user_email, source, created_at, invoice_count。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    if st.session_state.use_memory_mode:
        batches = [b for b in st.session_state.local_batches if b.get('user_email') == user_email]
        for b in batches:
            b['invoice_count'] = len([inv for inv in st.session_state.local_invoices 
                                      if inv.get('batch_id') == b.get('id') and inv.get('user_email', inv.get('user_id', '')) == user_email])
        batches.sort(key=lambda x: x.get('created_at', ''), reverse=True)
        return batches
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("""
            SELECT b.id, b.user_email, b.source, b.created_at,
                   (SELECT COUNT(*) FROM invoices i WHERE i.batch_id = b.id AND i.user_email = b.user_email) AS invoice_count
            FROM batches b WHERE b.user_email = ? ORDER BY b.created_at DESC
        """, (user_email,))
        rows = cursor.fetchall()
        conn.close()
        return [{'id': r[0], 'user_email': r[1], 'source': r[2], 'created_at': r[3], 'invoice_count': r[4] or 0} for r in rows]
    except Exception:
        return []


def get_invoices_by_batch(batch_id, user_email=None):
    """取得指定 Batch 下的發票（說明書 § 三：按組顯示用）。回傳已重命名欄位的 DataFrame。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    mapping = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編",
               "subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間"}
    if st.session_state.use_memory_mode:
        rows = [inv for inv in st.session_state.local_invoices 
                if inv.get('batch_id') == batch_id and inv.get('user_email', inv.get('user_id', '')) == user_email]
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        df = df.rename(columns=mapping)
        return df
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM invoices WHERE batch_id = ? AND user_email = ? ORDER BY id", (batch_id, user_email))
        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchall()
        conn.close()
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=cols)
        df = df.rename(columns=mapping)
        return df
    except Exception:
        return pd.DataFrame()


def delete_batch_cascade(batch_id, user_email=None):
    """刪除一組 Batch 及其下所有發票（說明書 Cascade Delete）。回傳 (success, deleted_invoices_count, error_message)。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    if st.session_state.use_memory_mode:
        before = len(st.session_state.local_invoices)
        st.session_state.local_invoices = [
            inv for inv in st.session_state.local_invoices
            if not (inv.get('batch_id') == batch_id and inv.get('user_email', inv.get('user_id', '')) == user_email)
        ]
        deleted = before - len(st.session_state.local_invoices)
        st.session_state.local_batches = [b for b in st.session_state.local_batches if not (b.get('id') == batch_id and b.get('user_email') == user_email)]
        return True, deleted, None
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM invoices WHERE batch_id = ? AND user_email = ?", (batch_id, user_email))
        deleted = cursor.rowcount
        cursor.execute("DELETE FROM batches WHERE id = ? AND user_email = ?", (batch_id, user_email))
        conn.commit()
        conn.close()
        return True, deleted, None
    except Exception as e:
        return False, 0, str(e)


def load_test_data(user_email=None):
    """載入測試數據：2 個 Batch（OCR 4 張 + 導入 3 張）+ 2 張未分組發票。回傳 (成功筆數, 錯誤訊息)。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    test_invoices = [
        # Batch 1 (OCR) - 4 張
        {"file_name": "發票1.jpg", "date": "2026/01/15", "invoice_number": "AB-12345678", "seller_name": "全家便利商店", "seller_ubn": "12345678", "subtotal": 95.24, "tax": 4.76, "total": 100, "category": "餐飲", "subject": "餐飲費", "status": "✅ 正常", "note": "", "tax_type": "5%"},
        {"file_name": "發票2.jpg", "date": "2026/01/16", "invoice_number": "AB-12345679", "seller_name": "統一超商", "seller_ubn": "87654321", "subtotal": 190.48, "tax": 9.52, "total": 200, "category": "餐飲", "subject": "餐飲費", "status": "✅ 正常", "note": "", "tax_type": "5%"},
        {"file_name": "發票3.jpg", "date": "2026/01/17", "invoice_number": "AB-12345680", "seller_name": "星巴克", "seller_ubn": "11112222", "subtotal": 142.86, "tax": 7.14, "total": 150, "category": "餐飲", "subject": "餐飲費", "status": "✅ 正常", "note": "咖啡", "tax_type": "5%"},
        {"file_name": "發票4.jpg", "date": "2026/01/18", "invoice_number": "AB-12345681", "seller_name": "麥當勞", "seller_ubn": "22223333", "subtotal": 380.95, "tax": 19.05, "total": 400, "category": "餐飲", "subject": "餐飲費", "status": "✅ 正常", "note": "會議餐", "tax_type": "5%"},
        # Batch 2 (import) - 3 張（含 5% 與 零稅率）
        {"file_name": "導入數據", "date": "2026/01/20", "invoice_number": "CD-88880001", "seller_name": "文具王", "seller_ubn": "33334444", "subtotal": 476.19, "tax": 23.81, "total": 500, "category": "辦公用品", "subject": "辦公用品", "status": "✅ 正常", "note": "影印紙", "tax_type": "5%"},
        {"file_name": "導入數據", "date": "2026/01/21", "invoice_number": "CD-88880002", "seller_name": "台灣大車隊", "seller_ubn": "55556666", "subtotal": 285.71, "tax": 14.29, "total": 300, "category": "交通", "subject": "交通費", "status": "✅ 正常", "note": "計程車", "tax_type": "5%"},
        {"file_name": "導入數據", "date": "2026/01/22", "invoice_number": "CD-88880003", "seller_name": "出口供應商", "seller_ubn": "66667777", "subtotal": 1000, "tax": 0, "total": 1000, "category": "其他", "subject": "採購", "status": "✅ 正常", "note": "零稅率", "tax_type": "零稅率"},
        # 未分組 - 2 張（免稅 / 0%）
        {"file_name": "舊資料", "date": "2026/01/10", "invoice_number": "EF-00000001", "seller_name": "免稅店", "seller_ubn": "77778888", "subtotal": 100, "tax": 0, "total": 100, "category": "其他", "subject": "雜項", "status": "✅ 正常", "note": "未分組", "tax_type": "免稅"},
        {"file_name": "舊資料", "date": "2026/01/11", "invoice_number": "EF-00000002", "seller_name": "零稅率供應商", "seller_ubn": "99990000", "subtotal": 200, "tax": 0, "total": 200, "category": "其他", "subject": "雜項", "status": "✅ 正常", "note": "未分組", "tax_type": "0%"},
    ]
    batch_sources = ["ocr", "ocr", "ocr", "ocr", "import", "import", "import", None, None]  # 前 4 屬 batch1, 5-7 屬 batch2, 8-9 未分組

    if st.session_state.use_memory_mode:
        bid_ocr = len(st.session_state.local_batches) + 1
        st.session_state.local_batches.append({"id": bid_ocr, "user_email": user_email, "source": "ocr", "created_at": now, "invoice_count": 4})
        bid_import = len(st.session_state.local_batches) + 1
        st.session_state.local_batches.append({"id": bid_import, "user_email": user_email, "source": "import", "created_at": now, "invoice_count": 3})
        base_id = len(st.session_state.local_invoices)
        for i, inv in enumerate(test_invoices):
            rec = dict(inv)
            rec["id"] = base_id + i + 1
            rec["user_email"] = user_email
            rec["image_path"] = None
            rec["created_at"] = now
            rec["modified_at"] = None
            rec["batch_id"] = bid_ocr if batch_sources[i] == "ocr" else (bid_import if batch_sources[i] == "import" else None)
            st.session_state.local_invoices.append(rec)
        return len(test_invoices), None

    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO batches (user_email, source) VALUES (?, 'ocr')", (user_email,))
        bid_ocr = cursor.lastrowid
        cursor.execute("INSERT INTO batches (user_email, source) VALUES (?, 'import')", (user_email,))
        bid_import = cursor.lastrowid
        for i, inv in enumerate(test_invoices):
            batch_id = None
            if batch_sources[i] == "ocr":
                batch_id = bid_ocr
            elif batch_sources[i] == "import":
                batch_id = bid_import
            cursor.execute("""
                INSERT INTO invoices (user_email, file_name, date, invoice_number, seller_name, seller_ubn, subtotal, tax, total, category, subject, status, note, batch_id, tax_type)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                user_email, inv["file_name"], inv["date"], inv["invoice_number"], inv["seller_name"], inv["seller_ubn"],
                inv["subtotal"], inv["tax"], inv["total"], inv["category"], inv["subject"], inv["status"], inv["note"], batch_id, inv["tax_type"]
            ))
        conn.commit()
        conn.close()
        return len(test_invoices), None
    except Exception as e:
        return 0, str(e)


def get_ungrouped_invoices(user_email=None):
    """取得未分組發票（batch_id 為 NULL）。回傳已重命名欄位的 DataFrame。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    mapping = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編",
               "subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間"}
    if st.session_state.use_memory_mode:
        rows = [inv for inv in st.session_state.local_invoices 
                if (inv.get('batch_id') is None or inv.get('batch_id') == '') and inv.get('user_email', inv.get('user_id', '')) == user_email]
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows)
        df = df.rename(columns=mapping)
        return df
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM invoices WHERE (batch_id IS NULL OR batch_id = '') AND user_email = ? ORDER BY id", (user_email,))
        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchall()
        conn.close()
        if not rows:
            return pd.DataFrame()
        df = pd.DataFrame(rows, columns=cols)
        df = df.rename(columns=mapping)
        return df
    except Exception:
        return pd.DataFrame()


def get_invoice_by_id(invoice_id, user_email=None):
    """依 id 取得單筆發票（用於詳情彈出框）。回傳 dict（中文欄位名）或 None。"""
    user_email = user_email or st.session_state.get('user_email', 'default_user')
    mapping = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編",
               "subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間","modified_at":"修改時間"}
    if st.session_state.use_memory_mode:
        for inv in st.session_state.local_invoices:
            if inv.get('id') == invoice_id and inv.get('user_email', inv.get('user_id', '')) == user_email:
                d = dict(inv)
                return {mapping.get(k, k): v for k, v in d.items()}
        return None
    try:
        path = get_db_path()
        is_uri = path.startswith("file:") and "mode=memory" in path
        conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM invoices WHERE id = ? AND user_email = ?", (invoice_id, user_email))
        row = cursor.fetchone()
        cols = [d[0] for d in cursor.description]
        conn.close()
        if not row:
            return None
        d = dict(zip(cols, row))
        return {mapping.get(k, k): v for k, v in d.items()}
    except Exception:
        return None


def validate_ubn(val):
    """台灣統編驗證：8 位數字（選填時空值視為通過）。回傳 (ok, message)。"""
    if val is None or (isinstance(val, str) and not str(val).strip()):
        return True, ""
    s = str(val).strip()
    if len(s) != 8:
        return False, "統編應為 8 位數字"
    if not s.isdigit():
        return False, "統編應為 8 位數字"
    return True, ""


def normalize_invoice_number(raw):
    """發票號碼正規化：回傳 8 碼數字字串，無法解析則回傳 None。"""
    if raw is None:
        return None
    s = str(raw)
    # 僅保留數字
    digits = re.sub(r"\D", "", s)
    if len(digits) < 8:
        return None
    # 取最後 8 碼（避免前面有年份等干擾）
    return digits[-8:]


# 財政部統一發票中獎號碼（新站）：僅兩頁
# https://invoice.etax.nat.gov.tw/index.html = 最新一期，lastNumber.html = 上一期
LOTTERY_ETAX_LATEST = "https://invoice.etax.nat.gov.tw/index.html"
LOTTERY_ETAX_PREVIOUS = "https://invoice.etax.nat.gov.tw/lastNumber.html"

# 爬蟲用 headers，模擬瀏覽器避免被擋
LOTTERY_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}


def _parse_lottery_draw_from_html(text):
    """從開獎頁 HTML 文字解析開獎號碼。不依賴標籤結構，只依「特別獎」後依序的 8 碼數字。
    回傳 (draw_dict, error_message)。成功時 error_message 為 None。
    """
    if not text or "特別獎" not in text:
        return None, "頁面內容中找不到「特別獎」，可能非開獎頁或格式已變更。"
    # 從「特別獎」出現處往後取整段，依序抓所有 8 碼數字（特別獎、特獎、頭獎×N）
    idx = text.find("特別獎")
    rest = text[idx:]
    nums = re.findall(r"\d{8}", rest)
    if len(nums) < 3:
        return None, "在「特別獎」區塊後找不到足夠的 8 碼數字（需至少特別獎、特獎、頭獎×1）。"
    special_prize = nums[0]
    top_prize = nums[1]
    first_prizes = nums[2:12]  # 頭獎最多取 10 組
    # 期別：114年09-10月 或 114年 11 ~ 12 月
    period_match = re.search(r"(\d{3}年\s*\d{1,2}\s*[~～\-]\s*\d{1,2}\s*月)", text)
    period_label = period_match.group(1).strip() if period_match else ""
    # 領獎期間
    claim_match = re.search(r"領獎期間自\s*(.+?止)", text)
    claim_period_text = (claim_match.group(1).strip() if claim_match else "").replace("\n", "")
    draw = {
        "period_label": period_label or "開獎期",
        "special_prize": special_prize,
        "top_prize": top_prize,
        "first_prizes": first_prizes,
        "extra_six": [],
        "claim_period_text": claim_period_text,
    }
    return draw, None


def fetch_lottery_draw_from_etax(slot):
    """爬蟲：直接取得財政部開獎頁 HTML，並解析開獎號碼。
    slot: 0 = 最新一期（index.html），1 = 上一期（lastNumber.html）。
    回傳 (draw_dict, error_message)。draw_dict 結構同 parse_lottery_text。
    """
    if slot not in (0, 1):
        return None, "僅支援「最新一期」或「上一期」自動取得，其餘請改用手動貼上。"
    url = LOTTERY_ETAX_LATEST if slot == 0 else LOTTERY_ETAX_PREVIOUS
    try:
        r = requests.get(url, timeout=15, headers=LOTTERY_HEADERS)
        r.raise_for_status()
        # 強制以 utf-8 解讀，避免編碼錯亂
        r.encoding = "utf-8"
        text = r.text
    except requests.RequestException as e:
        return None, f"無法取得開獎頁面：{e}"
    return _parse_lottery_draw_from_html(text)


def parse_lottery_text(raw_text):
    """從財政部『統一發票中獎號碼』頁面貼上的文字中解析開獎結果。
    回傳 (draw_dict, error_message)。draw_dict 結構：
      {
        "period_label": "114年 11 ~ 12 月",
        "special_prize": "97023797",
        "top_prize": "00507588",
        "first_prizes": ["92377231", "05232592", "78125249"],
        "extra_six": [],
        "claim_period_text": "115年2月6日起至115年5月5日止"
      }
    若解析失敗，draw_dict 為 None，error_message 為錯誤說明。
    """
    if raw_text is None or not str(raw_text).strip():
        return None, "請先貼上財政部開獎頁面的文字內容。"
    text = str(raw_text)
    # 1) 抓所有 8 碼數字（通常依序為 特別獎、特獎、頭獎...）
    nums = re.findall(r"(\d{8})", text)
    if len(nums) < 3:
        return None, "無法在文字中找到足夠的 8 碼號碼，請確認是否貼上正確頁面內容。"
    special_prize = nums[0]
    top_prize = nums[1] if len(nums) >= 2 else ""
    first_prizes = nums[2:5]  # 至多取三組頭獎，多出的忽略
    # 2) 期別文字（例如：114年 11 ~ 12 月）
    period_match = re.search(r"(\d{3}年\s*\d{1,2}\s*[~～]\s*\d{1,2}\s*月)", text)
    period_label = period_match.group(1).strip() if period_match else ""
    # 3) 領獎期間文字
    claim_match = re.search(r"領獎期間自(.+?止)", text)
    claim_period_text = claim_match.group(1).strip() if claim_match else ""
    draw = {
        "period_label": period_label,
        "special_prize": special_prize,
        "top_prize": top_prize,
        "first_prizes": first_prizes,
        "extra_six": [],
        "claim_period_text": claim_period_text,
    }
    return draw, None


def match_lottery_prize(inv_num8, draw):
    """依台灣統一發票規則回傳中獎結果。
    inv_num8: 8 位數字字串（已正規化）
    draw: parse_lottery_text 回傳的 dict
    回傳 (prize_name, amount)，未中獎則 ("未中獎", 0)
    """
    if not inv_num8 or not draw:
        return "未中獎", 0
    # 特別獎
    if inv_num8 == draw.get("special_prize"):
        return "特別獎", 10_000_000
    # 特獎
    if inv_num8 == draw.get("top_prize"):
        return "特獎", 2_000_000
    first_list = draw.get("first_prizes") or []
    # 頭獎
    for f in first_list:
        if inv_num8 == f:
            return "頭獎", 200_000
    # 二～五獎（末 7/6/5/4 碼比對任一頭獎）
    for f in first_list:
        if len(f) != 8:
            continue
        if inv_num8[-7:] == f[-7:]:
            return "二獎", 40_000
        if inv_num8[-6:] == f[-6:]:
            return "三獎", 10_000
        if inv_num8[-5:] == f[-5:]:
            return "四獎", 4_000
        if inv_num8[-4:] == f[-4:]:
            return "五獎", 1_000
    # 六獎：末 3 碼比對任一頭獎或增開六獎
    for f in first_list:
        if len(f) == 8 and inv_num8[-3:] == f[-3:]:
            return "六獎", 200
    extra_six = draw.get("extra_six") or []
    if inv_num8[-3:] in extra_six:
        return "六獎", 200
    return "未中獎", 0


def save_edited_data(ed_df, original_df, user_email=None):
    """自動保存編輯後的數據；含 modified_at 更新與統編驗證提示。回傳 (saved_count, errors, warnings)。"""
    saved_count = 0
    errors = []
    warnings = []
    
    # 將列名映射回數據庫字段名（含稅率類型，供 0%/免稅 編輯）
    reverse_mapping = {"檔案名稱":"file_name","日期":"date","發票號碼":"invoice_number",
                      "賣方名稱":"seller_name","賣方統編":"seller_ubn","銷售額":"subtotal",
                      "稅額":"tax","總計":"total","類型":"category","會計科目":"subject","狀態":"status","備註":"note",
                      "稅率類型":"tax_type"}
    
    for idx, row in ed_df.iterrows():
        if 'id' not in row or pd.isna(row['id']):
            continue
        
        record_id = int(row['id'])
        
        # 檢查是否有變更
        if idx < len(original_df):
            orig_row = original_df.iloc[idx]
            if orig_row.get('id') == record_id:
                # 比較關鍵字段是否有變化
                changed = False
                for col in ed_df.columns:
                    if col not in ['id', '選取']:
                        orig_val = orig_row.get(col, '')
                        new_val = row.get(col, '')
                        if str(orig_val) != str(new_val):
                            changed = True
                            break
                
                if not changed:
                    continue
        
        # 準備更新數據
        update_data = {}
        for display_col, db_col in reverse_mapping.items():
            if display_col in row:
                update_data[db_col] = row[display_col]
        
        # 稅率類型為空時預設 5%
        if "tax_type" in update_data and (update_data["tax_type"] is None or str(update_data.get("tax_type", "")).strip() == ""):
            update_data["tax_type"] = "5%"
        # 統編驗證（僅提示，不阻擋儲存）
        if "seller_ubn" in update_data and update_data["seller_ubn"]:
            ok_ubn, msg_ubn = validate_ubn(update_data["seller_ubn"])
            if not ok_ubn:
                warnings.append(f"記錄 ID {record_id} 賣方統編：{msg_ubn}（已儲存，僅供參考）")
        
        # 處理數值字段
        for num_col in ['subtotal', 'tax', 'total']:
            if num_col in update_data:
                try:
                    val = str(update_data[num_col]).replace(',', '').replace('$', '')
                    update_data[num_col] = float(val) if val else 0.0
                except:
                    update_data[num_col] = 0.0
        
        # 審計：每次寫回時更新 modified_at
        update_data["modified_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 保存到數據庫或內存
        try:
            if st.session_state.use_memory_mode:
                # 更新內存中的記錄
                for i, inv in enumerate(st.session_state.local_invoices):
                    if inv.get('id') == record_id:
                        for key, val in update_data.items():
                            st.session_state.local_invoices[i][key] = val
                        saved_count += 1
                        break
            else:
                # 更新數據庫（多用戶版本：使用 user_email）
                set_clause = ", ".join([f"{k} = ?" for k in update_data.keys()])
                user_email = st.session_state.get('user_email', 'default_user')
                query = f"UPDATE invoices SET {set_clause} WHERE id = ? AND user_email = ?"
                params = list(update_data.values()) + [record_id, user_email]
                result = run_query(query, tuple(params), is_select=False)
                if result:
                    saved_count += 1
                else:
                    errors.append(f"記錄 ID {record_id} 更新失敗")
        except Exception as e:
            errors.append(f"記錄 ID {record_id} 更新錯誤: {str(e)}")
    
    return saved_count, errors, warnings

def process_ocr(image_obj, file_name, model_name, api_key_val):
    try:
        if image_obj.mode != "RGB": image_obj = image_obj.convert("RGB")
        # 稍微降低解析度以加快速度並減少 Token，但保持足夠清晰度
        image_obj.thumbnail((1600, 1600), Image.Resampling.LANCZOS)
        img_byte = io.BytesIO(); image_obj.save(img_byte, format="JPEG", quality=85)
        img_base64 = base64.b64encode(img_byte.getvalue()).decode()
        
        # 優化 Prompt：明確要求純 JSON，不要 Markdown
        prompt = """You are a receipt OCR assistant. Extract data from this image.
        Output ONLY a valid JSON object. Do NOT use Markdown code blocks.
        Fields required:
        - date (Format: YYYY/MM/DD, convert ROC year to AD if needed)
        - invoice_no (Invoice number)
        - seller_name (Store name)
        - seller_ubn (Unified Business Number / Tax ID)
        - subtotal (Amount before tax, number only)
        - tax (Tax amount, number only)
        - total (Total amount, number only)
        - type (Invoice type, e.g., "電子發票", "收據")
        - category_suggest (Category, e.g., "餐飲", "交通", "辦公用品")
        
        If a field is missing, use null or 0.
        """
        
        payload = {
            "contents": [{"parts": [{"text": prompt}, {"inline_data": {"mime_type": "image/jpeg", "data": img_base64}}]}], 
            "generationConfig": {"temperature": 0.1, "maxOutputTokens": 1024}
        }
        
        # 優先順序調整：先試 v1beta (支援較新模型)，再試 v1
        configs = [
            ("v1beta", model_name),
            ("v1beta", f"models/{model_name}"),
            ("v1", model_name),
            ("v1", f"models/{model_name}")
        ]
        
        session = requests.Session(); session.trust_env = False
        os.environ.pop('HTTP_PROXY', None); os.environ.pop('HTTPS_PROXY', None)
        
        # 修復 Bug #5: 添加重試機制
        try:
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry
            
            # 配置重試策略
            retry_strategy = Retry(
                total=3,
                backoff_factor=1,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["POST"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            session.mount("http://", adapter)
            session.mount("https://", adapter)
        except ImportError:
            # 如果 urllib3 版本不支持，跳過重試配置
            pass
        
        last_err = ""
        debug_info = []
        
        for ver, m_name in configs:
            # 智能處理 models/ 前綴
            final_model_name = m_name if "models/" in m_name else f"models/{m_name}"
            # 修正 URL 結構：v1beta/models/gemini-pro:generateContent
            # 移除多餘的 models/ 如果 API 版本路徑已經隱含
            if ver == "v1beta" or ver == "v1":
                 # Google API 規範: https://generativelanguage.googleapis.com/v1beta/models/gemini-pro:generateContent
                 # 如果 m_name 已經包含 models/，則直接使用
                 pass

            url = f"https://generativelanguage.googleapis.com/{ver}/{m_name}:generateContent?key={api_key_val}"
            if "models/" not in m_name:
                 url = f"https://generativelanguage.googleapis.com/{ver}/models/{m_name}:generateContent?key={api_key_val}"
            
            try:
                # 修復 Bug #5: 使用帶重試的請求
                resp = session.post(url, json=payload, timeout=25)
                if resp.status_code == 200:
                    try:
                        resp_json = resp.json()
                        if 'candidates' not in resp_json or not resp_json['candidates']:
                            last_err = "API 回傳結構異常 (無 candidates)"
                            continue
                            
                        text = resp_json['candidates'][0]['content']['parts'][0]['text']
                        raw = extract_json(text)
                        
                        if raw:
                            data = {
                                "file_name": file_name,
                                "date": raw.get("date") or raw.get("日期") or datetime.now().strftime("%Y/%m/%d"),
                                "invoice_no": raw.get("invoice_no") or raw.get("invoice_number") or "N/A",
                                "seller_name": raw.get("seller_name") or "N/A",
                                "seller_ubn": raw.get("seller_ubn") or "N/A",
                                "subtotal": raw.get("subtotal") or 0, "tax": raw.get("tax") or 0, "total": raw.get("total") or 0,
                                "type": raw.get("type") or "其他", "category_suggest": raw.get("category_suggest") or "雜項"
                            }
                            data["status"] = "✅ 正常" if data["total"] else "⚠️ 缺漏"
                            return data, None
                        else:
                            last_err = f"JSON 解析失敗. 原始文本: {text[:100]}..."
                            debug_info.append(f"{ver}/{m_name}: {last_err}")
                    except Exception as parse_err:
                        last_err = f"解析異常: {str(parse_err)}"
                        debug_info.append(f"{ver}/{m_name}: {last_err}")
                else:
                    last_err = f"HTTP {resp.status_code}: {resp.text[:100]}"
                    debug_info.append(f"{ver}/{m_name}: {last_err}")
            except requests.exceptions.RequestException as e:
                # 修復 Bug #5: 區分網絡錯誤和其他錯誤
                last_err = f"網絡錯誤: {str(e)}"
                debug_info.append(f"{ver}/{m_name}: {last_err}")
                continue
            except Exception as e: 
                last_err = str(e)
                debug_info.append(f"{ver}/{m_name}: {last_err}")
                continue
                
        return None, f"所有嘗試皆失敗。最後錯誤: {last_err} | 歷程: {'; '.join(debug_info)}"
    except Exception as e: return None, f"系統錯誤: {str(e)}"


# --- AI 報帳小助理：對話與自然語言記帳 ---
ASSISTANT_SYSTEM_PROMPT = """你是「發票報帳小秘笈」的 AI 報帳小助理，使用繁體中文回答。
你會回答關於發票報帳、會計科目、本系統操作的簡單問題。
本系統會計科目範例：餐飲費、交通費、辦公用品、差旅費、雜項；類型範例：餐飲、交通、辦公用品、其他。
若用戶用一句話描述一筆支出（例如「今天午餐 120 元 全家」「昨天計程車 200」），請先簡短回覆確認，然後在回覆「最後一行」單獨寫 [EXPENSE] 並換行，下一行只放一個 JSON 物件，欄位：date(YYYY/MM/DD), seller_name, total(數字), category(類型), subject(會計科目)。若明顯只是問問題而非記帳則不要加 [EXPENSE]。"""


def call_gemini_chat(messages, api_key_val, model_name, system_instruction=None):
    """呼叫 Gemini 多輪對話 API（純文字），回傳 (reply_text, error)。"""
    if not api_key_val or not messages:
        return None, "缺少 API Key 或訊息"
    try:
        contents = []
        for m in messages:
            role = (m.get("role") or "user").strip().lower()
            if role == "model" or role == "assistant":
                role = "model"
            else:
                role = "user"
            text = (m.get("content") or "").strip()
            if not text:
                continue
            contents.append({"role": role, "parts": [{"text": text}]})
        if not contents:
            return None, "無有效訊息"
        payload = {
            "contents": contents,
            "generationConfig": {"temperature": 0.3, "maxOutputTokens": 1024},
        }
        if system_instruction:
            payload["systemInstruction"] = {"parts": [{"text": system_instruction}]}
        model_id = model_name if "models/" in model_name else f"models/{model_name}"
        url = f"https://generativelanguage.googleapis.com/v1beta/{model_id}:generateContent?key={api_key_val}"
        session = requests.Session()
        session.trust_env = False
        resp = session.post(url, json=payload, timeout=30)
        if resp.status_code != 200:
            return None, f"API 錯誤: {resp.status_code} {resp.text[:200]}"
        data = resp.json()
        if not data.get("candidates") or not data["candidates"][0].get("content", {}).get("parts"):
            return None, "API 回傳無內容"
        text = data["candidates"][0]["content"]["parts"][0].get("text", "").strip()
        return text, None
    except requests.exceptions.RequestException as e:
        return None, f"網路錯誤: {str(e)}"
    except Exception as e:
        return None, f"錯誤: {str(e)}"


def parse_expense_from_assistant_reply(reply_text):
    """從助理回覆中解析 [EXPENSE] 後的 JSON，回傳 dict 或 None。"""
    if not reply_text or "[EXPENSE]" not in reply_text:
        return None
    try:
        start = reply_text.find("[EXPENSE]") + len("[EXPENSE]")
        rest = reply_text[start:].strip()
        obj = extract_json(rest)
        if not obj:
            return None
        # 正規化欄位：date, seller_name, total, category, subject
        date_val = obj.get("date") or obj.get("日期") or datetime.now().strftime("%Y/%m/%d")
        seller = (obj.get("seller_name") or obj.get("賣方") or obj.get("store") or "未知").strip()
        total_val = obj.get("total") or obj.get("總計") or obj.get("金額") or 0
        try:
            total_val = float(total_val)
        except (TypeError, ValueError):
            total_val = 0
        category = (obj.get("category") or obj.get("類型") or "其他").strip()
        subject = (obj.get("subject") or obj.get("會計科目") or "雜項").strip()
        return {
            "date": date_val,
            "seller_name": seller,
            "total": total_val,
            "category": category,
            "subject": subject,
        }
    except Exception:
        return None


def insert_assistant_draft(draft, user_email):
    """將 AI 助理解析的一筆草稿寫入 invoices 表。回傳 (success, error_message)。"""
    def safe_str(v, d=""):
        if v is None or (isinstance(v, str) and not v.strip()): return d
        return str(v).strip()[:500]
    def safe_float(v):
        try: return float(v)
        except: return 0.0
    try:
        q = "INSERT INTO invoices (user_email, file_name, date, invoice_number, seller_name, seller_ubn, subtotal, tax, total, category, subject, status, note, tax_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
        total = safe_float(draft.get("total", 0))
        tax = round(total / 1.05 * 0.05, 2) if total else 0
        subtotal = round(total - tax, 2)
        params = (
            user_email,
            "AI助理新增",
            safe_str(draft.get("date"), datetime.now().strftime("%Y/%m/%d")),
            "AI-" + datetime.now().strftime("%Y%m%d%H%M%S"),
            safe_str(draft.get("seller_name"), "未知"),
            "",
            subtotal,
            tax,
            total,
            safe_str(draft.get("category"), "其他"),
            safe_str(draft.get("subject"), "雜項"),
            "✅ 正常",
            "由 AI 報帳小助理新增",
            "5%",
        )
        result = run_query(q, params, is_select=False)
        return bool(result), None if result else "寫入失敗"
    except Exception as e:
        return False, str(e)


# --- 4. 介面渲染 ---
# 這裡不再硬編碼 Key，防止洩漏。預設為空，強迫使用 Secrets 或手動輸入。
DEFAULT_KEY = "" 

# --- 主應用入口：檢查登入狀態（多用戶版本）---
# OAuth callback：若 URL 帶 code 與 state，先處理第三方登入
qp = st.query_params
if qp.get("code") and qp.get("state"):
    state = qp.get("state", "")
    provider = state.split(":", 1)[0] if ":" in state else ""
    success, msg = False, "登入失敗"
    if provider == "google":
        success, msg = handle_oauth_callback_google(qp["code"], state)
    elif provider == "line":
        success, msg = handle_oauth_callback_line(qp["code"], state)
    elif provider == "facebook":
        success, msg = handle_oauth_callback_facebook(qp["code"], state)
    if success:
        st.session_state.authenticated = True
        st.session_state.user_email = msg
        st.session_state.login_at = datetime.now().isoformat()
        st.session_state.pop("login_csrf_token", None)
        st.session_state.pop("oauth_state", None)
        try:
            st.query_params.clear()
        except Exception:
            pass
        st.rerun()
    else:
        st.error(f"❌ {msg}")
        st.info("請關閉此頁或返回登入頁重試。")
        st.stop()

# AUTH-04：Session 過期（預設 24 小時）
_SESSION_EXPIRE_HOURS = 24
if st.session_state.authenticated and st.session_state.user_email and st.session_state.get("login_at"):
    try:
        login_at = datetime.fromisoformat(st.session_state.login_at)
        if (datetime.now() - login_at).total_seconds() > _SESSION_EXPIRE_HOURS * 3600:
            st.session_state.authenticated = False
            st.session_state.user_email = None
            st.session_state.login_at = None
    except Exception:
        st.session_state.login_at = None

if not st.session_state.authenticated or not st.session_state.user_email:
    login_page()
    st.stop()  # 未登入時停止執行後續代碼

# 已登入，顯示側邊欄（還原先前樣式：標題 + 選單 + 用戶 + 登出 + 進階設定）
with st.sidebar:
    st.title("🛠️ 小工具")
    tool_options = [
        ("invoice", "📑 發票報帳小秘笈"),
        ("contract", "⚖️ AI 合約比對"),
        ("customer_service", "📧 AI 客服小秘"),
        ("meeting", "📅 AI 會議精華"),
    ]
    current = st.session_state.current_tool
    idx = next((i for i, (k, _) in enumerate(tool_options) if k == current), 0)
    choice = st.radio(
        "選擇工具",
        options=[label for _, label in tool_options],
        index=idx,
        key="sidebar_tool_radio",
        label_visibility="collapsed",
    )
    st.session_state.current_tool = next(k for k, label in tool_options if label == choice)
    
    st.markdown("---")
    user_email = st.session_state.get("user_email", "未登入")
    st.caption(f"👤 {user_email}")
    if st.button("🚪 登出", use_container_width=True, key="sidebar_logout"):
        st.session_state.authenticated = False
        st.session_state.user_email = None
        st.session_state.login_at = None
        st.rerun()
    
    st.markdown("---")
    with st.expander("⚙️ 進階設定", expanded=False):
        model = st.selectbox(
            "辨識模型",
            ["gemini-2.0-flash", "gemini-2.5-flash", "gemini-1.5-flash", "gemini-1.5-pro"],
            key="sidebar_model",
        )
        st.session_state.gemini_api_key = _safe_secrets_get("GEMINI_API_KEY")
        st.session_state.gemini_model = model
    
    if st.session_state.current_tool == "invoice":
        with st.expander("🧪 測試數據", expanded=False):
            if st.button("📋 載入測試數據", use_container_width=True, key="load_test_data_btn"):
                n, err = load_test_data(st.session_state.get("user_email", "default_user"))
                if err:
                    st.error(f"載入失敗：{err}")
                else:
                    st.success(f"已載入 {n} 筆測試發票（2 組 Batch：OCR 4 張 + 導入 3 張，未分組 2 張）")
                    time.sleep(0.5)
                    st.rerun()
    
    st.session_state.use_memory_mode = False

# 依所選小工具顯示主內容
if st.session_state.current_tool != "invoice":
    _tool = st.session_state.current_tool
    api_key = st.session_state.get("gemini_api_key") or _safe_secrets_get("GEMINI_API_KEY")
    model = st.session_state.get("gemini_model") or "gemini-2.0-flash"
    
    # --- 📅 AI 會議精華（最小可用版）---
    if _tool == "meeting":
        st.subheader("📅 AI 會議精華")
        st.caption("貼上會議逐字稿或紀錄，由 AI 產出結論與待辦事項。")
        if not api_key:
            st.warning("此功能需要 API 金鑰，請聯絡管理員設定。")
            st.stop()
        transcript = st.text_area("會議逐字稿或紀錄", height=200, placeholder="貼上會議內容…", key="meeting_transcript")
        if st.button("產出會議精華", type="primary", key="meeting_btn"):
            if not (transcript and transcript.strip()):
                st.error("請先貼上會議內容。")
            else:
                with st.spinner("正在產出精華…"):
                    sys_inst = "你是會議紀錄助手。根據使用者提供的會議逐字稿或紀錄，用繁體中文產出：1) 會議結論（簡短條列）；2) 待辦事項（誰／做什麼／期限若有的話）。結構清晰、條列式。"
                    reply, err = call_gemini_chat(
                        [{"role": "user", "content": transcript.strip()[:15000]}],
                        api_key, model, system_instruction=sys_inst,
                    )
                if err:
                    st.error(err)
                else:
                    st.success("已產出")
                    st.markdown(reply or "")
        st.stop()
    
    # --- ⚖️ AI 合約比對（最小可用版）---
    if _tool == "contract":
        st.subheader("⚖️ AI 合約比對")
        st.caption("貼上兩份合約或條款內容，由 AI 標示差異與重點。")
        if not api_key:
            st.warning("此功能需要 API 金鑰，請聯絡管理員設定。")
            st.stop()
        c1, c2 = st.columns(2)
        with c1:
            text_a = st.text_area("合約／條款 A", height=180, placeholder="貼上第一份內容…", key="contract_a")
        with c2:
            text_b = st.text_area("合約／條款 B", height=180, placeholder="貼上第二份內容…", key="contract_b")
        if st.button("開始比對", type="primary", key="contract_btn"):
            if not (text_a and text_b and text_a.strip() and text_b.strip()):
                st.error("請在 A、B 兩欄都貼上內容。")
            else:
                with st.spinner("正在比對…"):
                    sys_inst = "你是合約比對助手。根據使用者提供的兩份合約或條款（以 [A] 與 [B] 標示），用繁體中文產出：1) 主要差異（條列）；2) 需注意的條款或風險提示。簡潔明確。"
                    content = f"[A]\n{text_a.strip()[:8000]}\n\n[B]\n{text_b.strip()[:8000]}"
                    reply, err = call_gemini_chat(
                        [{"role": "user", "content": content}],
                        api_key, model, system_instruction=sys_inst,
                    )
                if err:
                    st.error(err)
                else:
                    st.success("比對結果")
                    st.markdown(reply or "")
        st.stop()
    
    # --- 📧 AI 客服小秘：維持佔位 ---
    st.subheader({"customer_service": "📧 AI 客服小秘"}.get(_tool, "小工具"))
    st.info("🛠️ 此工具即將推出，敬請期待。")
    st.caption("您可先使用「📑 發票報帳小秘笈」「📅 AI 會議精華」或「⚖️ AI 合約比對」。")
    st.stop()

# --- 發票報帳小秘笈主內容 ---
# 發票 OCR 與 AI 小助理需使用 Gemini API，在此統一取得金鑰與模型
api_key = st.session_state.get("gemini_api_key") or _safe_secrets_get("GEMINI_API_KEY")
model = st.session_state.get("gemini_model") or "gemini-2.0-flash"

# --- Hero：單一標題 + 副標 + 主操作入口 ---
with st.container():
    title_col1, title_col2 = st.columns([2.5, 1.5])
    with title_col1:
        st.title("發票報帳")
        st.caption("上傳辨識、導入、報表導出")
    with title_col2:
        st.write("")
        btn_row1, btn_row2, btn_row3 = st.columns(3)
        with btn_row1:
            if st.button("📷 上傳發票圖", type="primary", use_container_width=True):
                st.session_state.show_upload_dialog = True
                st.session_state.upload_mode = "ocr"
        with btn_row2:
            if st.button("📥 CSV／Excel 導入", type="primary", use_container_width=True):
                st.session_state.show_upload_dialog = True
                st.session_state.upload_mode = "import"
        with btn_row3:
            if st.button("🤖 AI 報帳小助理", type="secondary", use_container_width=True):
                st.session_state.show_assistant_dialog = True
# 查詢當前用戶的數據（多用戶版本：使用 user_email）
user_email = st.session_state.get('user_email', 'default_user')
df_raw = run_query("SELECT * FROM invoices WHERE user_email = ? ORDER BY id DESC", (user_email,))

st.markdown("---")
# ========== 1. 統計指標區（報表標題 + KPI）==========
with st.container():
    df_stats = df_raw.copy()
    if not df_stats.empty:
        # 先重命名列以便統計報表使用
        mapping = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編","subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間"}
        df_stats = df_stats.rename(columns=mapping)
        
        if "總計" in df_stats.columns:
            for c in ["總計", "稅額"]: 
                if c in df_stats.columns:
                    df_stats[c] = pd.to_numeric(df_stats[c].astype(str).str.replace(',',''), errors='coerce').fillna(0)
            
            # 計算「本月」數據
            today = datetime.now().date()
            month_start = today.replace(day=1)
            
            # 篩選本月的發票（支援多種日期格式：YYYY/MM/DD、YYYY-MM-DD、含時間等）
            if "日期" in df_stats.columns:
                try:
                    # 先不指定 format，讓 pandas 自動推斷多種格式
                    df_stats['日期_parsed'] = pd.to_datetime(df_stats['日期'], errors='coerce')
                    if df_stats['日期_parsed'].isna().all():
                        for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M:%S'):
                            df_stats['日期_parsed'] = pd.to_datetime(df_stats['日期'], errors='coerce', format=fmt)
                            if not df_stats['日期_parsed'].isna().all():
                                break
                    df_month = df_stats[df_stats['日期_parsed'].notna() & (df_stats['日期_parsed'].dt.date >= month_start)].copy()
                except Exception:
                    month_str = today.strftime("%Y/%m")
                    df_month = df_stats[df_stats['日期'].astype(str).str.contains(month_str, na=False)].copy()
            else:
                df_month = df_stats.copy()
            
            # 本月無發票時改顯示「全部」統計，避免 KPI 全為 0
            if df_month.empty and not df_stats.empty:
                df_month = df_stats.copy()
                _kpi_use_all = True
            else:
                _kpi_use_all = False
            
            # 計算統計數據（本月或全部）
            month_total = pd.to_numeric(df_month['總計'], errors='coerce').fillna(0).sum() if not df_month.empty else 0
            month_tax = pd.to_numeric(df_month['稅額'], errors='coerce').fillna(0).sum() if not df_month.empty and '稅額' in df_month.columns else 0
            month_invoice_count = len(df_month) if not df_month.empty else 0
            month_missing_count = len(df_month[df_month['狀態'].astype(str).str.contains('缺失', na=False)]) if not df_month.empty and '狀態' in df_month.columns else 0
            
            # 本月無發票時改顯示「全部」統計（在計算 df_month 後已設定 _kpi_use_all）
            kpi_pill = "全部" if _kpi_use_all else "本月份"
            # 報表標題區（參考 Planetaria：左 標題+說明，右 pill）
            st.markdown(
                '<div class="report-header">'
                '<div class="report-header-left">'
                '<p class="report-header-title"><span class="report-header-dot"></span> 發票報帳</p>'
                '<p class="report-header-desc">來自上傳與導入的發票明細</p>'
                '</div>'
                f'<div class="report-header-right"><span class="report-pill">{kpi_pill}</span></div>'
                '</div>',
                unsafe_allow_html=True,
            )
            # 四個 KPI 卡片（標籤在上、大數字在下，無邊框卡片）
            stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
            with stat_col1:
                st.markdown(f'<div class="kpi-card"><span class="kpi-label">本月總計</span><span class="kpi-value">${month_total:,.0f}</span></div>', unsafe_allow_html=True)
            with stat_col2:
                st.markdown(f'<div class="kpi-card"><span class="kpi-label">預計稅額</span><span class="kpi-value">${month_tax:,.0f}</span></div>', unsafe_allow_html=True)
            with stat_col3:
                st.markdown(f'<div class="kpi-card"><span class="kpi-label">發票總數</span><span class="kpi-value">{month_invoice_count:,} 筆</span></div>', unsafe_allow_html=True)
            with stat_col4:
                st.markdown(f'<div class="kpi-card"><span class="kpi-label">缺失件數</span><span class="kpi-value">{month_missing_count:,} 筆</span></div>', unsafe_allow_html=True)
            if _kpi_use_all:
                st.caption("本月尚無發票，以上為**全部**數據。")
            elif month_invoice_count == 0:
                st.caption("尚無本月發票，請先上傳或導入。")
    else:
        # 無數據時：報表標題 + 空 KPI 卡片
        st.markdown(
            '<div class="report-header">'
            '<div class="report-header-left">'
            '<p class="report-header-title"><span class="report-header-dot"></span> 發票報帳</p>'
            '<p class="report-header-desc">來自上傳與導入的發票明細</p>'
            '</div>'
            '<div class="report-header-right"><span class="report-pill">本月份</span></div>'
            '</div>',
            unsafe_allow_html=True,
        )
        stat_col1, stat_col2, stat_col3, stat_col4 = st.columns(4)
        with stat_col1:
            st.markdown('<div class="kpi-card"><span class="kpi-label">本月總計</span><span class="kpi-value">$0</span></div>', unsafe_allow_html=True)
        with stat_col2:
            st.markdown('<div class="kpi-card"><span class="kpi-label">預計稅額</span><span class="kpi-value">$0</span></div>', unsafe_allow_html=True)
        with stat_col3:
            st.markdown('<div class="kpi-card"><span class="kpi-label">發票總數</span><span class="kpi-value">0 筆</span></div>', unsafe_allow_html=True)
        with stat_col4:
            st.markdown('<div class="kpi-card"><span class="kpi-label">缺失件數</span><span class="kpi-value">0 筆</span></div>', unsafe_allow_html=True)

# ========== 發票對獎（簡化版）==========
def _run_lottery_and_match(slot):
    """取得開獎號碼並對所有發票對獎，結果寫入 session_state.lottery_last_checked。回傳 (error_message or None)。"""
    draw, err = fetch_lottery_draw_from_etax(slot)
    if err:
        return err
    st.session_state["lottery_draw"] = draw
    winners = []
    total_prize = 0
    checked_count = 0
    for _, row in df_raw.iterrows():
        inv_num8 = normalize_invoice_number(row.get("invoice_number"))
        if not inv_num8:
            continue
        prize, amount = match_lottery_prize(inv_num8, draw)
        checked_count += 1
        if amount > 0:
            winners.append({
                "日期": row.get("date"),
                "發票號碼": row.get("invoice_number"),
                "賣方名稱": row.get("seller_name"),
                "獎別": prize,
                "獎金": amount,
            })
            total_prize += amount
    st.session_state["lottery_last_checked"] = {
        "draw": draw,
        "winners": winners,
        "checked_count": checked_count,
        "total_prize": total_prize,
    }
    return None

# ========== 發票對獎（明顯位置，吸引用戶）==========
with st.container():
    if df_raw.empty:
        st.subheader("🎰 發票對獎")
        st.caption("目前沒有發票資料，請先上傳或導入後再進行對獎。")
    else:
        _auto_err = None
        if not st.session_state.get("lottery_last_checked"):
            with st.spinner("正在自動對獎最新一期…"):
                _auto_err = _run_lottery_and_match(0)
        last = st.session_state.get("lottery_last_checked")
        draw = st.session_state.get("lottery_draw")
        label = (last.get("draw") or draw or {}).get("period_label") if (last or draw) else "本期"
        if not label and draw:
            label = draw.get("period_label") or "本期"
        _claim = (last.get("draw") or draw or {}).get("claim_period_text")
        _winners = (last or {}).get("winners") or []
        st.subheader("🎰 發票對獎")
        if label or _claim:
            st.caption(f"{label}" + (f" · 領獎至 {_claim}" if _claim else ""))
        st.markdown('<style>div[data-testid="column"] { align-items: stretch; }</style>', unsafe_allow_html=True)
        _left, _right = st.columns([1, 1])
        with _left:
            st.markdown("**對獎結果**")
            if _auto_err and not last:
                st.warning("無法取得開獎，請使用下方手動貼上或點「對獎」重試。")
            elif last:
                checked_count = last.get("checked_count", 0)
                total_prize = last.get("total_prize", 0)
                if _winners:
                    st.success(f"對獎 {checked_count} 張，中獎 **{len(_winners)}** 張，共 **{total_prize:,}** 元。")
                else:
                    st.info(f"對獎 {checked_count} 張，未中獎。")
            else:
                st.caption("尚未對獎")
            if last and _winners:
                with st.expander("中獎明細", expanded=False):
                    st.dataframe(pd.DataFrame(last["winners"]), use_container_width=True, hide_index=True)
            st.markdown("**操作**")
            _b1, _b2 = st.columns(2)
            with _b1:
                if st.button("對獎（本期）", type="secondary", use_container_width=True, key="lottery_btn_latest"):
                    with st.spinner("對獎中…"):
                        err = _run_lottery_and_match(0)
                    if err:
                        st.error(err)
                    else:
                        st.rerun()
            with _b2:
                if st.button("對獎（上期）", type="secondary", use_container_width=True, key="lottery_btn_prev"):
                    with st.spinner("對獎中…"):
                        err = _run_lottery_and_match(1)
                    if err:
                        st.error(err)
                    else:
                        st.rerun()
            with st.expander("手動貼上開獎號碼（備用）", expanded=False):
                st.caption("當自動取得失敗或要對更早期別時，可至 [財政部開獎頁](https://invoice.etax.nat.gov.tw/) 複製整頁貼上後解析並對獎。")
                raw_lottery = st.text_area("貼上財政部「統一發票中獎號碼」頁面文字", value=st.session_state.get("lottery_raw_text", ""), height=100, key="lottery_raw_text")
                if st.button("解析並對獎", key="lottery_parse_and_match_btn"):
                    draw, err = parse_lottery_text(raw_lottery)
                    if err:
                        st.error(err)
                    else:
                        st.session_state["lottery_draw"] = draw
                        winners = []
                        total_prize = 0
                        checked_count = 0
                        for _, row in df_raw.iterrows():
                            inv_num8 = normalize_invoice_number(row.get("invoice_number"))
                            if not inv_num8:
                                continue
                            prize, amount = match_lottery_prize(inv_num8, draw)
                            checked_count += 1
                            if amount > 0:
                                winners.append({"日期": row.get("date"), "發票號碼": row.get("invoice_number"), "賣方名稱": row.get("seller_name"), "獎別": prize, "獎金": amount})
                                total_prize += amount
                        st.session_state["lottery_last_checked"] = {"draw": draw, "winners": winners, "checked_count": checked_count, "total_prize": total_prize}
                        st.rerun()
        with _right:
            st.markdown("**本期開獎號碼**")
            if draw:
                _rows = [("特別獎", draw.get("special_prize") or "—"), ("特獎", draw.get("top_prize") or "—"), ("頭獎", "、".join(draw.get("first_prizes") or []) or "—")]
                st.dataframe(pd.DataFrame(_rows, columns=["獎別", "中獎號碼"]), use_container_width=True, hide_index=True)
                if draw.get("claim_period_text"):
                    st.caption(f"領獎期間 {draw['claim_period_text']}")
            else:
                st.caption("尚無開獎資料，請點左側「對獎」取得。")

# 初始化 dialog 狀態
if "show_upload_dialog" not in st.session_state:
    st.session_state.show_upload_dialog = False
# AI 報帳小助理：對話紀錄、是否顯示對話框、待確認草稿
if "assistant_chat_history" not in st.session_state:
    st.session_state.assistant_chat_history = []
if "show_assistant_dialog" not in st.session_state:
    st.session_state.show_assistant_dialog = False
if "assistant_pending_draft" not in st.session_state:
    st.session_state.assistant_pending_draft = None

# 上傳對話框函數
@st.dialog("📤 上傳辨識", width="medium")
def upload_dialog():
    # 根據模式顯示不同內容
    upload_mode = st.session_state.get("upload_mode", "ocr")
    
    if upload_mode == "ocr":
        # OCR識別區域
        st.markdown("### 📷 上傳發票圖")
        if not api_key:
            st.warning("⚠️ 圖片辨識需要 API 金鑰。請在 **Manage app → Settings → Secrets** 中設定 `GEMINI_API_KEY`，設定後重新載入頁面。")
        st.caption("支援 JPG、PNG；建議單張清晰、光線充足，以利辨識。")
        files = st.file_uploader("批次選擇照片", type=["jpg","png","jpeg"], accept_multiple_files=True)
        if files:
            st.caption(f"已選擇 {len(files)} 個文件")
        
        if files and st.button("開始辨識 🚀", type="primary", use_container_width=True):
            # 先將檔案內容讀入 session state，避免 rerun 後 Streamlit 清除上傳檔案導致無法辨識
            try:
                st.session_state.upload_file_data = [(f.name, f.getvalue()) for f in files]
                st.session_state.start_ocr = True
                st.rerun()
            except Exception as e:
                st.error(f"讀取檔案失敗，請重試: {e}")
    else:
        # 數據導入區域
        st.markdown("### 📥 CSV／Excel 數據導入")
        st.info("💡 支援 Excel (.xlsx) 或 CSV (.csv)；必填欄位：日期、發票號碼、總計。可先下載模板再填寫。")
        
        # 下載導入模板
        template_data = {
            "檔案名稱": ["範例1.jpg", "範例2.jpg"],
            "日期": ["2025/01/01", "2025/01/02"],
            "發票號碼": ["AB12345678", "CD87654321"],
            "賣方名稱": ["範例商店", "範例公司"],
            "賣方統編": ["12345678", "87654321"],
            "銷售額": [1000, 2000],
            "稅額": [50, 100],
            "總計": [1050, 2100],
            "類型": ["餐飲", "交通"],
            "會計科目": ["餐飲費", "交通費"]
        }
        template_df = pd.DataFrame(template_data)
        template_csv = template_df.to_csv(index=False).encode('utf-8-sig')
        st.download_button("📥 下載導入模板 (CSV)", template_csv, "invoice_import_template.csv", 
                         mime="text/csv", use_container_width=True)
        
        uploaded_file = st.file_uploader("選擇要導入的文件", type=["csv", "xlsx"], key="import_file_dialog")
        
        if uploaded_file and st.button("開始導入", type="primary", use_container_width=True, key="import_btn_dialog"):
            st.session_state.import_file = uploaded_file
            st.session_state.start_import = True
            st.rerun()

# 顯示上傳對話框（優先，以便空狀態按鈕可觸發）
if st.session_state.show_upload_dialog:
    upload_dialog()
    st.session_state.show_upload_dialog = False

# 若有待處理的 OCR 或導入，不要停在空狀態，讓下方 OCR/導入區塊執行
_has_pending_ocr = st.session_state.get("start_ocr") and ("upload_file_data" in st.session_state or "upload_files" in st.session_state)
_has_pending_import = st.session_state.get("start_import") and "import_file" in st.session_state

# 發票模組：尚無資料時顯示空狀態與操作引導（有待處理 OCR/導入時不停止，讓辨識先跑）
if df_raw.empty and not _has_pending_ocr and not _has_pending_import:
    st.markdown("---")
    st.subheader("📋 發票明細")
    st.info("尚無發票資料，請先上傳或導入。完成後即可在此查看總覽、編輯與導出報表。")
    ec1, ec2 = st.columns(2)
    with ec1:
        if st.button("📷 上傳發票圖", type="primary", use_container_width=True, key="empty_upload_ocr"):
            st.session_state.show_upload_dialog = True
            st.session_state.upload_mode = "ocr"
            st.rerun()
    with ec2:
        if st.button("📥 CSV／Excel 導入", type="primary", use_container_width=True, key="empty_upload_import"):
            st.session_state.show_upload_dialog = True
            st.session_state.upload_mode = "import"
            st.rerun()
    st.caption("支援發票照片 AI 辨識或批次匯入既有資料。")
    st.stop()

# 設定公司資訊（用於 PDF 導出）
with st.expander("📋 設定公司資訊（用於 PDF 導出）", expanded=False):
    cn = st.text_input("報支公司名稱", value=st.session_state.get("company_name", ""), key="company_name_input", placeholder="例：○○有限公司")
    ub = st.text_input("公司統編", value=st.session_state.get("company_ubn", ""), key="company_ubn_input", placeholder="8 碼數字")
    st.session_state.company_name = cn if cn is not None else st.session_state.get("company_name", "")
    st.session_state.company_ubn = ub if ub is not None else st.session_state.get("company_ubn", "")
    if ub and ub.strip():
        ok_ubn, msg_ubn = validate_ubn(ub)
        if not ok_ubn:
            st.caption(f"⚠️ {msg_ubn}（僅供參考，不影響導出）")
    st.caption("導出 PDF 時會顯示於報表上方；可不填。")

# AI 報帳小助理對話框
@st.dialog("🤖 AI 報帳小助理", width="large")
def assistant_dialog():
    api_key = st.session_state.get("gemini_api_key") or ""
    model = st.session_state.get("gemini_model") or "gemini-2.0-flash"
    history = st.session_state.assistant_chat_history
    pending = st.session_state.assistant_pending_draft

    st.caption("可問報帳、會計科目或系統操作；也可用一句話記一筆支出，例如：「今天午餐 120 元 全家」")
    if not api_key:
        st.warning("此功能需要 API 金鑰，請聯絡管理員設定。")
        return
    if history and st.button("🗑️ 清除對話", key="assistant_clear_chat"):
        st.session_state.assistant_chat_history = []
        st.session_state.assistant_pending_draft = None
        st.rerun()
    st.divider()

    # 待確認草稿卡片
    if pending:
        with st.container():
            st.markdown("**📋 是否新增以下報帳？**")
            st.json(pending)
            c1, c2 = st.columns(2)
            with c1:
                if st.button("✅ 確認新增", type="primary", key="assistant_confirm_draft"):
                    user_email = st.session_state.get("user_email", "default_user")
                    ok, err = insert_assistant_draft(pending, user_email)
                    if ok:
                        st.success("已新增一筆報帳。")
                        st.session_state.assistant_pending_draft = None
                    else:
                        st.error(err or "新增失敗")
                    st.rerun()
            with c2:
                if st.button("❌ 取消", key="assistant_cancel_draft"):
                    st.session_state.assistant_pending_draft = None
                    st.rerun()
        st.divider()

    # 對話紀錄
    for msg in history:
        role = "user" if msg.get("role") == "user" else "assistant"
        with st.chat_message(role):
            st.write(msg.get("content", ""))

    # 聊天輸入：送出後呼叫 Gemini 並解析 [EXPENSE]
    if prompt := st.chat_input("輸入問題或記一筆支出…"):
        history.append({"role": "user", "content": prompt})
        messages = [{"role": m["role"], "content": m["content"]} for m in history]
        reply, err = call_gemini_chat(messages, api_key, model, ASSISTANT_SYSTEM_PROMPT)
        if err:
            history.append({"role": "model", "content": f"⚠️ 發生錯誤：{err}"})
        else:
            expense = parse_expense_from_assistant_reply(reply)
            if expense:
                st.session_state.assistant_pending_draft = expense
                display_reply = reply.split("[EXPENSE]")[0].strip() if "[EXPENSE]" in reply else reply
            else:
                display_reply = reply
            history.append({"role": "model", "content": display_reply})
        st.session_state.assistant_chat_history = history
        st.rerun()


if st.session_state.show_assistant_dialog:
    assistant_dialog()

# 處理 OCR 識別（從 dialog 觸發；使用 upload_file_data 避免 rerun 後上傳檔案被清除）
if st.session_state.get("start_ocr", False) and ("upload_file_data" in st.session_state or "upload_files" in st.session_state):
    # 優先使用已讀入的 (檔名, bytes)，否則沿用舊的 upload_files（UploadedFile 在 rerun 後可能失效）
    if "upload_file_data" in st.session_state:
        file_data_list = st.session_state.upload_file_data
        del st.session_state.upload_file_data
    else:
        file_data_list = [(f.name, f.getvalue()) for f in st.session_state.upload_files]
        del st.session_state.upload_files
    st.session_state.start_ocr = False
    
    if not api_key:
        st.error("無法辨識：未設定 API 金鑰。請在 **Manage app → Settings → Secrets** 中設定 `GEMINI_API_KEY` 後重新上傳。")
    else:
        # 初始化 session_state 用於存儲結果報告
        if "ocr_report" not in st.session_state: 
            st.session_state.ocr_report = []
        
        success_count = 0
        fail_count = 0
        duplicate_count = 0  # 因重複而跳過，需單獨提示
        user_email = st.session_state.get('user_email', 'default_user')
        # 邏輯架構說明書：上傳前先建立 Batch，單張失敗不影響已寫入
        batch_id = create_batch(user_email, 'ocr')
        
        with st.status("AI 正在分析發票中...", expanded=True) as status:
            prog = st.progress(0)
            n_files = len(file_data_list)
            
            for idx, (fname, fbytes) in enumerate(file_data_list):
                status.update(label=f"正在處理: {fname} ({idx+1}/{n_files})", state="running")
                try:
                    image_obj = Image.open(io.BytesIO(fbytes))
                except Exception as img_err:
                    st.error(f"❌ {fname} 無法讀取圖片: {img_err}")
                    st.session_state.ocr_report.append(f"{fname}: 無法讀取圖片 {img_err}")
                    fail_count += 1
                    prog.progress((idx+1)/n_files)
                    continue
                data, err = process_ocr(image_obj, fname, model, api_key)
                
                if data:
                    def clean_n(v):
                        try: return float(str(v).replace(',','').replace('$',''))
                        except: return 0.0
                    
                    # 處理空值：確保所有字段都有值
                    def safe_value(val, default='No'):
                        if val is None or val == '' or val == 'N/A':
                            return default
                        return str(val)
                    
                    # 檢查數據是否完整，用於設置狀態
                    def check_data_complete(data):
                        key_fields = ['date', 'invoice_no', 'seller_name', 'total']
                        for field in key_fields:
                            val = data.get(field, '')
                            if not val or val == 'N/A' or val == '' or (isinstance(val, (int, float)) and val == 0 and field == 'total'):
                                return False
                        return True
                    
                    # 檢查重複發票（即使發票號碼為"No"也要檢查，因為可能是同一張發票重複上傳）
                    invoice_no = safe_value(data.get("invoice_no"), "No")
                    invoice_date = safe_value(data.get("date"), datetime.now().strftime("%Y/%m/%d"))
                    
                    # 檢查重複：如果發票號碼不是"No"，使用發票號碼+日期檢查；如果是"No"，使用日期+賣方名稱檢查
                    is_duplicate = False
                    dup_id = None
                    
                    if invoice_no and invoice_no != "No" and invoice_no != "N/A":
                        # 正常情況：使用發票號碼+日期檢查
                        is_duplicate, dup_id = check_duplicate_invoice(invoice_no, invoice_date, user_email)
                    else:
                        # 發票號碼為"No"的情況：使用日期+賣方名稱+檔案名稱檢查（避免同一檔案重複上傳）
                        seller_name = safe_value(data.get("seller_name"), "")
                        file_name = fname
                        
                        if st.session_state.use_memory_mode:
                            # 內存模式檢查
                            for inv in st.session_state.local_invoices:
                                inv_user = inv.get('user_email', inv.get('user_id', 'default_user'))
                                if (inv_user == user_email and 
                                    inv.get('date') == invoice_date and
                                    inv.get('seller_name') == seller_name and
                                    inv.get('file_name') == file_name):
                                    is_duplicate = True
                                    dup_id = inv.get('id')
                                    break
                        else:
                            # 數據庫模式檢查
                            query = "SELECT id FROM invoices WHERE user_email = ? AND date = ? AND seller_name = ? AND file_name = ?"
                            result = run_query(query, (user_email, invoice_date, seller_name, file_name), is_select=True)
                            if not result.empty:
                                is_duplicate = True
                                dup_id = result.iloc[0]['id']
                    
                    if is_duplicate:
                        st.warning(f"⚠️ {fname}: 疑似重複發票（發票號碼: {invoice_no}, 日期: {invoice_date}），已跳過不重複新增")
                        duplicate_count += 1
                        continue
                    
                    # 保存圖片（多用戶版本：使用 user_email）
                    image_path = save_invoice_image(image_obj.copy(), fname, user_email)
                    
                    # 根據存儲模式選擇不同的保存方式
                    if st.session_state.use_memory_mode:
                        # 使用內存模式（含 batch_id、tax_type）
                        invoice_record = {
                            'id': len(st.session_state.local_invoices) + 1,
                            'user_email': user_email,
                            'file_name': safe_value(data.get("file_name"), "未命名"),
                            'date': safe_value(data.get("date"), datetime.now().strftime("%Y/%m/%d")),
                            'invoice_number': safe_value(data.get("invoice_no"), "No"),
                            'seller_name': safe_value(data.get("seller_name"), "No"),
                            'seller_ubn': safe_value(data.get("seller_ubn"), "No"),
                            'subtotal': clean_n(data.get("subtotal", 0)),
                            'tax': clean_n(data.get("tax", 0)),
                            'total': clean_n(data.get("total", 0)),
                            'category': safe_value(data.get("type"), "其他"),
                            'subject': safe_value(data.get("category_suggest"), "雜項"),
                            'status': "❌ 缺失" if not check_data_complete(data) else safe_value(data.get("status"), "✅ 正常"),
                            'note': safe_value(data.get("note") or data.get("備註"), ""),
                            'image_path': image_path,
                            'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            'batch_id': batch_id,
                            'tax_type': '5%'
                        }
                        st.session_state.local_invoices.append(invoice_record)
                        st.session_state.data_saved = True
                    else:
                        # 使用數據庫 - 確保數據保存（含 batch_id、tax_type）
                        init_db()
                        
                        # 讀取圖片數據（如果圖片路徑存在）
                        image_data = None
                        if image_path and os.path.exists(image_path):
                            try:
                                with open(image_path, 'rb') as img_file:
                                    image_data = img_file.read()
                            except:
                                pass
                        
                        q = "INSERT INTO invoices (user_email, file_name, date, invoice_number, seller_name, seller_ubn, subtotal, tax, total, category, subject, status, note, image_path, image_data, batch_id, tax_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
                        insert_params = (
                            user_email, 
                            safe_value(data.get("file_name"), "未命名"),
                            safe_value(data.get("date"), datetime.now().strftime("%Y/%m/%d")),
                            safe_value(data.get("invoice_no"), "No"),
                            safe_value(data.get("seller_name"), "No"),
                            safe_value(data.get("seller_ubn"), "No"),
                            clean_n(data.get("subtotal", 0)),
                            clean_n(data.get("tax", 0)),
                            clean_n(data.get("total", 0)),
                            safe_value(data.get("type"), "其他"),
                            safe_value(data.get("category_suggest"), "雜項"),
                            "❌ 缺失" if not check_data_complete(data) else safe_value(data.get("status"), "✅ 正常"),
                            safe_value(data.get("note") or data.get("備註"), ""),
                            image_path,
                            image_data,
                            batch_id,
                            '5%'
                        )
                        
                        result = run_query(q, insert_params, is_select=False)
                        
                        if not result:
                            st.error(f"⚠️ 數據保存失敗，請檢查資料庫連線")
                            if st.session_state.db_error:
                                st.error(f"錯誤詳情: {st.session_state.db_error}")
                            # 如果數據庫保存失敗，嘗試切換到內存模式
                            st.warning("💡 嘗試切換到內存模式保存數據...")
                            invoice_record = {
                                'id': len(st.session_state.local_invoices) + 1,
                                'user_email': user_email,
                                'file_name': safe_value(data.get("file_name"), "未命名"),
                                'date': safe_value(data.get("date"), datetime.now().strftime("%Y/%m/%d")),
                                'invoice_number': safe_value(data.get("invoice_no"), "No"),
                                'seller_name': safe_value(data.get("seller_name"), "No"),
                                'seller_ubn': safe_value(data.get("seller_ubn"), "No"),
                                'subtotal': clean_n(data.get("subtotal", 0)),
                                'tax': clean_n(data.get("tax", 0)),
                                'total': clean_n(data.get("total", 0)),
                                'category': safe_value(data.get("type"), "其他"),
                                'subject': safe_value(data.get("category_suggest"), "雜項"),
                                'note': safe_value(data.get("note") or data.get("備註"), ""),
                                'status': "❌ 缺失" if not check_data_complete(data) else safe_value(data.get("status"), "✅ 正常"),
                                'image_path': image_path,
                                'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                'batch_id': batch_id,
                                'tax_type': '5%'
                            }
                            st.session_state.local_invoices.append(invoice_record)
                            st.session_state.use_memory_mode = True
                            st.session_state.data_saved = True
                        else:
                            st.session_state.data_saved = True
                    success_count += 1
                else:
                    st.error(f"❌ {fname} 失敗: {err}")
                    st.session_state.ocr_report.append(f"{fname}: {err}")
                    fail_count += 1
                
                prog.progress((idx+1)/n_files)
            
            status.update(label=f"處理完成! 成功: {success_count}, 跳過重複: {duplicate_count}, 失敗: {fail_count}", state="complete", expanded=True)
        
        # 重複發票：明確提示（避免用戶以為沒任何反應）
        if duplicate_count > 0:
            st.info(f"ℹ️ **已跳過 {duplicate_count} 張重複發票**：發票號碼與日期已存在於列表中，未重複新增。若要再次匯入請先刪除舊資料。")
        
        # 若有辨識失敗（API/網路/解析錯誤），在頂部顯示明確摘要
        if fail_count > 0:
            st.error(f"⚠️ 辨識失敗 {fail_count} 張。常見原因：API 金鑰錯誤或過期、網路問題、圖片不清晰。請確認 Secrets 中的 GEMINI_API_KEY 正確，或展開上方詳情查看具體錯誤。")
            if st.session_state.get("ocr_report"):
                with st.expander("查看失敗詳情", expanded=True):
                    for line in st.session_state.ocr_report:
                        st.text(line)
        
        # 簡化顯示識別結果（只顯示摘要，不顯示圖片預覽）
        if success_count > 0:
            st.success(f"✅ 成功辨識並新增 {success_count} 張發票")
            if fail_count > 0:
                st.warning(f"⚠️ {fail_count} 張辨識失敗")
            if duplicate_count > 0:
                st.caption(f"另有 {duplicate_count} 張因重複已跳過。")
            # 自動清空圖片預覽，節省空間
            if "ocr_images" in st.session_state:
                st.session_state.ocr_images = []
            time.sleep(0.5)
            st.rerun()

# 處理數據導入（從 dialog 觸發）
if st.session_state.get("start_import", False) and "import_file" in st.session_state:
    uploaded_file = st.session_state.import_file
    st.session_state.start_import = False
    del st.session_state.import_file
    
    try:
        # 讀取文件
        if uploaded_file.name.endswith('.csv'):
            import_df = pd.read_csv(uploaded_file)
        else:
            try:
                import_df = pd.read_excel(uploaded_file)
            except:
                st.error("請安裝 openpyxl 庫以支持 Excel 文件: pip install openpyxl")
                st.stop()
        
        if import_df.empty:
            st.error("文件為空，請檢查文件內容")
        else:
            # 列名映射（支持多種可能的列名）
            column_mapping = {
                "檔案名稱": ["檔案名稱", "file_name", "檔名", "文件名"],
                "日期": ["日期", "date", "Date"],
                "發票號碼": ["發票號碼", "invoice_number", "invoice_no", "發票號"],
                "賣方名稱": ["賣方名稱", "seller_name", "賣方", "商家名稱"],
                "賣方統編": ["賣方統編", "seller_ubn", "統編", "統一編號"],
                "銷售額": ["銷售額", "subtotal", "未稅金額"],
                "稅額": ["稅額", "tax", "Tax"],
                "總計": ["總計", "total", "Total", "金額"],
                "類型": ["類型", "category", "Category"],
                "會計科目": ["會計科目", "subject", "Subject", "科目"],
                "備註": ["備註", "note", "Note", "备注", "備注"]
            }
            
            # 標準化列名
            for standard_name, possible_names in column_mapping.items():
                for possible_name in possible_names:
                    if possible_name in import_df.columns:
                        import_df.rename(columns={possible_name: standard_name}, inplace=True)
                        break
            
            # 檢查必填字段
            required_fields = ["日期", "發票號碼", "總計"]
            missing_fields = [f for f in required_fields if f not in import_df.columns]
            if missing_fields:
                st.error(f"缺少必填字段: {', '.join(missing_fields)}")
            else:
                # 開始導入（邏輯架構說明書：上傳前先建立 Batch）
                user_email = st.session_state.get('user_email', 'default_user')
                batch_id = create_batch(user_email, 'import')
                imported_count = 0
                duplicate_count = 0
                error_count = 0
                
                with st.status("正在導入數據...", expanded=False) as status:
                    for idx, row in import_df.iterrows():
                        try:
                            # 檢查重複
                            invoice_no = str(row.get("發票號碼", "No"))
                            invoice_date = str(row.get("日期", ""))
                            is_dup, _ = check_duplicate_invoice(invoice_no, invoice_date, user_email)
                            
                            if is_dup:
                                duplicate_count += 1
                                continue
                            
                            # 處理數值
                            def safe_float(val):
                                try:
                                    return float(str(val).replace(',', '').replace('$', ''))
                                except:
                                    return 0.0
                            
                            def safe_str(val, default="No"):
                                val_str = str(val) if not pd.isna(val) else ""
                                return val_str if val_str.strip() else default
                            
                            # 保存數據（含 batch_id、tax_type）
                            if st.session_state.use_memory_mode:
                                invoice_record = {
                                    'id': len(st.session_state.local_invoices) + 1,
                                    'user_email': user_email,
                                    'file_name': safe_str(row.get("檔案名稱"), "導入數據"),
                                    'date': safe_str(row.get("日期"), datetime.now().strftime("%Y/%m/%d")),
                                    'invoice_number': safe_str(row.get("發票號碼"), "No"),
                                    'seller_name': safe_str(row.get("賣方名稱"), "No"),
                                    'seller_ubn': safe_str(row.get("賣方統編"), "No"),
                                    'subtotal': safe_float(row.get("銷售額", 0)),
                                    'tax': safe_float(row.get("稅額", 0)),
                                    'total': safe_float(row.get("總計", 0)),
                                    'category': safe_str(row.get("類型"), "其他"),
                                    'subject': safe_str(row.get("會計科目"), "雜項"),
                                    'status': "✅ 正常",
                                    'note': safe_str(row.get("備註"), ""),
                                    'image_path': None,
                                    'created_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                    'batch_id': batch_id,
                                    'tax_type': '5%'
                                }
                                st.session_state.local_invoices.append(invoice_record)
                                imported_count += 1
                            else:
                                init_db()
                                q = "INSERT INTO invoices (user_email, file_name, date, invoice_number, seller_name, seller_ubn, subtotal, tax, total, category, subject, status, note, batch_id, tax_type) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"
                                params = (
                                    user_email,
                                    safe_str(row.get("檔案名稱"), "導入數據"),
                                    safe_str(row.get("日期"), datetime.now().strftime("%Y/%m/%d")),
                                    safe_str(row.get("發票號碼"), "No"),
                                    safe_str(row.get("賣方名稱"), "No"),
                                    safe_str(row.get("賣方統編"), "No"),
                                    safe_float(row.get("銷售額", 0)),
                                    safe_float(row.get("稅額", 0)),
                                    safe_float(row.get("總計", 0)),
                                    safe_str(row.get("類型"), "其他"),
                                    safe_str(row.get("會計科目"), "雜項"),
                                    "✅ 正常",
                                    safe_str(row.get("備註"), ""),
                                    batch_id,
                                    '5%'
                                )
                                if run_query(q, params, is_select=False):
                                    imported_count += 1
                                else:
                                    error_count += 1
                        
                        except Exception as e:
                            error_count += 1
                
                # 顯示結果
                if imported_count > 0:
                    st.success(f"✅ 成功導入 {imported_count} 筆數據")
                if duplicate_count > 0:
                    st.warning(f"⚠️ 跳過 {duplicate_count} 筆重複數據")
                if error_count > 0:
                    st.error(f"❌ {error_count} 筆數據導入失敗")
                
                if imported_count > 0:
                    time.sleep(1)
                    st.rerun()
                    
    except Exception as e:
        st.error(f"導入失敗: {str(e)}")

# ========== 3. 圖表展示區 ==========
st.markdown("---")
st.subheader("分析圖表")
with st.container():
    # 準備數據（如果df_stats已定義，使用它；否則使用df_raw並重命名）
    if 'df_stats' in locals() and not df_stats.empty:
        df_chart = df_stats.copy()
    else:
        df_chart = df_raw.copy()
        if not df_chart.empty:
            mapping = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編","subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間"}
            df_chart = df_chart.rename(columns=mapping)
    
    if not df_chart.empty:
        # 三列布局，使图表更紧凑
        chart_col1, chart_col2, chart_col3 = st.columns(3)
        
        # 图表高度调整为更紧凑（参考图片样式）
        chart_height = 220
            
        with chart_col1:
            # 圓餅圖 - 會計科目分布
            st.markdown("**會計科目分布**")
            if "會計科目" in df_chart.columns:
                df_pie = df_chart[df_chart['會計科目'].notna() & (df_chart['會計科目'] != 'No')].copy()
                if not df_pie.empty:
                    # 使用参考图片的颜色方案（蓝色系）
                    chart = alt.Chart(df_pie).mark_arc(innerRadius=25).encode(
                        theta=alt.Theta("count()", type="quantitative"),
                        color=alt.Color("會計科目", type="nominal", 
                                       scale=alt.Scale(scheme='blues')),
                        tooltip=["會計科目", "count()"]
                    ).properties(
                        height=chart_height,
                        background='#2F2F2F'
                    ).configure_legend(
                        labelFontSize=14,
                        titleFontSize=14,
                        labelColor='#E0E0E0',
                        titleColor='#FFFFFF'
                    ).configure_axis(
                        labelFontSize=14,
                        titleFontSize=0,
                        labelColor='#E0E0E0',
                        titleColor='#FFFFFF',
                        gridColor='#3F3F3F',
                        domainColor='#5F5F5F'
                    ).configure_text(
                        fontSize=14
                    )
                    st.altair_chart(chart, use_container_width=True, theme='streamlit')
                else:
                    st.info("📊 暫無數據", icon="ℹ️")
            elif "subject" in df_raw.columns:
                df_pie = df_raw[df_raw['subject'].notna() & (df_raw['subject'] != 'No')].copy()
                if not df_pie.empty:
                    chart = alt.Chart(df_pie).mark_arc(innerRadius=30).encode(
                        theta=alt.Theta("count()", type="quantitative"),
                        color=alt.Color("subject", type="nominal"),
                        tooltip=["subject", "count()"]
                    ).properties(height=chart_height).configure_legend(
                        labelFontSize=9,
                        titleFontSize=10
                    ).configure_axis(
                        labelFontSize=9,
                        titleFontSize=10
                    )
                    st.altair_chart(chart, use_container_width=True)
                else:
                    st.info("📊 暫無數據", icon="ℹ️")
            else:
                st.info("📊 暫無數據", icon="ℹ️")
        
        with chart_col2:
            # 折線圖 - 每日支出趨勢
            st.markdown("**每日支出趨勢**")
            if "日期" in df_chart.columns and "總計" in df_chart.columns:
                df_line = df_chart.copy()
                df_line['日期'] = pd.to_datetime(df_line['日期'], errors='coerce', format='%Y/%m/%d')
                df_line = df_line.dropna(subset=['日期'])
                
                if not df_line.empty:
                    df_line_grouped = df_line.groupby('日期')['總計'].sum().reset_index()
                    df_line_grouped = df_line_grouped.sort_values('日期')
                    
                    # 使用参考图片的颜色（绿色线条）
                    line_chart = alt.Chart(df_line_grouped).mark_line(
                        point=True, 
                        strokeWidth=3,
                        color='#34A853'  # 绿色，参考图片
                    ).encode(
                        x=alt.X('日期:T', title='', axis=alt.Axis(format='%Y/%m/%d')),
                        y=alt.Y('總計:Q', title='', axis=alt.Axis(format='$,.0f')),
                        tooltip=[alt.Tooltip('日期:T', format='%Y/%m/%d', title='日期'), alt.Tooltip('總計:Q', format='$,.0f', title='金額')]
                    ).properties(
                        height=chart_height,
                        background='#2F2F2F'
                    ).configure_axis(
                        labelFontSize=14,
                        titleFontSize=0,
                        labelColor='#E0E0E0',
                        titleColor='#FFFFFF',
                        gridColor='#3F3F3F',
                        domainColor='#5F5F5F'
                    ).configure_text(
                        fontSize=14
                    ).configure_legend(
                        labelFontSize=14,
                        titleFontSize=14
                    )
                    st.altair_chart(line_chart, use_container_width=True, theme='streamlit')
                else:
                    st.info("📈 暫無數據", icon="ℹ️")
            else:
                st.info("📈 暫無數據", icon="ℹ️")
        
        with chart_col3:
            # 柱狀圖 - 類型分布
            st.markdown("**類型分布**")
            if "類型" in df_chart.columns:
                df_bar = df_chart[df_chart['類型'].notna() & (df_chart['類型'] != 'No')].copy()
                if not df_bar.empty:
                    df_bar_grouped = df_bar.groupby('類型').size().reset_index(name='數量')
                    df_bar_grouped = df_bar_grouped.sort_values('數量', ascending=False).head(10)  # 只顯示前10個
                    
                    # 使用参考图片的颜色（蓝色/青色柱状图）
                    bar_chart = alt.Chart(df_bar_grouped).mark_bar(
                        color='#4285F4',  # 蓝色，参考图片
                        cornerRadiusTopLeft=2,
                        cornerRadiusTopRight=2
                    ).encode(
                        x=alt.X('類型:N', title='', sort='-y', axis=alt.Axis(labelAngle=0)),
                        y=alt.Y('數量:Q', title=''),
                        tooltip=[alt.Tooltip('類型:N', title='類型'), alt.Tooltip('數量:Q', title='數量')]
                    ).properties(
                        height=chart_height,
                        background='#2F2F2F'
                    ).configure_axis(
                        labelFontSize=14,
                        titleFontSize=0,
                        labelColor='#E0E0E0',
                        titleColor='#FFFFFF',
                        gridColor='#3F3F3F',
                        domainColor='#5F5F5F'
                    ).configure_text(
                        fontSize=14
                    ).configure_legend(
                        labelFontSize=14,
                        titleFontSize=14
                    )
                    st.altair_chart(bar_chart, use_container_width=True, theme='streamlit')
                else:
                    st.info("📊 暫無數據", icon="ℹ️")
            elif "category" in df_raw.columns:
                df_bar = df_raw[df_raw['category'].notna() & (df_raw['category'] != 'No')].copy()
                if not df_bar.empty:
                    df_bar_grouped = df_bar.groupby('category').size().reset_index(name='數量')
                    df_bar_grouped = df_bar_grouped.sort_values('數量', ascending=False).head(10)
                    
                    bar_chart = alt.Chart(df_bar_grouped).mark_bar().encode(
                        x=alt.X('category:N', title='類型', sort='-y'),
                        y=alt.Y('數量:Q', title='數量'),
                        color=alt.Color('category:N', legend=None),
                        tooltip=['category', '數量']
                    ).properties(
                        height=chart_height
                    ).configure_axis(
                        labelFontSize=9,
                        titleFontSize=10
                    )
                    st.altair_chart(bar_chart, use_container_width=True)
                else:
                    st.info("📊 暫無數據", icon="ℹ️")
            else:
                st.info("📊 暫無數據", icon="ℹ️")
    else:
        st.info("📊 目前無數據可顯示圖表")

# ========== 4. 數據表格區（發票明細與編輯）==========
st.markdown("---")
with st.container():
    # st.markdown("### 📋 數據稽核報表")  # 隱藏表頭
    
    # 使用原始查詢結果（如果df_stats已定義，使用它；否則使用df_raw並重命名）
    if 'df_stats' in locals() and not df_stats.empty:
        df = df_stats.copy()
        # 保存帶ID的副本用於刪除功能（僅後端使用，不在前端顯示）
        df_with_id = df.copy() if 'id' in df.columns else None
    else:
        df = df_raw.copy()
        # 保存帶ID的副本用於刪除功能（在重命名前，僅後端使用）
        df_with_id = df.copy() if 'id' in df.columns else None
        # 如果使用df_raw，需要重命名列
        if not df.empty:
            # 確保 df_raw 不為空，添加調試信息
            if df_raw.empty:
                st.warning("⚠️ 數據庫查詢結果為空，請檢查數據是否已正確保存")
            mapping = {
                "file_name":"檔案名稱",
                "date":"日期",
                "invoice_number":"發票號碼",
                "seller_name":"賣方名稱",
                "seller_ubn":"賣方統編",
                "subtotal":"銷售額",
                "tax":"稅額",
                "total":"總計",
                "category":"類型",
                "subject":"會計科目",
                "status":"狀態",
                "note":"備註",
                "created_at":"建立時間",
                "tax_type":"稅率類型",
                "modified_at":"修改時間"
            }
            df = df.rename(columns=mapping)
            # 同時重命名df_with_id的列（如果存在）
            if df_with_id is not None and not df_with_id.empty:
                df_with_id = df_with_id.rename(columns=mapping)

    # 確保主資料表一定有「總計」欄位（由銷售額 + 稅額計算）
    if not df.empty:
        has_subtotal = "銷售額" in df.columns
        has_tax = "稅額" in df.columns
        has_total = "總計" in df.columns

        # 若沒有總計，但有銷售額或稅額，則自動計算：總計 = 銷售額 + 稅額
        if not has_total and (has_subtotal or has_tax):
            subtotal_series = pd.to_numeric(df["銷售額"], errors="coerce").fillna(0) if has_subtotal else pd.Series(0, index=df.index)
            tax_series = pd.to_numeric(df["稅額"], errors="coerce").fillna(0) if has_tax else pd.Series(0, index=df.index)
            df["總計"] = (subtotal_series + tax_series).round(0)
    
    # 去掉重複列：以「發票號碼+日期」為業務鍵（同一張發票可能被插入多次），再依 id、整列去重
    if not df.empty:
        if "發票號碼" in df.columns and "日期" in df.columns:
            df = df.drop_duplicates(subset=["發票號碼", "日期"], keep="first")
        if "id" in df.columns and df["id"].notna().any() and df.duplicated(subset=["id"]).any():
            df = df.drop_duplicates(subset=["id"], keep="first")
        elif len(df) != len(df.drop_duplicates(keep="first")):
            df = df.drop_duplicates(keep="first")
        if df_with_id is not None and not df_with_id.empty:
            df_with_id = df_with_id.loc[df.index.intersection(df_with_id.index)].copy()
    
    # 保留未篩選的完整數據（按組視圖與導出全部用）
    df_base = df.copy() if not df.empty else df
    
    # ========== 篩選與操作（Material 3：篩選條件 / 操作 分區）==========
    if "preview_selected_count" not in st.session_state:
        st.session_state.preview_selected_count = 0
    delete_button_top = False  # 預設為未點擊

    today = datetime.now().date()
    mapping_opt = {"file_name":"檔案名稱","date":"日期","invoice_number":"發票號碼","seller_name":"賣方名稱","seller_ubn":"賣方統編","subtotal":"銷售額","tax":"稅額","total":"總計","category":"類型","subject":"會計科目","status":"狀態","note":"備註","created_at":"建立時間"}
    df_opt = df_raw.rename(columns=mapping_opt) if not df_raw.empty else pd.DataFrame()
    subjects = sorted([x for x in df_opt["會計科目"].dropna().astype(str).unique().tolist() if x and str(x).strip() and str(x) != "No"]) if not df_opt.empty and "會計科目" in df_opt.columns else []
    categories = sorted([x for x in df_opt["類型"].dropna().astype(str).unique().tolist() if x and str(x).strip() and str(x) != "No"]) if not df_opt.empty and "類型" in df_opt.columns else []
    subject_options = sorted(set(list(subjects) + ["雜項", "餐飲費", "交通費", "辦公用品", "差旅費"]))
    category_options = sorted(set(list(categories) + ["其他", "餐飲", "交通", "辦公用品"]))

    filter_row1, filter_row2, filter_row3 = st.columns([2, 1, 1])
    with filter_row1:
        search = st.text_input(
            "搜尋發票號碼或賣方名稱",
            placeholder="輸入發票號碼、賣方名稱或檔名...",
            label_visibility="visible",
            key="main_search_input"
        )
    with filter_row2:
        st.caption("選擇開始與結束日期")
        date_start = st.session_state.get("date_range_start")
        date_end = st.session_state.get("date_range_end")
        # 默認顯示最近一個月（開始 = 今天減 30 天，結束 = 今天）
        if date_start is not None and date_end is not None:
            display_start, display_end = date_start, date_end
        else:
            display_end = today
            display_start = today - timedelta(days=30)
        if display_start > display_end:
            display_start, display_end = display_end, display_start
        date_range_value = st.date_input(
            "日期區間",
            value=(display_start, display_end),
            key="filter_date_range",
            label_visibility="collapsed",
            help="選擇開始與結束日期 (GMT+8)"
        )
        if isinstance(date_range_value, (list, tuple)) and len(date_range_value) == 2:
            dr_start, dr_end = date_range_value[0], date_range_value[1]
        else:
            dr_start = dr_end = date_range_value
        if dr_start and dr_end:
            if dr_start > dr_end:
                dr_start, dr_end = dr_end, dr_start
            # 僅在用戶已選過日期或本次選擇與「默認一個月」不同時才寫入，避免初次載入就篩掉全部數據
            default_range = (display_start, display_end)
            chosen = (dr_start, dr_end)
            if (date_start is not None and date_end is not None) or chosen != default_range:
                st.session_state.date_range_start = dr_start
                st.session_state.date_range_end = dr_end
            else:
                st.session_state.date_range_start = None
                st.session_state.date_range_end = None
    with filter_row3:
        status_filter = st.pills(
            "狀態",
            options=["全部", "正常", "缺失"],
            default="全部",
            label_visibility="visible",
            key="status_filter_pills"
        )

    adv1, adv2, adv3, adv4 = st.columns(4)
    with adv1:
        filter_subjects = st.multiselect("會計科目", options=subjects, default=st.session_state.get("filter_subjects", []), key="filter_subjects")
    with adv2:
        filter_categories = st.multiselect("類型", options=categories, default=st.session_state.get("filter_categories", []), key="filter_categories")
    with adv3:
        filter_amount_min = st.number_input("最小金額", min_value=0, value=int(st.session_state.get("filter_amount_min", 0)), step=100, key="filter_amount_min")
    with adv4:
        filter_amount_max = st.number_input("最大金額", min_value=0, value=int(st.session_state.get("filter_amount_max", 0)), step=100, key="filter_amount_max")

    # 視圖切換：按單張（可選擇、編輯）為預設；按組可導出全部
    view_mode = st.radio("視圖", ["📋 按單張", "📦 按組"], horizontal=True, key="invoice_view_mode", label_visibility="collapsed", index=0)
    is_group_view = view_mode == "📦 按組"
    if not is_group_view and not df_base.empty:
        df = df_base.copy()
        df_before_search = len(df)
        search_term = (search or "").strip().lower()
        if search_term:
            def _safe_search_val(val):
                if val is None: return ""
                try:
                    if pd.isna(val): return ""
                except Exception:
                    pass
                return str(val).strip()
            def match_row(row):
                parts = [_safe_search_val(row.get(col, "")) for col in ["發票號碼", "賣方名稱", "檔案名稱"]]
                return search_term in " ".join(parts).lower()
            df = df[df.apply(match_row, axis=1)]
            if len(df) == 0 and df_before_search > 0:
                st.info(f"💡 搜尋「{search}」沒有匹配到任何數據（已過濾 {df_before_search} 筆）")
        status_filter = st.session_state.get("status_filter_pills", "全部")
        if status_filter != "全部" and "狀態" in df.columns:
            if status_filter == "正常":
                df = df[df["狀態"].astype(str).str.contains("正常", na=False)]
            elif status_filter == "缺失":
                df = df[df["狀態"].astype(str).str.contains("缺失|缺漏|❌", na=False, regex=True)]
        date_start = st.session_state.get("date_range_start")
        date_end = st.session_state.get("date_range_end")
        if date_start is not None and date_end is not None and "日期" in df.columns:
            date_col = "日期"
            try:
                df[date_col] = pd.to_datetime(df[date_col], errors='coerce', format='%Y/%m/%d')
                valid_dates_mask = df[date_col].notna()
                date_filter_mask = (df[date_col].dt.date >= date_start) & (df[date_col].dt.date <= date_end)
                df = df[valid_dates_mask & date_filter_mask]
            except Exception:
                try:
                    def date_in_range(date_str):
                        try:
                            date_val = datetime.strptime(str(date_str), "%Y/%m/%d").date()
                            return date_start <= date_val <= date_end
                        except Exception:
                            return False
                    df = df[df[date_col].astype(str).apply(date_in_range)]
                except Exception:
                    pass
        filter_subjects = st.session_state.get("filter_subjects", [])
        if filter_subjects and "會計科目" in df.columns:
            df = df[df["會計科目"].astype(str).isin(filter_subjects)]
        filter_categories = st.session_state.get("filter_categories", [])
        if filter_categories and "類型" in df.columns:
            df = df[df["類型"].astype(str).isin(filter_categories)]
        if "總計" in df.columns:
            total_num = pd.to_numeric(df["總計"].astype(str).str.replace(",", ""), errors="coerce").fillna(0)
            amount_min = st.session_state.get("filter_amount_min", 0) or 0
            amount_max = st.session_state.get("filter_amount_max", 0) or 0
            mask = pd.Series(True, index=df.index)
            if amount_min > 0: mask = mask & (total_num >= amount_min)
            if amount_max > 0: mask = mask & (total_num <= amount_max)
            df = df[mask]

    st.caption("💡 篩選在 **按單張** 視圖生效；**按組** 可導出全部。")
    # 操作按鈕（刪除、CSV、Excel、PDF）已移至「按單張」視圖中「共 N 筆…」說明下方
    # 移除image相關的列
    if not df.empty:
        columns_to_drop = ['image_data', 'imageData', 'image_path']  # 移除所有圖片相關列
        for col in columns_to_drop:
            if col in df.columns:
                df = df.drop(columns=[col])
            if df_with_id is not None and col in df_with_id.columns:
                df_with_id = df_with_id.drop(columns=[col])
        
        # 移除ID列、user_id列與檔案名稱列（不在表格中顯示，但保留在df_with_id中）
        # 「銷售額(未稅)」「稅額」「總計」保留在前端表格中顯示
        columns_to_hide = ['id', 'user_id', 'user_email', '檔案名稱']
        for col in columns_to_hide:
            if col in df.columns:
                df = df.drop(columns=[col])
    
    # 修復 Bug #3: 在篩選前保存原始索引映射，以便刪除功能正常工作
    # 將df_with_id保存到session_state，確保刪除功能可以訪問
    if 'df_with_id' in locals() and df_with_id is not None:
        # 創建索引到ID的映射，保存到session_state
        if 'id' in df_with_id.columns:
            # 創建一個映射：df的索引 -> id
            index_to_id_map = {}
            for idx in df_with_id.index:
                if idx in df.index:
                    index_to_id_map[idx] = df_with_id.loc[idx, 'id']
            st.session_state.index_to_id_map = index_to_id_map
        st.session_state.df_with_id = df_with_id.copy()
    
    if not df.empty and df_with_id is not None:
        # 保存原始索引到df中（在篩選前），用於刪除功能
        df['_original_index'] = df.index
    
    # ========== 2. 發票明細與編輯 ==========
    st.subheader("📋 發票明細與編輯")
    _user_email = st.session_state.get('user_email', 'default_user')

    # 詳情彈出框：點「查看詳情」時以 dialog 顯示（按組／按單張共用）
    if st.session_state.get("detail_invoice_id") is not None:
        _inv_id = st.session_state.detail_invoice_id
        _row = get_invoice_by_id(_inv_id, _user_email)
        @st.dialog("發票詳情")
        def _invoice_detail_dialog():
            if not _row:
                st.warning("找不到該筆發票")
                if st.button("關閉", key="detail_dialog_close"):
                    st.session_state.detail_invoice_id = None
                    st.rerun()
                return
            def _esc(s):
                if s is None or (isinstance(s, float) and pd.isna(s)):
                    return ""
                return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")[:50]
            col_left, col_right = st.columns([2, 1])
            with col_left:
                date_str = str(_row.get("日期", "") or "")[:10]
                num_str = str(_row.get("發票號碼", "") or "")
                st.markdown('<div class="detail-section"><h3 class="detail-title">發票</h3>', unsafe_allow_html=True)
                st.markdown(f'<p class="detail-meta">開立日期 {date_str}　發票號碼 {num_str}</p>', unsafe_allow_html=True)
                st.markdown('<hr class="detail-divider">', unsafe_allow_html=True)
                seller_name = str(_row.get("賣方名稱", "") or "")
                seller_ubn = str(_row.get("賣方統編", "") or "")
                buyer_name = st.session_state.get("company_name", "") or ""
                buyer_ubn = st.session_state.get("company_ubn", "") or ""
                st.markdown(
                    '<div class="detail-from-to">'
                    '<div class="detail-block"><span class="detail-label">賣方</span><p class="detail-address">' + _esc(seller_name) + " " + _esc(seller_ubn) + '</p></div>'
                    '<div class="detail-block"><span class="detail-label">買方</span><p class="detail-address">' + (_esc(buyer_name) + " " + _esc(buyer_ubn)).strip() or "（請在「設定公司資訊」填寫）" + '</p></div>'
                    '</div>',
                    unsafe_allow_html=True,
                )
                st.markdown('<hr class="detail-divider">', unsafe_allow_html=True)
                st.markdown('<p class="detail-label">項目與金額</p>', unsafe_allow_html=True)
                detail_rows = []
                for label, key in [("銷售額", "銷售額"), ("稅額", "稅額"), ("未稅金額", "未稅金額"), ("總計", "總計")]:
                    v = _row.get(key)
                    if v is not None and str(v).strip() not in ("", "No"):
                        try:
                            detail_rows.append((label, f"{float(v):,.0f}"))
                        except Exception:
                            detail_rows.append((label, str(v)))
                if not detail_rows and _row.get("總計") is not None:
                    try:
                        detail_rows.append(("總計", f"{float(_row['總計']):,.0f}"))
                    except Exception:
                        detail_rows.append(("總計", str(_row.get("總計", ""))))
                if detail_rows:
                    tbl = '<table class="detail-amount-table"><thead><tr><th>項目</th><th class="text-right">金額</th></tr></thead><tbody>'
                    for lbl, amt in detail_rows:
                        tbl += f'<tr><td>{lbl}</td><td class="text-right amount-monospace">{amt}</td></tr>'
                    tbl += "</tbody></table>"
                    st.markdown(tbl, unsafe_allow_html=True)
                try:
                    total_num = float(_row.get("總計", 0) or 0)
                except Exception:
                    total_num = 0
                tax_val = _row.get("稅額") or _row.get("稅額 (5%)")
                try:
                    tax_num = float(tax_val) if tax_val is not None and pd.notna(tax_val) else 0
                except Exception:
                    tax_num = 0
                sub_num = total_num - tax_num
                st.markdown('<hr class="detail-divider">', unsafe_allow_html=True)
                st.markdown(
                    '<table class="detail-summary-table">'
                    f'<tr><td>小計</td><td class="text-right amount-monospace">{sub_num:,.0f}</td></tr>'
                    f'<tr><td>稅額</td><td class="text-right amount-monospace">{tax_num:,.0f}</td></tr>'
                    f'<tr><td class="detail-total-row">總計</td><td class="text-right amount-monospace detail-total-row">{total_num:,.0f}</td></tr>'
                    '</table>',
                    unsafe_allow_html=True,
                )
            with col_right:
                try:
                    total_num = float(_row.get("總計", 0) or 0)
                except Exception:
                    total_num = 0
                status_val = str(_row.get("狀態", ""))
                status_class = "detail-status-ok" if ("正常" in status_val or "✅" in status_val) else "detail-status-warn"
                st.markdown(
                    '<div class="detail-card">'
                    '<span class="detail-card-label">金額</span>'
                    f'<div class="detail-amount">${total_num:,.0f}</div>'
                    f'<span class="detail-status {status_class}">{status_val}</span>'
                    '</div>',
                    unsafe_allow_html=True,
                )
                created = str(_row.get("建立時間", "")) if _row.get("建立時間") is not None else ""
                modified = str(_row.get("修改時間", "")) if _row.get("修改時間") is not None else ""
                items = []
                if created:
                    items.append(("建立發票", created[:19] if len(created) > 19 else created))
                if modified:
                    items.append(("最後修改", modified[:19] if len(modified) > 19 else modified))
                if not items:
                    items.append(("—", ""))
                tl = '<div class="detail-card"><p class="detail-card-label">Activity</p><ul class="detail-timeline">'
                for desc, ts in items:
                    tl += f'<li class="detail-timeline-item"><span class="detail-timeline-dot"></span><span class="detail-timeline-text">{_esc(desc)}</span><span class="detail-timeline-time">{_esc(ts)}</span></li>'
                tl += "</ul></div>"
                st.markdown(tl, unsafe_allow_html=True)
                if st.button("關閉", key="detail_dialog_close"):
                    st.session_state.detail_invoice_id = None
                    st.rerun()
        _invoice_detail_dialog()

    if is_group_view:
        # ---------- 按組：組摘要表 + 可展開明細 + 刪除確認 dialog ----------
        batches_list = get_batches_for_user(_user_email)
        ungrouped_df = get_ungrouped_invoices(_user_email)
        if not batches_list and ungrouped_df.empty:
            st.info("📊 目前沒有數據，請上傳發票圖片或導入 CSV 數據。")
        else:
            st.caption("💡 切換至「按單張」可顯示並編輯數據表格。")
            # 組摘要表（一覽：建立時間、來源、張數、合計、稅額）
            summary_rows = []
            for b in batches_list:
                inv_df = get_invoices_by_batch(b['id'], _user_email)
                if inv_df.empty:
                    continue
                created = (b.get('created_at') or '')[:16].replace('T', ' ')
                src = 'OCR' if (b.get('source') or '') == 'ocr' else '導入'
                total_sum = pd.to_numeric(inv_df.get('總計', 0), errors='coerce').fillna(0).sum()
                tax_sum = pd.to_numeric(inv_df.get('稅額', 0), errors='coerce').fillna(0).sum() if '稅額' in inv_df.columns else 0
                summary_rows.append({"建立時間": created, "來源": src, "張數": len(inv_df), "合計": f"${total_sum:,.0f}", "稅額": f"${tax_sum:,.0f}"})
            if not ungrouped_df.empty:
                total_ug = pd.to_numeric(ungrouped_df.get('總計', 0), errors='coerce').fillna(0).sum()
                tax_ug = pd.to_numeric(ungrouped_df.get('稅額', 0), errors='coerce').fillna(0).sum() if '稅額' in ungrouped_df.columns else 0
                summary_rows.append({"建立時間": "未分組", "來源": "-", "張數": len(ungrouped_df), "合計": f"${total_ug:,.0f}", "稅額": f"${tax_ug:,.0f}"})
            if summary_rows:
                st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)
            for b in batches_list:
                inv_df = get_invoices_by_batch(b['id'], _user_email)
                if inv_df.empty:
                    continue
                created = (b.get('created_at') or '')[:16].replace('T', ' ')
                src = 'OCR' if (b.get('source') or '') == 'ocr' else '導入'
                total_sum = pd.to_numeric(inv_df.get('總計', 0), errors='coerce').fillna(0).sum()
                tax_sum = pd.to_numeric(inv_df.get('稅額', 0), errors='coerce').fillna(0).sum() if '稅額' in inv_df.columns else 0
                with st.expander(f"📦 {created} · {src} · {len(inv_df)} 張 · 合計 ${total_sum:,.0f}", expanded=False):
                    # 本組摘要：總計 | 稅額 | 張數（4px/8px 網格）
                    sum_col1, sum_col2, sum_col3 = st.columns(3)
                    with sum_col1:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">總計</span><span class="batch-summary-value">${:,.0f}</span></div>'.format(total_sum), unsafe_allow_html=True)
                    with sum_col2:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">稅額</span><span class="batch-summary-value">${:,.0f}</span></div>'.format(tax_sum), unsafe_allow_html=True)
                    with sum_col3:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">張數</span><span class="batch-summary-value">{}</span></div>'.format(len(inv_df)), unsafe_allow_html=True)
                    # 組內每條：日期、號碼、廠商、總計、狀態、查看詳情（點擊彈出框）
                    st.markdown('<div class="master-list-table-wrap"><table class="master-list-table"><thead><tr><th class="master-list-th col-date">日期</th><th class="master-list-th col-num">號碼</th><th class="master-list-th col-vendor">廠商</th><th class="master-list-th col-amount">總計</th><th class="master-list-th col-status">狀態</th><th class="master-list-th col-action">操作</th></tr></thead></table></div>', unsafe_allow_html=True)
                    for _, inv_row in inv_df.iterrows():
                        inv_id = inv_row.get('id')
                        date_val = str(inv_row.get('日期', ''))[:10] if pd.notna(inv_row.get('日期')) else ''
                        num_val = str(inv_row.get('發票號碼', '')) if pd.notna(inv_row.get('發票號碼')) else ''
                        vendor_val = str(inv_row.get('賣方名稱', ''))[:40] if pd.notna(inv_row.get('賣方名稱')) else ''
                        try:
                            total_fmt = f"{float(inv_row.get('總計', 0)):,.0f}" if pd.notna(inv_row.get('總計')) else '0'
                        except Exception:
                            total_fmt = '0'
                        status_val = str(inv_row.get('狀態', '')) if pd.notna(inv_row.get('狀態')) else ''
                        status_dot = 'status-ok' if ('正常' in status_val or '✅' in status_val) else 'status-warn'
                        sc0, sc1, sc2, sc3, sc4, sc5 = st.columns([1, 1.2, 2, 1, 1.2, 0.8])
                        with sc0:
                            st.markdown(f'<div class="master-list-row-cell">{date_val}</div>', unsafe_allow_html=True)
                        with sc1:
                            st.markdown(f'<div class="master-list-row-cell">{num_val}</div>', unsafe_allow_html=True)
                        with sc2:
                            st.markdown(f'<div class="master-list-row-cell">{vendor_val}</div>', unsafe_allow_html=True)
                        with sc3:
                            st.markdown(f'<div class="master-list-row-cell amount-monospace">{total_fmt}</div>', unsafe_allow_html=True)
                        with sc4:
                            st.markdown(f'<div class="master-list-row-cell master-list-status"><span class="status-dot {status_dot}"></span><span class="status-text">{status_val}</span></div>', unsafe_allow_html=True)
                        with sc5:
                            if st.button("查看詳情", key=f"detail_inv_{inv_id}", type="secondary"):
                                st.session_state.detail_invoice_id = inv_id
                                st.rerun()
                    if st.button("🗑️ 刪除此組", key=f"del_batch_{b['id']}", type="secondary"):
                        st.session_state["pending_delete_batch_id"] = b["id"]
                        st.rerun()
            if not ungrouped_df.empty:
                total_ug = pd.to_numeric(ungrouped_df.get('總計', 0), errors='coerce').fillna(0).sum()
                tax_ug = pd.to_numeric(ungrouped_df.get('稅額', 0), errors='coerce').fillna(0).sum() if '稅額' in ungrouped_df.columns else 0
                with st.expander(f"📄 未分組 ({len(ungrouped_df)} 張) · 合計 ${total_ug:,.0f}", expanded=False):
                    sum_col1, sum_col2, sum_col3 = st.columns(3)
                    with sum_col1:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">總計</span><span class="batch-summary-value">${:,.0f}</span></div>'.format(total_ug), unsafe_allow_html=True)
                    with sum_col2:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">稅額</span><span class="batch-summary-value">${:,.0f}</span></div>'.format(tax_ug), unsafe_allow_html=True)
                    with sum_col3:
                        st.markdown('<div class="batch-summary-item"><span class="batch-summary-label">張數</span><span class="batch-summary-value">{}</span></div>'.format(len(ungrouped_df)), unsafe_allow_html=True)
                    st.markdown('<div class="master-list-table-wrap"><table class="master-list-table"><thead><tr><th class="master-list-th col-date">日期</th><th class="master-list-th col-num">號碼</th><th class="master-list-th col-vendor">廠商</th><th class="master-list-th col-amount">總計</th><th class="master-list-th col-status">狀態</th><th class="master-list-th col-action">操作</th></tr></thead></table></div>', unsafe_allow_html=True)
                    for _, inv_row in ungrouped_df.iterrows():
                        inv_id = inv_row.get('id')
                        date_val = str(inv_row.get('日期', ''))[:10] if pd.notna(inv_row.get('日期')) else ''
                        num_val = str(inv_row.get('發票號碼', '')) if pd.notna(inv_row.get('發票號碼')) else ''
                        vendor_val = str(inv_row.get('賣方名稱', ''))[:40] if pd.notna(inv_row.get('賣方名稱')) else ''
                        try:
                            total_fmt = f"{float(inv_row.get('總計', 0)):,.0f}" if pd.notna(inv_row.get('總計')) else '0'
                        except Exception:
                            total_fmt = '0'
                        status_val = str(inv_row.get('狀態', '')) if pd.notna(inv_row.get('狀態')) else ''
                        status_dot = 'status-ok' if ('正常' in status_val or '✅' in status_val) else 'status-warn'
                        uc0, uc1, uc2, uc3, uc4, uc5 = st.columns([1, 1.2, 2, 1, 1.2, 0.8])
                        with uc0:
                            st.markdown(f'<div class="master-list-row-cell">{date_val}</div>', unsafe_allow_html=True)
                        with uc1:
                            st.markdown(f'<div class="master-list-row-cell">{num_val}</div>', unsafe_allow_html=True)
                        with uc2:
                            st.markdown(f'<div class="master-list-row-cell">{vendor_val}</div>', unsafe_allow_html=True)
                        with uc3:
                            st.markdown(f'<div class="master-list-row-cell amount-monospace">{total_fmt}</div>', unsafe_allow_html=True)
                        with uc4:
                            st.markdown(f'<div class="master-list-row-cell master-list-status"><span class="status-dot {status_dot}"></span><span class="status-text">{status_val}</span></div>', unsafe_allow_html=True)
                        with uc5:
                            if st.button("查看詳情", key=f"detail_ug_{inv_id}", type="secondary"):
                                st.session_state.detail_invoice_id = inv_id
                                st.rerun()
            # 刪除 Batch 確認：使用 dialog，避免置頂混淆
            if st.session_state.get("pending_delete_batch_id") is not None:
                _bid = st.session_state["pending_delete_batch_id"]
                @st.dialog("⚠️ 確認刪除此組")
                def _delete_batch_dialog():
                    st.warning("確定要刪除此組及其內所有發票？此操作不可恢復。")
                    c1, c2 = st.columns(2)
                    with c1:
                        if st.button("✅ 確認刪除此組", type="primary", use_container_width=True):
                            ok, cnt, err = delete_batch_cascade(_bid, _user_email)
                            st.session_state.pop("pending_delete_batch_id", None)
                            if ok:
                                st.success(f"已刪除此組，共 {cnt} 張發票。")
                                time.sleep(0.5)
                                st.rerun()
                            else:
                                st.error(f"刪除失敗：{err}")
                    with c2:
                        if st.button("❌ 取消", use_container_width=True):
                            st.session_state.pop("pending_delete_batch_id", None)
                            st.rerun()
                _delete_batch_dialog()
        # 按組時導出全部（使用 df_base）
        if not df_base.empty:
            st.markdown("**導出全部**")
            ec1, ec2, ec3 = st.columns(3)
            with ec1:
                csv_data = df_base.to_csv(index=False).encode('utf-8-sig')
                st.download_button("📥 CSV（全部）", csv_data, "invoice_all.csv", mime="text/csv", use_container_width=True, key="export_csv_group")
            with ec2:
                st.caption("Excel／PDF 請在「按單張」視圖使用當前篩選導出。")
    else:
        # 搜尋有字 → 按單張顯示（現有表格 + 篩選）
        if df.empty:
            if not df_raw.empty:
                # 有原始數據但篩選後為空：使用者導向提示
                with st.expander("📋 目前篩選結果為 0 筆", expanded=True):
                    st.write("**目前篩選條件：**")
                    st.write(f"- 關鍵字搜尋: {search if search else '無'}")
                    date_start = st.session_state.get("date_range_start")
                    date_end = st.session_state.get("date_range_end")
                    if date_start and date_end:
                        st.write(f"- 日期範圍: {date_start} ~ {date_end}")
                    else:
                        st.write("- 時間範圍: 全部")
                    st.write(f"- 狀態: {st.session_state.get('status_filter_pills', '全部')}")
                    st.caption("若需顯示更多資料，可放寬條件或清除篩選。")
                    if st.button("🔄 清除所有篩選條件", use_container_width=True, key="clear_filters_empty"):
                        if "time_filter_last_preset" in st.session_state:
                            st.session_state["time_filter_last_preset"] = "全部"
                        if "date_range_start" in st.session_state:
                            st.session_state.date_range_start = None
                        if "date_range_end" in st.session_state:
                            st.session_state.date_range_end = None
                        st.rerun()
            elif df_raw.empty:
                st.info("📊 目前沒有數據，請上傳發票圖片或導入CSV數據")
            else:
                st.info("📊 目前沒有數據，請上傳發票圖片或導入CSV數據")
        else:
            # 處理空值：用"No"替換
            def fill_empty(val):
                if pd.isna(val) or val == '' or val == 'N/A' or str(val).strip() == '':
                    return 'No'
                return str(val)
            
            # 對所有列應用空值處理（除了狀態列，狀態列需要特殊處理）
            for col in df.columns:
                if col not in ['選取', '狀態']:  # 跳過選取和狀態列
                    df[col] = df[col].apply(fill_empty)
            
            # 處理狀態列：檢查是否有缺失數據，如果有則顯示"缺失"
            if "狀態" in df.columns:
                def check_status(row):
                    # 先檢查關鍵字段是否為空或"No"（優先級最高）
                    key_fields = ['日期', '發票號碼', '賣方名稱', '總計']
                    has_missing = False
                    for field in key_fields:
                        if field in row:
                            val = str(row[field]).strip()
                            if pd.isna(row[field]) or val == '' or val == 'N/A' or val == 'No' or val == '未填':
                                has_missing = True
                                break
                    
                    # 如果有缺失，直接返回"缺失"（不考慮原始狀態）
                    if has_missing:
                        return '❌ 缺失'
                    
                    # 如果沒有缺失，再檢查原始狀態
                    original_status = str(row.get('狀態', '')).strip()
                    
                    # 如果原本的狀態已經是錯誤狀態，保持原樣（但確保有紅色X）
                    if '缺漏' in original_status or '缺失' in original_status or '錯誤' in original_status:
                        # 如果已經有❌，保持原樣；如果沒有，添加❌
                        if '❌' not in original_status and '⚠️' not in original_status:
                            return f'❌ {original_status}'
                        return original_status
                    
                    # 如果沒有缺失且原始狀態正常，返回"正常"
                    if original_status and ('正常' in original_status or '✅' in original_status):
                        return '✅ 正常'
                    
                    # 如果原始狀態為空，返回"正常"
                    return '✅ 正常'
            
                df['狀態'] = df.apply(check_status, axis=1)
            
            # 再次確保移除image相關的列
            columns_to_drop = ['image_data', 'imageData', 'image_path']
            for col in columns_to_drop:
                if col in df.columns:
                    df = df.drop(columns=[col])
            
            # 確保ID列保留在df中（用於刪除功能），但不在顯示中顯示
            # 從df_with_id中獲取id列（如果存在）
            if df_with_id is not None and 'id' in df_with_id.columns:
                # 確保df中有id列（用於刪除功能）
                if 'id' not in df.columns:
                    # 通過索引匹配，將id從df_with_id複製到df
                    df = df.copy()
                    df['id'] = None
                    for idx in df.index:
                        if idx in df_with_id.index:
                            df.loc[idx, 'id'] = df_with_id.loc[idx, 'id']
            
            # 移除其他不需要顯示的列（保留「總計」供前端表格使用）
            columns_to_hide = ['user_id', 'user_email', '檔案名稱']
            for col in columns_to_hide:
                if col in df.columns:
                    df = df.drop(columns=[col])
            
            # 自動計算「未稅金額」與「稅額 (5%)」；審計：依稅率類型支援 5%/0%/免稅
            if "總計" in df.columns:
                total_series = pd.to_numeric(df["總計"], errors="coerce").fillna(0)
                tax_type_col = df.get("稅率類型")
                if tax_type_col is None:
                    tax_type_col = pd.Series("5%", index=df.index)
                tax_type_str = tax_type_col.fillna("5%").astype(str).str.strip().str.lower()
                is_zero_or_exempt = tax_type_str.isin(["0%", "exempt", "零稅率", "免稅"])
                
                if "稅額" in df.columns:
                    existing_tax = pd.to_numeric(df["稅額"], errors="coerce").fillna(0)
                    calc_tax = pd.Series(0.0, index=df.index).where(is_zero_or_exempt, (total_series - (total_series / 1.05)).round(0))
                    tax_series = existing_tax.where((existing_tax > 0) | (total_series == 0), calc_tax)
                else:
                    tax_series = pd.Series(0.0, index=df.index).where(is_zero_or_exempt, (total_series - (total_series / 1.05)).round(0))
                
                subtotal_series = (total_series - tax_series).round(0)
                df["未稅金額"] = subtotal_series
                df["稅額 (5%)"] = tax_series
            
            # 為問題行添加警示圖示（發票號碼為 "No" 或狀態為 "缺失"）
            if "發票號碼" in df.columns:
                def add_warning_icon(invoice_no, status):
                    """為問題行添加警示圖示"""
                    invoice_str = str(invoice_no).strip() if pd.notna(invoice_no) else ""
                    status_str = str(status).strip() if pd.notna(status) else ""
                    
                    is_problem = (invoice_str == "No" or invoice_str == "" or 
                                 "缺失" in status_str or "❌" in status_str or "缺漏" in status_str)
                    
                    if is_problem:
                        return "⚠️ " + str(invoice_no) if invoice_str != "No" else "⚠️ No"
                    return str(invoice_no)
                
                df["發票號碼"] = df.apply(
                    lambda row: add_warning_icon(row.get("發票號碼", ""), row.get("狀態", "")), 
                    axis=1
                )
            
            # 調整列順序：選取 -> 狀態 -> 其他列（id列保留但不顯示）
            if "選取" not in df.columns: 
                df.insert(0, "選取", False)
            
            # 將狀態列移到選取列之後
            if "狀態" in df.columns:
                cols = df.columns.tolist()
                cols.remove("狀態")
                if "選取" in cols:
                    select_idx = cols.index("選取")
                    cols.insert(select_idx + 1, "狀態")
                else:
                    cols.insert(0, "狀態")
                df = df[cols]
        
            # 調整金額相關欄位的順序：銷售額 -> 未稅金額 -> 稅額 -> 稅額 (5%) -> 總計
            if "未稅金額" in df.columns and "稅額 (5%)" in df.columns:
                cols = df.columns.tolist()
                # 移除這些欄位
                for col in ["銷售額", "未稅金額", "稅額", "稅額 (5%)", "總計"]:
                    if col in cols:
                        cols.remove(col)
            
                # 找到合適的位置插入（在「狀態」之後，其他欄位之前）
                try:
                    status_idx = cols.index("狀態")
                    insert_pos = status_idx + 1
                except:
                    insert_pos = 1
            
                # 按順序插入金額欄位（包含總計變化）
                amount_cols = ["銷售額", "未稅金額", "稅額", "稅額 (5%)", "總計"]
                for i, col in enumerate(amount_cols):
                    if col in df.columns:
                        cols.insert(insert_pos + i, col)
            
                df = df[cols]
        
            # 在刪除功能使用後，移除 _original_index 列（如果存在）
            if '_original_index' in df.columns:
                df = df.drop(columns=['_original_index'])
        
            # 不再顯示標題和選中數量
            if st.session_state.get("show_delete_confirm", False):
                delete_records = st.session_state.get("delete_records", [])
                delete_count = st.session_state.get("delete_count", 0)
            
                # 使用裝飾器方式定義刪除確認對話框
                @st.dialog("⚠️ 確認刪除")
                def delete_confirm_dialog():
                    st.warning(f"確定要刪除選中的 {delete_count} 條數據嗎？")
                    st.error("⚠️ 此操作不可恢復！")
                
                    # 顯示要刪除的記錄預覽（顯示id、發票號碼和日期）
                    if delete_records:
                        with st.expander("查看要刪除的記錄", expanded=False):
                            # 準備預覽數據，將id、發票號碼、日期格式化顯示
                            preview_data = []
                            for rec in delete_records:
                                preview_row = {}
                                if 'id' in rec and rec['id'] is not None:
                                    preview_row['ID'] = rec['id']
                                if 'invoice_number' in rec and rec.get('invoice_number'):
                                    preview_row['發票號碼'] = rec['invoice_number']
                                else:
                                    preview_row['發票號碼'] = '(空)'
                                if 'date' in rec and rec.get('date'):
                                    preview_row['日期'] = rec['date']
                                else:
                                    preview_row['日期'] = '(空)'
                                preview_data.append(preview_row)
                        
                            if preview_data:
                                preview_df = pd.DataFrame(preview_data)
                                st.dataframe(preview_df, use_container_width=True, hide_index=True)
                            else:
                                st.info("無法顯示記錄詳情")
                
                    col1, col2 = st.columns(2)
                    with col1:
                        if st.button("✅ 確認刪除", type="primary", use_container_width=True):
                            # 執行刪除：使用發票號碼+日期+用戶郵箱組合刪除（最可靠的方式）
                            user_email = st.session_state.get('user_email', 'default_user')
                            deleted_count = 0
                            errors = []
                        
                            if st.session_state.use_memory_mode:
                                # 內存模式：從列表中刪除（優先使用id，否則使用發票號碼+日期）
                                original_count = len(st.session_state.local_invoices)
                            
                                def should_delete_invoice(inv):
                                    """判斷是否應該刪除此發票"""
                                    for rec in delete_records:
                                        # 優先使用id匹配
                                        if 'id' in rec and rec['id'] is not None:
                                            if inv.get('id') == rec['id'] and inv.get('user_email', inv.get('user_id', 'default_user')) == user_email:
                                                return True
                                        # 如果沒有id，使用發票號碼+日期組合
                                        elif 'invoice_number' in rec and 'date' in rec:
                                            inv_num = str(inv.get('invoice_number', '')).strip()
                                            inv_date = str(inv.get('date', '')).strip()
                                            rec_num = str(rec.get('invoice_number', '')).strip()
                                            rec_date = str(rec.get('date', '')).strip()
                                        
                                            if (inv_num == rec_num or (not inv_num and not rec_num)) and \
                                               (inv_date == rec_date or (not inv_date and not rec_date)) and \
                                               inv.get('user_email', inv.get('user_id', 'default_user')) == user_email:
                                                return True
                                        # 如果只有發票號碼（數據不完整）
                                        elif 'invoice_number' in rec and rec.get('invoice_number'):
                                            inv_num = str(inv.get('invoice_number', '')).strip()
                                            rec_num = str(rec.get('invoice_number', '')).strip()
                                            inv_date = str(inv.get('date', '')).strip()
                                        
                                            if inv_num == rec_num and (not inv_date or inv_date in ['', 'No', 'N/A']) and \
                                               inv.get('user_email', inv.get('user_id', 'default_user')) == user_email:
                                                return True
                                        # 如果只有日期（數據不完整）
                                        elif 'date' in rec and rec.get('date'):
                                            inv_date = str(inv.get('date', '')).strip()
                                            rec_date = str(rec.get('date', '')).strip()
                                            inv_num = str(inv.get('invoice_number', '')).strip()
                                        
                                            if inv_date == rec_date and (not inv_num or inv_num in ['', 'No', 'N/A']) and \
                                               inv.get('user_email', inv.get('user_id', 'default_user')) == user_email:
                                                return True
                                    return False
                            
                                st.session_state.local_invoices = [
                                    inv for inv in st.session_state.local_invoices 
                                    if not should_delete_invoice(inv)
                                ]
                                deleted_count = original_count - len(st.session_state.local_invoices)
                            else:
                                # 數據庫模式：優先使用id刪除（支持數據不完整），否則使用發票號碼+日期+用戶郵箱組合
                                try:
                                    path = get_db_path()
                                    is_uri = path.startswith("file:") and "mode=memory" in path
                                    conn = sqlite3.connect(path, timeout=30, uri=is_uri, check_same_thread=False)
                                    cursor = conn.cursor()
                                
                                    # 逐條刪除
                                    for rec in delete_records:
                                        try:
                                            # 優先使用id刪除（最可靠，支持數據不完整）
                                            if 'id' in rec and rec['id'] is not None:
                                                cursor.execute(
                                                    "DELETE FROM invoices WHERE id=? AND user_email=?",
                                                    (rec['id'], user_email)
                                                )
                                            # 如果沒有id，使用發票號碼+日期+用戶郵箱組合
                                            elif 'invoice_number' in rec and 'date' in rec and rec.get('invoice_number') and rec.get('date'):
                                                cursor.execute(
                                                    "DELETE FROM invoices WHERE user_email=? AND invoice_number=? AND date=?",
                                                    (user_email, rec['invoice_number'], rec['date'])
                                                )
                                            # 如果只有發票號碼（數據不完整）
                                            elif 'invoice_number' in rec and rec.get('invoice_number'):
                                                cursor.execute(
                                                    "DELETE FROM invoices WHERE user_email=? AND invoice_number=? AND (date IS NULL OR date='' OR date='No')",
                                                    (user_email, rec['invoice_number'])
                                                )
                                            # 如果只有日期（數據不完整）
                                            elif 'date' in rec and rec.get('date'):
                                                cursor.execute(
                                                    "DELETE FROM invoices WHERE user_email=? AND date=? AND (invoice_number IS NULL OR invoice_number='' OR invoice_number='No')",
                                                    (user_email, rec['date'])
                                                )
                                            else:
                                                errors.append("無法確定要刪除的記錄（缺少必要的標識信息）")
                                                continue
                                        
                                            if cursor.rowcount > 0:
                                                deleted_count += cursor.rowcount
                                            else:
                                                # 記錄未找到的記錄信息
                                                rec_info = f"ID: {rec.get('id', 'N/A')}, 發票號碼: {rec.get('invoice_number', 'N/A')}, 日期: {rec.get('date', 'N/A')}"
                                                errors.append(f"未找到記錄: {rec_info}")
                                        except Exception as e:
                                            rec_info = f"ID: {rec.get('id', 'N/A')}, 發票號碼: {rec.get('invoice_number', 'N/A')}, 日期: {rec.get('date', 'N/A')}"
                                            errors.append(f"刪除失敗（{rec_info}）: {str(e)}")
                                
                                    conn.commit()
                                    conn.close()
                                
                                    if deleted_count == 0 and not errors:
                                        errors.append("未找到要刪除的記錄，可能已被刪除或數據不匹配")
                                    
                                except Exception as e:
                                    errors.append(f"刪除失敗: {str(e)}")
                        
                            # 清理狀態
                            st.session_state.show_delete_confirm = False
                            if "delete_records" in st.session_state:
                                del st.session_state.delete_records
                            if "delete_count" in st.session_state:
                                del st.session_state.delete_count
                        
                            if deleted_count > 0:
                                st.success(f"✅ 已刪除 {deleted_count} 條數據")
                            else:
                                st.warning("⚠️ 未找到要刪除的記錄，可能已被刪除或數據不匹配")
                        
                            if errors:
                                for err in errors:
                                    st.error(err)
                        
                            time.sleep(0.5)
                            st.rerun()
                
                    with col2:
                        if st.button("❌ 取消", use_container_width=True):
                            # 取消刪除，清理狀態
                            st.session_state.show_delete_confirm = False
                            if "delete_records" in st.session_state:
                                del st.session_state.delete_records
                            if "delete_count" in st.session_state:
                                del st.session_state.delete_count
                            st.rerun()
            
                # 調用對話框函數
                delete_confirm_dialog()

            # ========== 單條數據：可選擇、可編輯 ==========
            # 保存原始數據的副本用於比較（不包含ID列）
            original_df_copy = df.copy()
        
            # 處理日期列：嘗試轉換為日期類型（先創建 df_for_editor）
            df_for_editor = df.copy()
            # 去掉表格中不展示的內部欄位（id 保留供刪除/儲存用，由 column_order 隱藏；_original_index 僅內部用）
            if "_original_index" in df_for_editor.columns:
                df_for_editor = df_for_editor.drop(columns=["_original_index"])
        
            # 準備列配置（不包含ID列、user_id列、檔案名稱列）
            # 金額類數字右對齊，文字類左對齊
            column_config = { 
                "選取": st.column_config.CheckboxColumn("選取", default=False),
                "銷售額": st.column_config.NumberColumn("銷售額", format="$%d"),
                "稅額": st.column_config.NumberColumn("稅額", format="$%d"),
                "未稅金額": st.column_config.NumberColumn("未稅金額", format="$%d"),
                "稅額 (5%)": st.column_config.NumberColumn("稅額 (5%)", format="$%d"),
                "總計": st.column_config.NumberColumn("總計", format="$%d"),
                "備註": st.column_config.TextColumn("備註", width="medium"),
                "建立時間": st.column_config.DatetimeColumn("建立時間", format="YYYY-MM-DD"),
                "稅率類型": st.column_config.SelectboxColumn("稅率類型", options=["5%", "0%", "免稅", "零稅率"], required=False)
            }
            column_config["會計科目"] = st.column_config.SelectboxColumn("會計科目", options=subject_options, required=False)
            # 類型 = 發票類型（三聯發票、二聯發票、電子發票、收銀機發票、其它）
            # 僅顯示發票「類型」本身，不再混入報帳分類（餐飲、交通等）
            invoice_type_options = ["三聯發票", "二聯發票", "電子發票", "收銀機發票", "其它"]
            column_config["類型"] = st.column_config.SelectboxColumn("類型", options=invoice_type_options, required=False)
        
            # 文字類欄位左對齊配置（會計科目、類型已用 SelectboxColumn）
            text_columns = ["賣方名稱", "發票號碼", "賣方統編", "狀態", "備註"]
            for col in text_columns:
                if col in df_for_editor.columns and col not in column_config:
                    column_config[col] = st.column_config.TextColumn(col, width="medium")
        
            # 確保id列在df_for_editor中（用於刪除功能），但不在column_config中配置（隱藏顯示）
            # 注意：如果列不在column_config中，Streamlit會自動隱藏它
            # 但為了確保ID列可用於刪除功能，我們需要確保它在df_for_editor中
            if "id" in df_for_editor.columns:
                # id列保留但不配置，這樣它會隱藏顯示但仍然可用於刪除功能
                # 不添加id到column_config，這樣它會被隱藏
                pass
        
            if "日期" in df_for_editor.columns:
                try:
                    # 嘗試將日期字符串轉換為日期類型
                    df_for_editor["日期"] = pd.to_datetime(df_for_editor["日期"], errors='coerce', format='%Y/%m/%d')
                    # 如果轉換成功（沒有全部為NaT），使用DateColumn
                    if not df_for_editor["日期"].isna().all():
                        column_config["日期"] = st.column_config.DateColumn("日期", format="YYYY-MM-DD")
                    else:
                        # 轉換失敗，使用TextColumn
                        column_config["日期"] = st.column_config.TextColumn("日期", width="medium")
                        df_for_editor["日期"] = df["日期"]  # 恢復原始字符串
                except:
                    # 轉換失敗，使用TextColumn
                    column_config["日期"] = st.column_config.TextColumn("日期", width="medium")
                    df_for_editor["日期"] = df["日期"]  # 確保使用原始字符串
        
            # 處理建立時間列（created_at）
            if "建立時間" in df_for_editor.columns:
                try:
                    # 嘗試將建立時間轉換為日期時間類型
                    df_for_editor["建立時間"] = pd.to_datetime(df_for_editor["建立時間"], errors='coerce')
                    if not df_for_editor["建立時間"].isna().all():
                        column_config["建立時間"] = st.column_config.DatetimeColumn("建立時間", format="YYYY-MM-DD")
                    else:
                        column_config["建立時間"] = st.column_config.TextColumn("建立時間", width="medium")
                        df_for_editor["建立時間"] = df["建立時間"]
                except:
                    column_config["建立時間"] = st.column_config.TextColumn("建立時間", width="medium")
                    df_for_editor["建立時間"] = df["建立時間"]
        
            # 操作按鈕列（刪除、CSV、Excel、PDF）
            act_col1, act_col2, act_col3, act_col4 = st.columns(4)
            with act_col1:
                if not df.empty:
                    preview_selected = st.session_state.get("preview_selected_count", 0)
                    if preview_selected > 0:
                        delete_button_top = st.button(
                            f"🗑️ 刪除 {preview_selected} 條",
                            type="primary",
                            use_container_width=True,
                            help="刪除已選中的數據",
                            key="delete_button_top"
                        )
                    else:
                        st.button(
                            "🗑️ 刪除",
                            disabled=True,
                            use_container_width=True,
                            help="請先勾選要刪除的記錄",
                            key="delete_button_top_disabled"
                        )
                        delete_button_top = False
            with act_col2:
                if not df.empty:
                    csv_data = df.to_csv(index=False).encode('utf-8-sig')
                    st.download_button(
                        "📥 CSV",
                        csv_data,
                        "invoice_report.csv",
                        mime="text/csv",
                        use_container_width=True,
                        help="導出當前篩選後的數據為 CSV"
                    )
            with act_col3:
                if not df.empty:
                    def _gen_excel():
                        export_df = df_stats.copy() if ('df_stats' in locals() and not getattr(df_stats, 'empty', True)) else df.copy()
                        if export_df.empty:
                            return b""
                        total_series = pd.to_numeric(export_df.get('總計', 0), errors='coerce').fillna(0)
                        subtotal_series = pd.to_numeric(export_df.get('銷售額', 0), errors='coerce').fillna(0)
                        tax_series = pd.to_numeric(export_df.get('稅額', 0), errors='coerce').fillna(0)
                        tax_type_col = export_df.get('稅率類型')
                        if tax_type_col is None:
                            tax_type_col = pd.Series('5%', index=export_df.index)
                        tax_type_str = tax_type_col.fillna('5%').astype(str).str.strip().str.lower()
                        is_zero_or_exempt = tax_type_str.isin(['0%', 'exempt', '零稅率', '免稅'])
                        need_recalc = ((subtotal_series == 0) | (tax_series == 0)) & (total_series > 0)
                        if need_recalc.any():
                            calc_tax = pd.Series(0.0, index=export_df.index).where(is_zero_or_exempt, (total_series - (total_series / 1.05)).round(0))
                            calc_subtotal = (total_series - calc_tax).round(0)
                            tax_series = tax_series.where(~need_recalc, calc_tax)
                            subtotal_series = subtotal_series.where(~need_recalc, calc_subtotal)
                        if '銷售額' in export_df.columns:
                            export_df = export_df.rename(columns={'銷售額': '銷售額(未稅)'})
                        else:
                            export_df = export_df.copy()
                            export_df['銷售額(未稅)'] = subtotal_series
                        if '稅額' not in export_df.columns:
                            export_df['稅額'] = tax_series
                        else:
                            export_df['稅額'] = tax_series
                        if '總計' not in export_df.columns:
                            export_df['總計'] = total_series
                        else:
                            export_df['總計'] = total_series
                        desired_order = ["日期", "發票號碼", "賣方名稱", "賣方統編", "銷售額(未稅)", "稅額", "總計", "會計科目", "類型", "備註"]
                        columns = []
                        seen = set()
                        for c in desired_order:
                            if c in export_df.columns and c not in seen:
                                columns.append(c)
                                seen.add(c)
                        for c in export_df.columns:
                            if c not in seen and c not in ['日期_parsed', 'date', 'Date']:
                                columns.append(c)
                                seen.add(c)
                        export_df = export_df[columns].copy()
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine="openpyxl") as writer:
                            export_df.to_excel(writer, index=False, sheet_name="發票報表")
                            ws = writer.sheets["發票報表"]
                            header_font = Font(bold=True)
                            for col_cells in ws.iter_cols(min_row=1, max_row=1):
                                for cell in col_cells:
                                    cell.font = header_font
                                    col_letter = cell.column_letter
                                    ws.column_dimensions[col_letter].width = max(12, len(str(cell.value)) + 4)
                            amount_headers = {"銷售額(未稅)", "稅額", "總計"}
                            header_map = {cell.value: cell.column for cell in ws[1] if cell.value}
                            for header in amount_headers:
                                col_idx = header_map.get(header)
                                if col_idx is not None:
                                    for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                                        c = cell[0]
                                        c.number_format = '#,##0'
                                        c.alignment = Alignment(horizontal='right')
                        return output.getvalue()
                    excel_data = _gen_excel()
                    st.download_button(
                        "📊 導出Excel",
                        excel_data,
                        f"invoice_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="導出符合國稅局欄位結構的 Excel 報表"
                    )
            with act_col4:
                if not df.empty:
                    if PDF_AVAILABLE:
                        def _gen_pdf():
                            pdf = FPDF()
                            pdf.set_auto_page_break(auto=True, margin=15)
                            pdf.add_page()
                            font_path = "NotoSansTC-Regular.ttf"
                            font_loaded = False
                            font_name = "NotoSansTC"
                            if os.path.exists(font_path):
                                try:
                                    pdf.add_font(font_name, '', font_path, uni=True)
                                    pdf.add_font(font_name, 'B', font_path, uni=True)
                                    font_loaded = True
                                except Exception:
                                    font_loaded = False
                            def safe_cell(pdf, w, h, txt, border=0, ln=0, align='', fill=False, link='', font_name_override=None):
                                try:
                                    if font_name_override:
                                        pdf.set_font(font_name_override[0], font_name_override[1], font_name_override[2])
                                    pdf.cell(w, h, txt, border, ln, align, fill, link)
                                except Exception:
                                    pdf.set_font('Arial', '', 10)
                                    pdf.cell(w, h, str(txt)[:50], border, ln, align, fill, link)
                            if font_loaded:
                                pdf.set_font(font_name, 'B', 16)
                                safe_cell(pdf, 0, 10, '發票報帳統計報表', ln=1, align='C')
                            else:
                                pdf.set_font('Arial', 'B', 16)
                                safe_cell(pdf, 0, 10, 'Invoice Report', ln=1, align='C')
                            pdf.ln(5)
                            if font_loaded:
                                pdf.set_font(font_name, '', 10)
                                safe_cell(pdf, 0, 5, f'生成時間: {datetime.now().strftime("%Y年%m月%d日 %H:%M:%S")}', ln=1, align='R')
                            else:
                                pdf.set_font('Arial', '', 10)
                                safe_cell(pdf, 0, 5, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', ln=1, align='R')
                            pdf.ln(5)
                            if font_loaded:
                                pdf.set_font(font_name, '', 10)
                                company_name = st.session_state.get('company_name', '')
                                company_ubn = st.session_state.get('company_ubn', '')
                                if company_name:
                                    safe_cell(pdf, 200, 10, txt=f"報支公司：{company_name}", ln=1)
                                if company_ubn:
                                    safe_cell(pdf, 200, 10, txt=f"公司統編：{company_ubn}", ln=1)
                            else:
                                pdf.set_font('Arial', '', 10)
                                company_name = st.session_state.get('company_name', '')
                                company_ubn = st.session_state.get('company_ubn', '')
                                if company_name:
                                    safe_cell(pdf, 200, 10, txt=f"Company: {company_name}", ln=1)
                                if company_ubn:
                                    safe_cell(pdf, 200, 10, txt=f"UBN: {company_ubn}", ln=1)
                            pdf.ln(5)
                            if font_loaded:
                                pdf.set_font(font_name, 'B', 12)
                                safe_cell(pdf, 0, 8, '統計摘要', ln=1)
                                pdf.set_font(font_name, '', 10)
                            else:
                                pdf.set_font('Arial', 'B', 12)
                                safe_cell(pdf, 0, 8, '統計摘要', ln=1)
                                pdf.set_font('Arial', '', 10)
                            safe_cell(pdf, 90, 6, '累計金額:', 1)
                            export_df_for_stats = df_stats.copy() if 'df_stats' in locals() and not getattr(df_stats, 'empty', True) else df.copy()
                            if "總計" in export_df_for_stats.columns:
                                total_sum = pd.to_numeric(export_df_for_stats['總計'], errors='coerce').fillna(0).sum()
                            else:
                                total_sum = 0
                            safe_cell(pdf, 90, 6, f"${total_sum:,.0f}", 1, ln=1)
                            safe_cell(pdf, 90, 6, '累計稅額:', 1)
                            if "稅額" in export_df_for_stats.columns:
                                tax_sum = pd.to_numeric(export_df_for_stats['稅額'], errors='coerce').fillna(0).sum()
                            else:
                                tax_sum = 0
                            safe_cell(pdf, 90, 6, f"${tax_sum:,.0f}", 1, ln=1)
                            safe_cell(pdf, 90, 6, '發票總數:', 1)
                            safe_cell(pdf, 90, 6, f"{len(export_df_for_stats)} 筆", 1, ln=1)
                            pdf.ln(5)
                            export_df = df.copy()
                            col_widths = [25, 30, 30, 30, 25, 25, 25]
                            if font_loaded:
                                pdf.set_font(font_name, 'B', 10)
                                headers = ["日期", "發票號碼", "賣方統編", "銷售額(未稅)", "稅額", "總計", "備註"]
                            else:
                                pdf.set_font('Arial', 'B', 10)
                                headers = ["Date", "Invoice No", "Seller UBN", "Net Amount (Excl. Tax)", "Tax", "Total", "Note"]
                            for i, header in enumerate(headers):
                                safe_cell(pdf, col_widths[i], 7, header, 1, align='C')
                            pdf.ln()
                            if font_loaded:
                                pdf.set_font(font_name, '', 8)
                            else:
                                pdf.set_font('Arial', '', 8)
                            def pdf_safe_value(val, default='No'):
                                if pd.isna(val) or val == '' or val == 'N/A' or str(val).strip() == '':
                                    return default
                                return str(val)
                            for _, row in export_df.iterrows():
                                total_val = pd.to_numeric(row.get('總計', row.get('total', 0)), errors='coerce')
                                subtotal_val = pd.to_numeric(row.get('銷售額', row.get('subtotal', 0)), errors='coerce')
                                tax_val = pd.to_numeric(row.get('稅額', row.get('tax', 0)), errors='coerce')
                                if pd.isna(total_val):
                                    total_val = 0
                                if (pd.isna(subtotal_val) or subtotal_val == 0) or (pd.isna(tax_val) or tax_val == 0):
                                    if total_val > 0:
                                        tax_type_val = str(row.get('稅率類型', row.get('tax_type', '5%')) or '5%').strip().lower()
                                        if tax_type_val in ('0%', 'exempt', '零稅率', '免稅'):
                                            tax_val = 0
                                            subtotal_val = total_val
                                        else:
                                            tax_val = round(total_val - (total_val / 1.05))
                                            subtotal_val = total_val - tax_val
                                    else:
                                        subtotal_val = 0
                                        tax_val = 0
                                date_str = pdf_safe_value(row.get('日期', ''), 'No')[:10]
                                invoice_no = pdf_safe_value(row.get('發票號碼', ''), 'No')[:15]
                                seller_ubn = pdf_safe_value(row.get('賣方統編', ''), 'No')[:15]
                                note = pdf_safe_value(row.get('備註', '') or row.get('會計科目', '') or row.get('類型', ''), '')[:15]
                                net_amount_str = f"${subtotal_val:,.0f}"
                                tax_str = f"${tax_val:,.0f}"
                                total_str = f"${total_val:,.0f}"
                                safe_cell(pdf, col_widths[0], 6, date_str, 1)
                                safe_cell(pdf, col_widths[1], 6, invoice_no, 1)
                                safe_cell(pdf, col_widths[2], 6, seller_ubn, 1)
                                safe_cell(pdf, col_widths[3], 6, net_amount_str, 1, align='R')
                                safe_cell(pdf, col_widths[4], 6, tax_str, 1, align='R')
                                safe_cell(pdf, col_widths[5], 6, total_str, 1, align='R')
                                safe_cell(pdf, col_widths[6], 6, note, 1, ln=1)
                                if pdf.get_y() > 270:
                                    pdf.add_page()
                                    if font_loaded:
                                        pdf.set_font(font_name, 'B', 10)
                                    else:
                                        pdf.set_font('Arial', 'B', 10)
                                    for i, header in enumerate(headers):
                                        safe_cell(pdf, col_widths[i], 7, header, 1, align='C')
                                    pdf.ln()
                                    if font_loaded:
                                        pdf.set_font(font_name, '', 8)
                                    else:
                                        pdf.set_font('Arial', '', 8)
                            pdf_bytes = pdf.output(dest='S')
                            return bytes(pdf_bytes) if isinstance(pdf_bytes, bytearray) else pdf_bytes
                        pdf_data = _gen_pdf()
                        st.download_button(
                            "📄 導出PDF",
                            pdf_data,
                            f"invoice_report_{datetime.now().strftime('%Y%m%d')}.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            help="導出當前數據為PDF報告"
                        )
                    else:
                        st.info("📄 PDF", help="需要安裝 fpdf2")

            # 檢查並清理 DataFrame 的列名（確保沒有重複或無效列名），然後顯示數據表格
            try:
                if df_for_editor.empty:
                    st.info("📊 目前沒有數據可顯示")
                    ed_df = pd.DataFrame()
                else:
                    if df_for_editor.columns.duplicated().any():
                        cols = pd.Series(df_for_editor.columns)
                        for dup in cols[cols.duplicated()].unique():
                            cols[cols[cols == dup].index.values.tolist()] = [dup if i == 0 else f"{dup}_{i}" 
                                                                             for i in range(sum(cols == dup))]
                        df_for_editor.columns = cols
                    def clean_column_name(name):
                        if name is None: return "unnamed"
                        if not isinstance(name, str): name = str(name)
                        name = name.strip()
                        if name == "": return "unnamed"
                        return name.replace('\x00', '').replace('\n', ' ').replace('\r', ' ')
                    df_for_editor.columns = [clean_column_name(col) for col in df_for_editor.columns]
                    if df_for_editor.columns.duplicated().any():
                        cols, seen, new_cols = list(df_for_editor.columns), {}, []
                        for col in cols:
                            seen[col] = seen.get(col, 0) + 1
                            new_cols.append(col if seen[col] == 1 else f"{col}_{seen[col]}")
                        df_for_editor.columns = new_cols
                    # 表格僅顯示：選取、狀態放最前，其餘依序
                    table_columns_order = [
                        "選取", "狀態", "日期", "發票號碼", "賣方名稱", "賣方統編",
                        "會計科目", "類型", "銷售額", "稅額", "稅額 (5%)", "未稅金額", "總計",
                        "稅率類型", "備註", "建立時間"
                    ]
                    visible_columns = [c for c in table_columns_order if c in df_for_editor.columns]
                    def is_valid_column_name(name):
                        return name is not None and (isinstance(name, str) and name.strip() != "")
                    visible_columns = [c for c in visible_columns if is_valid_column_name(c)]
                    visible_columns = list(dict.fromkeys(visible_columns))
                    valid_column_config = {}
                    for k, v in column_config.items():
                        cleaned_key = clean_column_name(k)
                        if cleaned_key in df_for_editor.columns and is_valid_column_name(cleaned_key):
                            valid_column_config[cleaned_key] = v
                    try:
                        ed_df = st.data_editor(
                            df_for_editor,
                            use_container_width=True,
                            hide_index=True,
                            height=500,
                            column_config=valid_column_config if valid_column_config else None,
                            column_order=visible_columns if visible_columns else None,
                            key="invoice_data_editor_single"
                        )
                    except Exception as e:
                        st.error(f"表格顯示錯誤: {str(e)}")
                        st.dataframe(df_for_editor, use_container_width=True, height=500)
                        ed_df = df_for_editor.copy()
            except Exception as e:
                st.error(f"數據處理錯誤: {str(e)}")
                ed_df = pd.DataFrame()
                if not df_for_editor.empty:
                    st.dataframe(df_for_editor, use_container_width=True, height=500)

            # 添加 JavaScript 來高亮問題行並設置列對齊（在表格渲染後執行）
            st.markdown("""
            <script>
            (function() {
                function formatTable() {
                    const editor = document.querySelector('[data-testid="stDataEditor"]');
                    if (editor) {
                        const rows = editor.querySelectorAll('tbody tr');
                        const headerRow = editor.querySelector('thead tr');
                    
                        // 獲取表頭列名，用於確定列索引
                        const headers = [];
                        if (headerRow) {
                            headerRow.querySelectorAll('th').forEach(function(th) {
                                headers.push(th.textContent.trim());
                            });
                        }
                    
                        // 定義金額類欄位（需要右對齊）
                        const amountColumns = ['銷售額', '稅額', '未稅金額', '稅額 (5%)', '總計'];
                        // 定義變化百分比欄位（需要居中對齊）
                        const changeColumns = [];
                    
                        rows.forEach(function(row) {
                            const cells = row.querySelectorAll('td');
                            let isWarning = false;
                        
                            cells.forEach(function(cell, index) {
                                const text = cell.textContent || cell.innerText || '';
                            
                                // 檢查是否為問題行
                                if (text.includes('⚠️') || text.includes('❌ 缺失') || text.includes('❌ 缺漏')) {
                                    isWarning = true;
                                }
                            
                                // 設置列對齊與樣式 class（Stripe 風格：狀態綠標籤、金額等寬右對齊）
                                const columnName = headers[index] || '';
                            
                                if (columnName === '狀態' && (text.indexOf('正常') !== -1 || text.indexOf('✅') !== -1)) {
                                    cell.classList.add('status-ok');
                                }
                                if (amountColumns.includes(columnName)) {
                                    cell.classList.add('amount-cell');
                                    cell.style.textAlign = 'right';
                                }
                                else if (changeColumns.includes(columnName)) {
                                    cell.style.textAlign = 'center';
                                    cell.style.fontSize = '13px';
                                }
                                else {
                                    cell.style.textAlign = 'left';
                                }
                            });
                        
                            // 高亮問題行
                            if (isWarning) {
                                row.style.backgroundColor = 'rgba(234, 67, 53, 0.15)';
                                row.style.borderLeft = '4px solid #EA4335';
                                row.addEventListener('mouseenter', function() {
                                    this.style.backgroundColor = 'rgba(234, 67, 53, 0.25)';
                                });
                                row.addEventListener('mouseleave', function() {
                                    this.style.backgroundColor = 'rgba(234, 67, 53, 0.15)';
                                });
                            }
                        });
                    }
                }
            
                // 等待表格渲染完成後執行
                setTimeout(formatTable, 500);
                // 監聽表格更新
                const observer = new MutationObserver(formatTable);
                const targetNode = document.querySelector('[data-testid="stDataEditor"]');
                if (targetNode) {
                    observer.observe(targetNode, { childList: true, subtree: true });
                }
            })();
            </script>
            """, unsafe_allow_html=True)
        
            # 如果日期被轉換為日期類型，需要轉回字符串格式以便保存
            if not ed_df.empty and "日期" in ed_df.columns and ed_df["日期"].dtype != object:
                ed_df["日期"] = ed_df["日期"].dt.strftime("%Y/%m/%d").fillna(df["日期"] if not df.empty else "")
        
            # 處理選取列（如果存在）
            if not ed_df.empty and "選取" in ed_df.columns:
                if "選取" in df.columns:
                    df["選取"] = ed_df["選取"]
            elif "選取" not in df.columns:
                df["選取"] = False
        
            # 檢查是否有選中的行
            selected_count = ed_df["選取"].sum() if not ed_df.empty and "選取" in ed_df.columns else 0
            # 保存到session_state，用於下次顯示（不自動觸發rerun，避免頻繁刷新）
            current_selected = st.session_state.get("preview_selected_count", 0)
            if current_selected != selected_count:
                st.session_state.preview_selected_count = int(selected_count)
                # 只在用戶明確點擊刪除按鈕時才觸發rerun，不自動刷新
                # 移除自動 rerun，避免數據報表快速消失
        
            # 統一處理刪除邏輯（使用當前的選中數量）
            delete_button = delete_button_top
        
            if selected_count > 0 and delete_button:
                selected_rows = ed_df[ed_df["選取"]==True]
                # 收集要刪除的記錄信息（使用發票號碼+日期）
                records_to_delete = []
                user_email = st.session_state.get('user_email', 'default_user')
            
                for idx, row in selected_rows.iterrows():
                    # 優先從df_with_id獲取原始數據（未經過fill_empty處理，避免"No"值）
                    record_id = None
                    invoice_number = None
                    date = None
                
                    # 方法1: 優先從df_with_id獲取id（最可靠的方式，支持數據不完整的記錄）
                    if df_with_id is not None and idx in df_with_id.index:
                        # 優先獲取id字段（如果存在）
                        if 'id' in df_with_id.columns:
                            record_id = df_with_id.loc[idx, 'id']
                            if pd.isna(record_id):
                                record_id = None
                            else:
                                record_id = int(record_id) if record_id else None
                    
                        # 同時獲取發票號碼和日期（用於備選刪除方式）
                        if 'invoice_number' in df_with_id.columns:
                            invoice_number = df_with_id.loc[idx, 'invoice_number']
                        elif '發票號碼' in df_with_id.columns:
                            invoice_number = df_with_id.loc[idx, '發票號碼']
                    
                        if 'date' in df_with_id.columns:
                            date = df_with_id.loc[idx, 'date']
                        elif '日期' in df_with_id.columns:
                            date = df_with_id.loc[idx, '日期']
                
                    # 方法2: 如果df_with_id中沒有，從df獲取（df已經重命名為中文列名）
                    if record_id is None and df_with_id is not None and idx in df_with_id.index:
                        # 嘗試從df獲取id（如果df中有id列）
                        if 'id' in df.columns and idx in df.index:
                            record_id = df.loc[idx, 'id']
                            if pd.isna(record_id):
                                record_id = None
                            else:
                                record_id = int(record_id) if record_id else None
                
                    if (not invoice_number or pd.isna(invoice_number) or str(invoice_number).strip() in ['', 'No', 'N/A', 'nan', 'None']):
                        if idx in df.index and '發票號碼' in df.columns:
                            invoice_number = df.loc[idx, '發票號碼']
                
                    if (not date or pd.isna(date) or str(date).strip() in ['', 'No', 'N/A', 'nan', 'None']):
                        if idx in df.index and '日期' in df.columns:
                            date = df.loc[idx, '日期']
                
                    # 方法3: 如果還是沒有，從ed_df獲取（最後備選）
                    if (not invoice_number or pd.isna(invoice_number) or str(invoice_number).strip() in ['', 'No', 'N/A', 'nan', 'None']):
                        if '發票號碼' in row.index:
                            invoice_number = row.get('發票號碼')
                
                    if (not date or pd.isna(date) or str(date).strip() in ['', 'No', 'N/A', 'nan', 'None']):
                        if '日期' in row.index:
                            date = row.get('日期')
                
                    # 轉換為字符串並清理
                    if invoice_number is not None and not pd.isna(invoice_number):
                        invoice_number = str(invoice_number).strip()
                        invoice_number = invoice_number.replace('No', '').replace('N/A', '').replace('nan', '').replace('None', '').strip()
                    else:
                        invoice_number = ''
                
                    if date is not None and not pd.isna(date):
                        # 如果日期是日期類型，轉換為字符串
                        if isinstance(date, pd.Timestamp) or hasattr(date, 'strftime'):
                            try:
                                date = date.strftime("%Y/%m/%d")
                            except:
                                date = str(date).strip()
                        else:
                            date = str(date).strip()
                        date = date.replace('No', '').replace('N/A', '').replace('nan', '').replace('None', '').strip()
                    else:
                        date = ''
                
                    # 允許刪除數據不完整的記錄：優先使用id，如果沒有id則使用發票號碼+日期組合
                    # 如果都沒有，仍然嘗試添加（使用空值），讓刪除邏輯處理
                    delete_record = {}
                    if record_id is not None:
                        delete_record['id'] = record_id
                    if invoice_number:
                        delete_record['invoice_number'] = invoice_number
                    if date:
                        delete_record['date'] = date
                
                    # 只要有任何一個標識符（id、發票號碼+日期、或至少一個字段），就允許刪除
                    if delete_record:
                        records_to_delete.append(delete_record)
            
                if records_to_delete:
                    # 顯示刪除確認對話框
                    st.session_state.show_delete_confirm = True
                    st.session_state.delete_records = records_to_delete
                    st.session_state.delete_count = len(records_to_delete)
                    st.rerun()
                else:
                    st.warning("⚠️ 無法確定要刪除的記錄。請確保數據已正確加載。")
                    # 調試信息
                    with st.expander("🔍 調試信息", expanded=False):
                        st.write("**選中的行數:**", len(selected_rows))
                        st.write("**ed_df的列名:**", list(ed_df.columns))
                        st.write("**df的列名:**", list(df.columns) if 'df' in locals() else 'df未定義')
                        st.write("**df_with_id的列名:**", list(df_with_id.columns) if df_with_id is not None and not df_with_id.empty else 'df_with_id為None或空')
                        st.write("**選中的行數據（前3行）:**")
                        if not selected_rows.empty:
                            # 只顯示前3行，避免過多數據
                            display_cols = ['發票號碼', '日期'] if '發票號碼' in selected_rows.columns and '日期' in selected_rows.columns else list(selected_rows.columns)[:5]
                            st.dataframe(selected_rows[display_cols].head(3))
                        st.write("**提示:** 現在支持刪除數據不完整的記錄（即使發票號碼或日期為空）。如果仍然無法刪除，請檢查調試信息。")
        
            # 檢測是否有變更並自動保存（比較關鍵字段）
            # 使用 session_state 來追蹤是否已經檢查過變更，避免無限循環
            if "data_editor_checked" not in st.session_state:
                st.session_state.data_editor_checked = False
        
            # 只在第一次加載或明確需要檢查時才檢測變更
            if not st.session_state.data_editor_checked:
                has_changes = False
                try:
                    # 比較關鍵字段是否有變化（不包含ID和選取列）
                    # 只比較實際的數據列，跳過計算列
                    comparison_cols = [col for col in ed_df.columns 
                                      if col not in ['選取'] 
                                      and col in original_df_copy.columns]
                
                    for col in comparison_cols:
                        try:
                            # 使用更寬鬆的比較，忽略數據類型差異
                            ed_series = ed_df[col].astype(str).fillna('')
                            orig_series = original_df_copy[col].astype(str).fillna('')
                            if not ed_series.equals(orig_series):
                                has_changes = True
                                break
                        except:
                            # 如果比較失敗，跳過這一列
                            continue
                except:
                    # 如果比較失敗，不進行自動保存
                    has_changes = False
            
                # 標記為已檢查，避免重複檢查
                st.session_state.data_editor_checked = True
            
                # 只在確實有變更時才保存（且不是第一次加載）
                if has_changes and st.session_state.get("data_editor_initialized", False):
                    # 有變更，自動保存
                    # 多用戶版本：使用 user_email
                    user_email = st.session_state.get('user_email', 'default_user')
                    saved_count, errors, warnings = save_edited_data(ed_df, original_df_copy, user_email)
                    if saved_count > 0:
                        st.success(f"✅ 已自動保存 {saved_count} 筆數據變更")
                        # 統編驗證提示（僅提示，不阻擋）
                        if warnings:
                            st.warning("⚠️ 部分賣方統編非 8 位數字，已儲存僅供參考。")
                            if len(warnings) <= 3:
                                for w in warnings:
                                    st.caption(w)
                            else:
                                with st.expander("查看統編提示", expanded=False):
                                    for w in warnings:
                                        st.caption(w)
                        # 修復 Bug #4: 改進錯誤顯示，使用 expander 顯示所有錯誤
                        if errors:
                            if len(errors) > 3:
                                with st.expander(f"⚠️ 發現 {len(errors)} 個錯誤（點擊查看詳情）", expanded=False):
                                    for err in errors:
                                        st.error(err)
                            else:
                                for err in errors:
                                    st.error(err)
                        # 重置檢查標誌，允許下次檢查
                        st.session_state.data_editor_checked = False
                        time.sleep(0.5)
                        st.rerun()
                    elif errors:
                        # 如果全部失敗，顯示所有錯誤
                        if len(errors) > 3:
                            st.error(f"保存失敗: {errors[0]}")
                            with st.expander(f"查看所有 {len(errors)} 個錯誤", expanded=False):
                                for err in errors:
                                    st.error(err)
                        else:
                            for err in errors:
                                st.error(f"保存失敗: {err}")
                        # 重置檢查標誌
                        st.session_state.data_editor_checked = False
                else:
                    # 第一次加載，標記為已初始化
                    if not st.session_state.get("data_editor_initialized", False):
                        st.session_state.data_editor_initialized = True
